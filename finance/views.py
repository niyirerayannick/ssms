from collections import OrderedDict
from decimal import Decimal
from urllib.parse import urlencode

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Sum, Count
from django.conf import settings
from django.core.paginator import Paginator
from django.forms import formset_factory
from django.urls import reverse
from django.utils import timezone
from core.activity import set_audit_context
from core.models import District, AcademicYear, School
from core.export_utils import (
    ExportNumberedCanvas,
    add_export_header,
    autosize_worksheet_columns,
    build_export_pdf_document,
    build_export_table,
    resolve_logo_path,
    style_excel_header,
    style_excel_table_rows,
    write_excel_report_header,
)
from core.utils import normalize_identifier_value
from django.http import JsonResponse
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import SchoolFee, SchoolFeePayment, SchoolFeeDisbursement
from .forms import (
    FeeForm,
    FamilyInsuranceForm,
    BulkFeeFilterForm,
    BulkStudentFeeForm,
    SchoolFeePaymentForm,
    SchoolFeeDisbursementMarkPaidForm,
    SchoolFeeReconciliationForm,
)
from insurance.models import FamilyInsurance
from families.models import Family, FamilyStudent
from students.models import Student, StudentEnrollmentHistory
from .services import (
    SchoolFeeScope,
    filter_school_fee_queryset,
    get_bulk_enrollment_queryset,
    get_fee_queryset_with_context,
    get_or_create_fee_enrollment,
    get_or_create_school_fee_for_enrollment,
    record_school_fee_payment,
    reconcile_fee_scope,
    reconcile_disbursement_scope,
)


def _apply_default_fee_scope(params, *, default_term='1'):
    """Ensure fee pages have a stable default academic year/term scope for display only."""
    data = params.copy()
    if not data.get('academic_year'):
        active_year = AcademicYear.objects.filter(is_active=True).order_by('-name').first()
        if active_year:
            data['academic_year'] = str(active_year.id)
    if not data.get('term'):
        data['term'] = default_term
    return data


def _get_filtered_fees(request):
    """Return school fees queryset with the same filters used by the list page."""
    params = _apply_default_fee_scope(request.GET)
    return filter_school_fee_queryset(get_fee_queryset_with_context(), params=params)


def _get_filtered_disbursements(request):
    """Return the payout queue derived directly from unpaid school fee balances."""
    params = _apply_default_fee_scope(request.GET, default_term='')
    fees, filters = filter_school_fee_queryset(get_fee_queryset_with_context(), params=params)
    disbursements = fees.filter(balance__gt=0)
    return disbursements, filters


def _sync_fee_disbursement_queue(*, academic_year, term, school=None):
    """Reconcile fee snapshots and cached balances for a defined scope."""
    return reconcile_fee_scope(academic_year=academic_year, term=term, school=school)


def _resolve_logo_path():
    """Return the first existing local logo asset path."""
    return resolve_logo_path()


def _get_disbursement_totals(disbursements):
    """Aggregate summary totals for the disbursement queue."""
    totals = disbursements.aggregate(total=Sum('balance'))
    return {
        'listed_count': disbursements.count(),
        'pending_count': disbursements.count(),
        'exported_count': 0,
        'paid_count': 0,
        'total_amount_to_pay': totals['total'] or Decimal('0'),
    }


def _get_disbursement_labels(filters):
    academic_year_label = (
        AcademicYear.objects.filter(id=filters.get('academic_year_filter')).values_list('name', flat=True).first()
        if filters.get('academic_year_filter') else 'All Years'
    )
    district_label = (
        District.objects.filter(id=filters.get('district_filter')).values_list('name', flat=True).first()
        if filters.get('district_filter') else 'All Districts'
    )
    school_label = (
        School.objects.filter(id=filters.get('school_filter')).values_list('name', flat=True).first()
        if filters.get('school_filter') else 'All Schools'
    )
    return {
        'academic_year_label': academic_year_label or 'All Years',
        'district_label': district_label or 'All Districts',
        'school_label': school_label or 'All Schools',
    }


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def fees_list(request):
    """List all school fees with filters."""
    fees, filters = _get_filtered_fees(request)
    
    # Pagination
    paginator = Paginator(fees.order_by('-created_at'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'fees': page_obj,
        'page_obj': page_obj,
        **filters,
        'districts': District.objects.order_by('name'),
        'academic_years': AcademicYear.objects.order_by('-name'),
        'schools': School.objects.order_by('name'),
    }
    return render(request, 'finance/fees_list.html', context)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def export_fees_excel(request):
    """Export filtered school fee records with student and school banking details."""

    fees, _filters = _get_filtered_fees(request)
    fees = fees.order_by('student__last_name', 'student__first_name', '-academic_year__name', 'term')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'School Fees Export'

    headers = [
        'No.',
        'Student Name',
        'Academic Year',
        'Term',
        'Level',
        'School',
        'Bank Name',
        'Bank Account Name',
        'Bank Account Number',
        'Required Amount',
        'Amount Paid',
        'Balance To Pay',
        'Status',
    ]
    header_row = write_excel_report_header(worksheet, 'School Fees Export', 'Filtered school fee records', len(headers))
    worksheet.append(headers)
    style_excel_header(worksheet, header_row)

    data_start_row = header_row + 1
    for index, fee in enumerate(fees, start=1):
        worksheet.append([
            index,
            fee.student.full_name if fee.student else 'N/A',
            fee.academic_year.name if fee.academic_year else 'N/A',
            fee.get_term_display(),
            fee.class_level or 'N/A',
            fee.school_name or 'N/A',
            fee.bank_name or 'N/A',
            fee.bank_account_name or 'N/A',
            normalize_identifier_value(fee.bank_account_number, 'N/A'),
            float(fee.total_fees),
            float(fee.amount_paid),
            float(fee.balance),
            fee.get_payment_status_display(),
        ])

    try:
        autosize_worksheet_columns(worksheet, max_width=30)
    except AttributeError:
        pass
    style_excel_table_rows(
        worksheet,
        header_row_idx=header_row,
        data_start_row=data_start_row,
        data_end_row=worksheet.max_row,
        max_col=len(headers),
        centered_columns=[1, 3, 4, 13],
        right_aligned_columns=[10, 11, 12],
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="school_fees_students_export.xlsx"'
    workbook.save(response)
    return response


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def fee_disbursement_queue(request):
    """Queue of school fee balances finance still needs to pay out."""

    disbursements, filters = _get_filtered_disbursements(request)
    disbursements = disbursements.order_by('school_name', 'student__last_name', 'student__first_name')
    paginator = Paginator(disbursements, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    totals = _get_disbursement_totals(disbursements)
    labels = _get_disbursement_labels(filters)
    reconciliation_initial = _apply_default_fee_scope(request.GET, default_term='all')
    reconciliation_form = SchoolFeeReconciliationForm(initial={
        'academic_year': reconciliation_initial.get('academic_year'),
        'term': reconciliation_initial.get('term') or 'all',
        'school': reconciliation_initial.get('school'),
    })

    context = {
        'disbursements': page_obj,
        'page_obj': page_obj,
        'mark_paid_form': SchoolFeeDisbursementMarkPaidForm(),
        'reconciliation_form': reconciliation_form,
        'districts': District.objects.order_by('name'),
        'academic_years': AcademicYear.objects.order_by('-name'),
        'schools': School.objects.order_by('name'),
        **totals,
        **labels,
        **filters,
    }
    return render(request, 'finance/fee_disbursement_queue.html', context)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def sync_fee_disbursement_queue(request):
    """Reconcile fee snapshots and balances for a defined year/term scope."""

    if request.method == 'POST':
        form = SchoolFeeReconciliationForm(request.POST)
    else:
        initial_scope = _apply_default_fee_scope(request.GET, default_term='all')
        form = SchoolFeeReconciliationForm(initial={
            'academic_year': initial_scope.get('academic_year'),
            'term': initial_scope.get('term') or 'all',
            'school': initial_scope.get('school'),
        })

    if not form.is_valid():
        messages.error(request, 'Select a valid academic year and term scope before running reconciliation.')
        redirect_url = reverse('finance:fee_disbursement_queue')
        return redirect(redirect_url)

    academic_year = form.cleaned_data['academic_year']
    term = form.cleaned_data['term']
    school = form.cleaned_data.get('school')
    results = reconcile_disbursement_scope(academic_year=academic_year, term=term, school=school)
    term_label = 'All Terms' if term == 'all' else dict(SchoolFee.TERM_CHOICES).get(term, term)
    school_label = school.name if school else 'All Schools'
    messages.success(
        request,
        (
            f'You reconciled disbursements for {academic_year.name}, {term_label}, {school_label}. '
            f'Processed {results["processed_count"]} fee record(s), '
            f'created {results["created_count"]} disbursement row(s), '
            f'updated {results["updated_count"]} row(s), '
            f'cancelled {results["cancelled_count"]} row(s).'
        )
    )
    set_audit_context(
        request,
        action='Reconciled fee disbursements',
        description=(
            f'Reconciled {academic_year.name}, {term_label}, {school_label}. '
            f'Processed {results["processed_count"]}, created {results["created_count"]}, '
            f'updated {results["updated_count"]}, cancelled {results["cancelled_count"]}.'
        ),
    )
    query = urlencode({
        'academic_year': academic_year.id,
        'term': '' if term == 'all' else term,
        'school': school.id if school else '',
    })
    redirect_url = reverse('finance:fee_disbursement_queue')
    return redirect(f"{redirect_url}?{query}" if query else redirect_url)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def export_fee_disbursement_excel(request):
    """Export the derived payout queue for finance processing."""

    disbursements, filters = _get_filtered_disbursements(request)
    disbursements = disbursements.order_by('school_name', 'student__last_name', 'student__first_name')
    totals = _get_disbursement_totals(disbursements)
    labels = _get_disbursement_labels(filters)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Fee Disbursements'

    subtitle = (
        f"Academic Year: {labels['academic_year_label']} | "
        f"District: {labels['district_label']} | "
        f"School: {labels['school_label']}"
    )
    headers = [
        'No.',
        'Student Name',
        'Academic Year',
        'School Term',
        'District',
        'School',
        'Class Level',
        'Bank Name',
        'Bank Account Name',
        'Bank Account Number',
        'Amount To Pay',
        'Status',
    ]
    header_row = write_excel_report_header(worksheet, 'School Fee Disbursement Queue', subtitle, len(headers))
    worksheet.append(headers)
    style_excel_header(worksheet, header_row)

    data_start_row = header_row + 1
    for index, fee in enumerate(disbursements, start=1):
        student = fee.student
        district_name = student.partner.district.name if student and student.partner and student.partner.district else (
            student.family.district.name if student and student.family and student.family.district else (
                fee.school.district.name if fee.school and fee.school.district else 'N/A'
            )
        )
        worksheet.append([
            index,
            student.full_name if student else 'N/A',
            fee.academic_year.name if fee.academic_year else 'N/A',
            fee.get_term_display(),
            district_name,
            fee.school_name or 'N/A',
            fee.class_level or 'N/A',
            fee.bank_name or 'N/A',
            fee.bank_account_name or 'N/A',
            normalize_identifier_value(fee.bank_account_number, 'N/A'),
            float(fee.balance),
            fee.get_payment_status_display(),
        ])

    try:
        autosize_worksheet_columns(worksheet, max_width=32)
    except AttributeError:
        pass

    worksheet.append([])
    worksheet.append(['TOTAL RECORDS', totals['listed_count']])
    worksheet.append(['TOTAL AMOUNT TO PAY', float(totals['total_amount_to_pay'])])
    worksheet.append(['UNPAID RECORDS', totals['pending_count']])
    style_excel_table_rows(
        worksheet,
        header_row_idx=header_row,
        data_start_row=data_start_row,
        data_end_row=data_start_row + max(disbursements.count(), 0) - 1 if hasattr(disbursements, 'count') else worksheet.max_row,
        max_col=len(headers),
        centered_columns=[1, 3, 4, 12],
        right_aligned_columns=[11],
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="school_fee_disbursement_queue.xlsx"'
    workbook.save(response)
    return response


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def export_fee_disbursement_pdf(request):
    """Export the derived payout queue as a signable PDF."""

    disbursements, filters = _get_filtered_disbursements(request)
    disbursements = list(disbursements.order_by('school_name', 'student__last_name', 'student__first_name'))
    totals = _get_disbursement_totals(SchoolFee.objects.filter(id__in=[item.id for item in disbursements]))
    labels = _get_disbursement_labels(filters)

    buffer = HttpResponse(content_type='application/pdf')
    buffer['Content-Disposition'] = 'attachment; filename="school_fee_disbursement_queue.pdf"'

    doc = build_export_pdf_document(
        buffer,
        'School Fee Disbursement Queue',
        pagesize=landscape(A4),
        left_margin=20,
        right_margin=20,
        top_margin=20,
        bottom_margin=20,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        'DisbursementCell',
        parent=styles['BodyText'],
        fontSize=6.4,
        leading=7.4,
        textColor=colors.HexColor('#0f172a'),
    )
    amount_style = ParagraphStyle(
        'DisbursementAmount',
        parent=cell_style,
        alignment=2,
        fontName='Helvetica-Bold',
    )
    elements = []
    add_export_header(
        elements,
        'School Fee Disbursement Queue',
        (
            f"Academic Year: {labels['academic_year_label']} | "
            f"District: {labels['district_label']} | "
            f"School: {labels['school_label']}"
        ),
        generated_label=f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}",
    )

    summary_data = [
        ['Total Records', str(totals['listed_count']), 'Unpaid', str(totals['pending_count']), '', '', '', ''],
        ['Total Amount', f"FRW {totals['total_amount_to_pay']:,.2f}", '', '', '', '', '', ''],
    ]
    summary_table = build_export_table(
        summary_data,
        col_widths=[78, 70, 58, 55, 60, 55, 42, 42],
        body_font_size=8,
        centered_columns=[1, 3, 5, 7],
    )
    summary_table.setStyle(TableStyle([
        ('SPAN', (1, 1), (-1, 1)),
        ('ALIGN', (1, 1), (-1, 1), 'LEFT'),
        ('TEXTCOLOR', (0, 1), (1, 1), colors.HexColor('#dc2626')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 12))

    table_data = [[
        'No.',
        'Student',
        'District',
        'Academic Year',
        'School Term',
        'School',
        'Bank',
        'Account Name',
        'Account Number',
        'Amount To Pay',
    ]]
    for index, fee in enumerate(disbursements, start=1):
        student = fee.student
        district_name = student.partner.district.name if student and student.partner and student.partner.district else (
            student.family.district.name if student and student.family and student.family.district else (
                fee.school.district.name if fee.school and fee.school.district else 'N/A'
            )
        )
        table_data.append([
            Paragraph(str(index), cell_style),
            Paragraph(student.full_name if student else 'N/A', cell_style),
            Paragraph(district_name, cell_style),
            Paragraph(fee.academic_year.name if fee.academic_year else 'N/A', cell_style),
            Paragraph(fee.get_term_display(), cell_style),
            Paragraph(fee.school_name or 'N/A', cell_style),
            Paragraph(fee.bank_name or 'N/A', cell_style),
            Paragraph(fee.bank_account_name or 'N/A', cell_style),
            Paragraph(normalize_identifier_value(fee.bank_account_number, 'N/A'), cell_style),
            Paragraph(f"{fee.balance:,.2f}", amount_style),
        ])

    table_data.append([
        '',
        Paragraph('TOTAL', amount_style),
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        Paragraph(f"{totals['total_amount_to_pay']:,.2f}", amount_style),
    ])
    table = build_export_table(
        table_data,
        col_widths=[24, 96, 58, 64, 52, 102, 60, 88, 80, 64],
        body_font_size=6.5,
        centered_columns=[0],
        right_aligned_columns=[9],
        total_row_indexes=[len(table_data) - 1],
    )
    table.setStyle(TableStyle([
        ('TEXTCOLOR', (9, -1), (9, -1), colors.HexColor('#b45309')),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 24))

    signature_table = Table([
        [
            Paragraph('Prepared By<br/><br/>_________________________<br/>Program Coordinator', styles['BodyText']),
            Paragraph('Verified By<br/><br/>_________________________<br/>Corporate and HR Verification', styles['BodyText']),
            Paragraph('Approved By<br/><br/>_________________________<br/>Executive Secretary', styles['BodyText']),
        ]
    ], colWidths=[240, 240, 240])
    signature_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(signature_table)

    doc.build(elements, canvasmaker=ExportNumberedCanvas)
    return buffer


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
@transaction.atomic
def mark_fee_disbursements_paid(request):
    """Mark selected unpaid fee balances as paid and post SchoolFeePayment records."""

    if request.method != 'POST':
        return redirect('finance:fee_disbursement_queue')

    selected_ids = request.POST.getlist('selected_fees')
    form = SchoolFeeDisbursementMarkPaidForm(request.POST)
    if not selected_ids:
        messages.error(request, 'Select at least one fee record to mark as paid.')
        return redirect('finance:fee_disbursement_queue')
    if not form.is_valid():
        messages.error(request, 'Provide a valid payment date before marking payouts as paid.')
        return redirect('finance:fee_disbursement_queue')

    payment_date = form.cleaned_data['payment_date']
    payment_reference = form.cleaned_data['payment_reference']
    notes = form.cleaned_data['notes']

    fees = list(
        get_fee_queryset_with_context()
        .filter(id__in=selected_ids, balance__gt=0)
    )

    updated_count = 0
    for fee in fees:
        amount_paid = fee.balance
        try:
            _, created = record_school_fee_payment(
                fee=fee,
                amount_paid=amount_paid,
                payment_date=payment_date,
                payment_method='bank',
                reference_number=payment_reference,
                recorded_by=request.user,
                notes=notes or f'Processed from finance payout queue on {payment_date.isoformat()}.',
                idempotency_key=(
                    f'queue-payment:{fee.pk}:{payment_date.isoformat()}:'
                    f'{normalize_identifier_value(payment_reference) or "no-ref"}:{amount_paid}'
                ),
            )
        except ValidationError as exc:
            messages.error(request, '; '.join(exc.messages))
            continue
        if created:
            updated_count += 1

    messages.success(request, f'{updated_count} fee record(s) marked as paid and posted to school fee payments.')
    set_audit_context(
        request,
        action='Marked school fee disbursements paid',
        description=f'Marked {updated_count} payout queue record(s) as paid.',
    )
    return redirect('finance:fee_disbursement_queue')


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def school_fees_dashboard(request):
    """Dashboard focused on school fees only."""
    fees_queryset, dashboard_filters = _get_filtered_fees(request)
    total_school_fees = fees_queryset.count()
    students_paid_count = fees_queryset.filter(payment_status='paid').values('student_id').distinct().count()
    academic_year_count = fees_queryset.values('academic_year_id').distinct().count()
    term_record_count = total_school_fees

    total_fees_required = fees_queryset.aggregate(Sum('total_fees'))['total_fees__sum'] or 0
    total_fees_collected = fees_queryset.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    total_fees_outstanding = fees_queryset.aggregate(Sum('balance'))['balance__sum'] or 0
    fees_collection_percentage = round((total_fees_collected / total_fees_required * 100) if total_fees_required > 0 else 0, 1)

    from django.utils import timezone
    from datetime import timedelta
    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_school_fees = fees_queryset.filter(updated_at__gte=seven_days_ago).count()

    school_groups_map = OrderedDict()
    for fee in fees_queryset:
        student = fee.student
        school = fee.school
        school_name = school.name if school else (fee.school_name or 'Unassigned School')
        key = ('school', school.id) if school else ('external', school_name.lower())
        if key not in school_groups_map:
            school_groups_map[key] = {
                'school_id': school.id if school else None,
                'school_name': school_name,
                'student_ids': set(),
                'total_required': Decimal('0'),
                'total_paid': Decimal('0'),
                'total_balance': Decimal('0'),
                'status_counts': {
                    'paid': 0,
                    'partial': 0,
                    'pending': 0,
                    'overdue': 0,
                },
            }

        entry = school_groups_map[key]
        entry['student_ids'].add(student.id if student else None)
        entry['total_required'] += fee.total_fees
        entry['total_paid'] += fee.amount_paid
        entry['total_balance'] += fee.balance
        status = fee.payment_status or 'pending'
        if status in entry['status_counts']:
            entry['status_counts'][status] += 1

    school_fee_groups = []
    for entry in school_groups_map.values():
        school_fee_groups.append({
            'school_id': entry['school_id'],
            'school_name': entry['school_name'],
            'student_count': len({sid for sid in entry['student_ids'] if sid}),
            'total_required': entry['total_required'],
            'total_paid': entry['total_paid'],
            'total_balance': entry['total_balance'],
            'status_counts': entry['status_counts'],
            'has_linked_school': entry['school_id'] is not None,
        })

    paginator = Paginator(school_fee_groups, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'students_paid_count': students_paid_count,
        'academic_year_count': academic_year_count,
        'term_record_count': term_record_count,
        'total_fees_required': total_fees_required,
        'total_fees_collected': total_fees_collected,
        'total_fees_outstanding': total_fees_outstanding,
        'fees_collection_percentage': fees_collection_percentage,
        'recent_school_fees': recent_school_fees,
        'school_fee_groups': page_obj,
        'page_obj': page_obj,
        'districts': District.objects.order_by('name'),
        'academic_years': AcademicYear.objects.order_by('-name'),
        'schools': School.objects.order_by('name'),
        **dashboard_filters,
    }
    return render(request, 'finance/school_fees_dashboard.html', context)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def school_fee_students(request, school_id):
    """Display aggregated fee information for students within a school."""
    school = get_object_or_404(School, pk=school_id)

    school_fees = (
        get_fee_queryset_with_context()
        .filter(school=school)
        .order_by('student__first_name', 'student__last_name')
    )

    student_map = OrderedDict()
    for fee in school_fees:
        student = fee.student
        if not student:
            continue
        entry = student_map.setdefault(
            student.id,
            {
                'student': student,
                'total_required': Decimal('0'),
                'total_paid': Decimal('0'),
                'total_balance': Decimal('0'),
                'latest_status': fee.payment_status,
                'status_display': fee.get_payment_status_display(),
            },
        )
        entry['total_required'] += fee.total_fees
        entry['total_paid'] += fee.amount_paid
        entry['total_balance'] += fee.balance
        entry['latest_status'] = fee.payment_status
        entry['status_display'] = fee.get_payment_status_display()

    student_summaries = list(student_map.values())
    school_totals = {
        'required': sum((item['total_required'] for item in student_summaries), Decimal('0')),
        'paid': sum((item['total_paid'] for item in student_summaries), Decimal('0')),
        'balance': sum((item['total_balance'] for item in student_summaries), Decimal('0')),
        'student_count': len(student_summaries),
    }

    context = {
        'school': school,
        'student_summaries': student_summaries,
        'school_totals': school_totals,
    }
    return render(request, 'finance/school_fee_students.html', context)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def bulk_fee_entry(request):
    """Allow finance admins to record or update many school fee entries at once."""

    data_source = request.POST if request.method == 'POST' else request.GET
    filter_form = BulkFeeFilterForm(data_source or None)
    formset = None
    table_rows = []
    students_loaded = False
    selected_filters = {}
    selected_school = None
    selected_district = None
    selected_partner = None
    selected_scope_school = None
    selected_year = None
    term_display = None
    category_label = None
    summary = {
        'student_count': 0,
        'total_required': Decimal('0'),
        'total_paid': Decimal('0'),
    }

    if filter_form.is_valid():
        students_loaded = True
        academic_year = filter_form.cleaned_data['academic_year']
        district = filter_form.cleaned_data['district']
        partner = filter_form.cleaned_data.get('partner')
        school = filter_form.cleaned_data.get('school')
        term = filter_form.cleaned_data['term']
        category = filter_form.cleaned_data['category']
        payment_date = filter_form.cleaned_data.get('payment_date') or timezone.now().date()

        selected_district = district
        selected_partner = partner
        selected_scope_school = school
        selected_year = academic_year
        term_display = dict(SchoolFee.TERM_CHOICES).get(term, term)
        category_label = dict(filter_form.fields['category'].choices).get(category, category)

        scope = SchoolFeeScope(
            academic_year=academic_year,
            term=term,
            district=district,
            partner=partner,
            school=school,
            category=category,
        )
        enrollments = list(get_bulk_enrollment_queryset(scope))
        student_ids = [enrollment.student_id for enrollment in enrollments]
        existing_fee_map = {
            fee.student_id: fee
            for fee in get_fee_queryset_with_context().filter(
                student_id__in=student_ids,
                academic_year=academic_year,
                term=term,
            )
        }

        initial_data = []
        total_required_sum = Decimal('0')
        total_paid_sum = Decimal('0')
        for enrollment in enrollments:
            existing_fee = existing_fee_map.get(enrollment.student_id)
            default_total = (
                existing_fee.total_fees
                if existing_fee
                else (enrollment.school.fee_amount if enrollment.school else Decimal('0'))
            )
            default_paid = existing_fee.amount_paid if existing_fee else Decimal('0')
            total_required_sum += default_total or Decimal('0')
            total_paid_sum += default_paid or Decimal('0')
            initial_data.append({
                'student_id': enrollment.student_id,
                'total_fees': default_total,
                'amount_paid': default_paid,
            })

        BulkFormSet = formset_factory(BulkStudentFeeForm, extra=0)
        formset_valid = False
        if request.method == 'POST':
            formset = BulkFormSet(request.POST)
            formset_valid = formset.is_valid()
            if formset_valid:
                created_count = 0
                updated_count = 0
                payment_delta_count = 0
                enrollment_lookup = {enrollment.student_id: enrollment for enrollment in enrollments}
                bulk_has_errors = False
                with transaction.atomic():
                    for form in formset:
                        student_id = form.cleaned_data.get('student_id')
                        total_fees = form.cleaned_data.get('total_fees')
                        desired_amount_paid = form.cleaned_data.get('amount_paid')

                        if not student_id or student_id not in enrollment_lookup:
                            continue
                        if total_fees is None:
                            continue
                        if desired_amount_paid is None:
                            desired_amount_paid = Decimal('0')

                        enrollment = enrollment_lookup[student_id]
                        fee, created = get_or_create_school_fee_for_enrollment(
                            enrollment,
                            term,
                            total_fees=total_fees,
                            actor=request.user,
                        )
                        fee.payment_date = payment_date
                        fee.recorded_by = request.user
                        fee.save()

                        current_paid = fee.amount_paid or Decimal('0')
                        if desired_amount_paid < current_paid:
                            form.add_error(
                                'amount_paid',
                                'Bulk entry cannot reduce already recorded payments. Use payment history for corrections.'
                            )
                            bulk_has_errors = True
                            break

                        payment_delta = desired_amount_paid - current_paid
                        if payment_delta > 0:
                            try:
                                record_school_fee_payment(
                                    fee=fee,
                                    amount_paid=payment_delta,
                                    payment_date=payment_date,
                                    payment_method='bank',
                                    reference_number='',
                                    recorded_by=request.user,
                                    notes=f'Bulk entry adjustment for {academic_year.name} {fee.get_term_display()}.',
                                    idempotency_key=(
                                        f'bulk-entry:{fee.pk}:{payment_date.isoformat()}:'
                                        f'{desired_amount_paid}'
                                    ),
                                )
                            except ValidationError as exc:
                                form.add_error('amount_paid', '; '.join(exc.messages))
                                bulk_has_errors = True
                                break
                            payment_delta_count += 1
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                    if bulk_has_errors:
                        transaction.set_rollback(True)

                if not bulk_has_errors:
                    scope_label = partner.name if partner else (school.name if school else district.name)
                    set_audit_context(
                        request,
                        action='Saved bulk school fees',
                        description=(
                            f'Saved bulk fees for {scope_label} in {academic_year.name} '
                            f'{dict(SchoolFee.TERM_CHOICES).get(term, term)}: '
                            f'{created_count} created, {updated_count} updated, '
                            f'{payment_delta_count} payment adjustments.'
                        ),
                    )
                    messages.success(
                        request,
                        (
                            f"Bulk fees saved for {scope_label}: "
                            f"{created_count} new, {updated_count} updated, {payment_delta_count} payment adjustment(s) posted."
                        ),
                    )
                    params_dict = {
                        'academic_year': academic_year.id,
                        'term': term,
                        'category': category,
                        'payment_date': payment_date.isoformat(),
                    }
                    if partner:
                        params_dict['partner'] = partner.id
                    if school:
                        params_dict['school'] = school.id
                    elif district:
                        params_dict['district'] = district.id
                    params = urlencode(params_dict)
                    return redirect(f"{reverse('finance:bulk_fee_entry')}?{params}")
        else:
            formset = BulkFormSet(initial=initial_data)

        if formset is not None:
            for idx, enrollment in enumerate(enrollments):
                if idx >= len(formset.forms):
                    break
                table_rows.append({
                    'student': enrollment.student,
                    'enrollment': enrollment,
                    'form': formset.forms[idx],
                    'existing_fee': existing_fee_map.get(enrollment.student_id),
                })

        summary = {
            'student_count': len(enrollments),
            'total_required': total_required_sum,
            'total_paid': total_paid_sum,
        }
        selected_filters = {
            'academic_year': academic_year.id,
            'term': term,
            'category': category,
            'payment_date': payment_date.isoformat(),
        }
        if partner:
            selected_filters['partner'] = partner.id
        if school:
            selected_filters['school'] = school.id
        elif district:
            selected_filters['district'] = district.id

    context = {
        'filter_form': filter_form,
        'formset': formset,
        'table_rows': table_rows,
        'students_loaded': students_loaded,
        'selected_filters': selected_filters,
        'selected_school': selected_school,
        'selected_district': selected_district,
        'selected_partner': selected_partner,
        'selected_scope_school': selected_scope_school,
        'selected_year': selected_year,
        'term_display': term_display,
        'category_label': category_label,
        'summary': summary,
    }
    return render(request, 'finance/bulk_fee_entry.html', context)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def fee_create(request):
    """Create a new fee record - for both school fees and Mutuelle de Santé."""
    payment_type = request.GET.get('type', 'school_fee')  # 'school_fee' or 'insurance'
    
    if request.method == 'POST':
        if payment_type == 'insurance':
            # Handle Mutuelle de Santé (Family Insurance) payment
            form = FamilyInsuranceForm(request.POST)
            # Allow any family to be selected (since they are loaded dynamically via JS)
            # This ensures validation passes even if the initial queryset was empty
            form.fields['family'].queryset = Family.objects.all()
            
            if form.is_valid():
                insurance = form.save()
                messages.success(request, f'Mutuelle de Santé payment recorded for family {insurance.family.family_code}!')
                return redirect('insurance:mutuelle_dashboard')
        else:
            # Handle School Fee payment
            form = FeeForm(request.POST)
            # Allow any student to be selected (since they are loaded dynamically via JS)
            form.fields['student'].queryset = Student.objects.all()
            
            if form.is_valid():
                fee = form.save(commit=False)
                if not fee.payment_date:
                    fee.payment_date = timezone.now().date()
                fee.recorded_by = request.user
                fee.save()
                set_audit_context(
                    request,
                    action='Created school fee plan',
                    description=f'Created school fee plan for {fee.student.full_name} in {fee.academic_year.name} {fee.get_term_display()}.',
                )
                messages.success(request, f'School fee plan recorded for {fee.student.full_name}.')
                return redirect('finance:school_fees_dashboard')
    else:
        if payment_type == 'insurance':
            form = FamilyInsuranceForm()
            # Initialize with empty queryset - families will be loaded via AJAX based on district
            form.fields['family'].queryset = Family.objects.none()
        else:
            form = FeeForm()
            # Initialize with empty queryset - students will be loaded via AJAX based on district
            form.fields['student'].queryset = Student.objects.none()
    
    context = {
        'form': form,
        'payment_type': payment_type,
        'districts': District.objects.order_by('name'),
        'title': 'Add Mutuelle Payment' if payment_type == 'insurance' else 'Add School Fee Plan'
    }
    return render(request, 'finance/fee_form.html', context)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def fee_edit(request, pk):
    """Edit an existing fee record."""
    fee = get_object_or_404(SchoolFee, pk=pk)
    
    if request.method == 'POST':
        form = FeeForm(request.POST, instance=fee)
        if form.is_valid():
            fee = form.save(commit=False)
            if not fee.payment_date:
                fee.payment_date = timezone.now().date()
            if not fee.recorded_by:
                fee.recorded_by = request.user
            fee.save()
            set_audit_context(
                request,
                action='Updated school fee plan',
                description=f'Updated school fee plan for {fee.student.full_name} in {fee.academic_year.name} {fee.get_term_display()}.',
            )
            messages.success(request, f'Fee plan updated for {fee.student.full_name}.')
            return redirect('finance:fees_list')
    else:
        form = FeeForm(instance=fee)
    
    return render(request, 'finance/fee_form.html', {'form': form, 'fee': fee, 'title': 'Edit Fee Plan'})


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def overdue_fees(request):
    """List all overdue fees."""
    overdue = SchoolFee.objects.filter(payment_status='overdue').select_related('student')
    
    context = {
        'overdue_fees': overdue,
    }
    return render(request, 'finance/overdue_fees.html', context)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def finance_dashboard(request):
    """Legacy finance dashboard endpoint redirected to School Fees dashboard."""
    return redirect('finance:school_fees_dashboard')


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def student_payment_history(request, student_id):
    """View all school fee payments for a specific student."""
    student = get_object_or_404(Student, pk=student_id)
    
    # Get all fees for this student
    fees_qs = (
        get_fee_queryset_with_context().filter(student=student)
        .prefetch_related('payments')
        .order_by('-academic_year__name', '-term')
    )
    
    # Calculate totals for this student
    total_required = fees_qs.aggregate(Sum('total_fees'))['total_fees__sum'] or 0
    total_paid = fees_qs.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    total_outstanding = fees_qs.aggregate(Sum('balance'))['balance__sum'] or 0
    
    # Payment percentage
    payment_percentage = round((total_paid / total_required * 100) if total_required > 0 else 0, 1)

    fees = list(fees_qs)
    payment_records = list(
        SchoolFeePayment.objects.filter(school_fee__student=student)
        .select_related('school_fee', 'school_fee__academic_year')
        .order_by('-payment_date', '-created_at')
    )
    
    context = {
        'student': student,
        'fees': fees,
        'total_required': total_required,
        'total_paid': total_paid,
        'total_outstanding': total_outstanding,
        'payment_percentage': payment_percentage,
        'payment_records': payment_records,
    }
    return render(request, 'finance/student_payment_history.html', context)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def add_student_payment(request, student_id):
    """Add a school fee payment transaction for a specific student."""
    student = get_object_or_404(Student, pk=student_id)
    selected_history = None
    
    if request.method == 'POST':
        form = SchoolFeePaymentForm(request.POST, student=student)
        if form.is_valid():
            academic_year = form.cleaned_data['academic_year']
            term = form.cleaned_data['term']
            total_fees = form.cleaned_data['total_fees']
            selected_history = get_or_create_fee_enrollment(student, academic_year)
            if not selected_history:
                form.add_error('academic_year', 'This student has no enrollment history for the selected academic year.')
            else:
                fee, _created = get_or_create_school_fee_for_enrollment(
                    selected_history,
                    term,
                    total_fees=total_fees,
                    actor=request.user,
                )
                try:
                    record_school_fee_payment(
                        fee=fee,
                        amount_paid=form.cleaned_data['amount_paid'],
                        payment_date=form.cleaned_data['payment_date'],
                        payment_method=form.cleaned_data['payment_method'],
                        reference_number=form.cleaned_data.get('reference_number') or '',
                        recorded_by=request.user,
                        notes=form.cleaned_data.get('notes') or '',
                        idempotency_key=(
                            f'manual-payment:{fee.pk}:{form.cleaned_data["payment_date"].isoformat()}:'
                            f'{normalize_identifier_value(form.cleaned_data.get("reference_number")) or "no-ref"}:'
                            f'{form.cleaned_data["amount_paid"]}'
                        ),
                    )
                except ValidationError as exc:
                    form.add_error('amount_paid', '; '.join(exc.messages))
                else:
                    set_audit_context(
                        request,
                        action='Recorded school fee payment',
                        description=(
                            f'Recorded school fee payment for {student.full_name} in {academic_year.name} '
                            f'{dict(SchoolFee.TERM_CHOICES).get(term, term)}.'
                        ),
                    )
                    messages.success(request, f'School fee payment recorded for {student.full_name}!')
                    return redirect('finance:student_payments', student_id=student.id)
    else:
        form = SchoolFeePaymentForm(student=student)
    
    if not selected_history and form.is_bound and form.is_valid():
        selected_history = get_or_create_fee_enrollment(student, form.cleaned_data['academic_year'])

    if not selected_history:
        active_year = AcademicYear.objects.filter(is_active=True).first()
        if active_year:
            selected_history = StudentEnrollmentHistory.objects.filter(
                student=student,
                academic_year=active_year,
            ).select_related('school').first()

    display_school = (
        selected_history.display_school_name
        if selected_history else (student.school.name if student.school else student.school_name)
    )
    display_class = (
        selected_history.class_level
        if selected_history else student.class_level
    )
    display_school_obj = selected_history.school if selected_history else student.school

    context = {
        'form': form,
        'student': student,
        'title': f'Add School Fee Payment - {student.full_name}',
        'student_school_name': display_school,
        'student_class_level': display_class,
        'student_bank_name': display_school_obj.bank_name if display_school_obj and display_school_obj.bank_name else '',
        'student_bank_account_name': display_school_obj.bank_account_name if display_school_obj and display_school_obj.bank_account_name else '',
        'student_bank_account_number': normalize_identifier_value(
            display_school_obj.bank_account_number if display_school_obj and display_school_obj.bank_account_number else ''
        ),
    }
    return render(request, 'finance/student_fee_form.html', context)


from core.utils import encode_id, decode_id


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def get_student_details(request, student_id):
    """API endpoint to fetch student details including school and class information."""
    try:
        # Support both raw integer ID and HashID
        student_id = decode_id(student_id)
        student = get_object_or_404(Student, pk=student_id)
        academic_year_id = request.GET.get('academic_year')
        academic_year = AcademicYear.objects.filter(pk=academic_year_id).first() if academic_year_id else AcademicYear.objects.filter(is_active=True).first()
        history = (
            student.enrollment_history.filter(academic_year=academic_year).select_related('school').first()
            if academic_year else None
        )
        latest_fee = (
            SchoolFee.objects.filter(student=student, academic_year=academic_year).order_by('-created_at').first()
            if academic_year else SchoolFee.objects.filter(student=student).order_by('-created_at').first()
        )
        payment_dates = latest_fee.payment_dates if latest_fee else ''
        school = history.school if history and history.school else student.school
        school_name = history.display_school_name if history else (student.school_name or (student.school.name if student.school else ''))
        class_level = history.class_level if history else student.class_level
        
        data = {
            'success': True,
            'student_id': encode_id(student.id),
            'student_name': student.full_name,
            'school_name': school_name,
            'school_id': encode_id(school.id) if school else None,
            'class_level': class_level,
            'enrollment_status': student.get_enrollment_status_display(),
            'total_fees': str(school.fee_amount) if school else '0',
            'payment_dates': payment_dates,
            'bank_name': school.bank_name if school and school.bank_name else '',
            'bank_account_name': school.bank_account_name if school and school.bank_account_name else '',
            'bank_account_number': normalize_identifier_value(
                school.bank_account_number if school and school.bank_account_number else ''
            ),
        }
        return JsonResponse(data)
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def get_family_insurance_details(request, family_id):
    """API endpoint to fetch family insurance details including amounts."""
    try:
        # Support both raw integer ID and HashID
        family_id = decode_id(family_id)
        family = get_object_or_404(Family, pk=family_id)
        
        # Get the latest insurance record for this family
        latest_insurance = FamilyInsurance.objects.filter(family=family).order_by('-insurance_year__name').first()
        
        # Calculate total contribution across all years
        total_contribution = FamilyInsurance.objects.filter(family=family).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
        
        # Calculate total required across all years
        total_required_all_years = FamilyInsurance.objects.filter(family=family).aggregate(Sum('required_amount'))['required_amount__sum'] or 0
        
        data = {
            'success': True,
            'family_id': encode_id(family.id),
            'family_code': family.family_code,
            'family_name': (family.head_of_family or '').strip(),
            'total_members': family.total_family_members or 0,
            'payment_ability': family.payment_ability,
            'payment_ability_display': family.get_payment_ability_display(),
            'mutuelle_support_status': family.mutuelle_support_status,
            'mutuelle_support_status_display': family.get_mutuelle_support_status_display(),
            'total_contribution': str(total_contribution),
            'total_required_all_years': str(total_required_all_years),
        }
        
        if latest_insurance:
            data['insurance_year_id'] = encode_id(latest_insurance.insurance_year_id) if latest_insurance.insurance_year_id else None
            data['insurance_year'] = latest_insurance.insurance_year.name if latest_insurance.insurance_year else ''
            data['required_amount'] = str(latest_insurance.required_amount)
            data['amount_paid'] = str(latest_insurance.amount_paid)
            data['balance'] = str(latest_insurance.balance)
            data['coverage_status'] = latest_insurance.get_coverage_status_display()
            data['payment_dates'] = latest_insurance.payment_dates
            data['has_existing_record'] = True
        else:
            active_year = AcademicYear.objects.filter(is_active=True).order_by('-name').first()
            data['insurance_year_id'] = encode_id(active_year.id) if active_year else None
            data['insurance_year'] = active_year.name if active_year else ''
            data['required_amount'] = str(family.total_contribution)
            data['amount_paid'] = '0'
            data['balance'] = str(family.total_contribution)
            data['coverage_status'] = 'Not Covered'
            data['payment_dates'] = ''
            data['has_existing_record'] = False
        
        return JsonResponse(data)
    except Family.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Family not found'}, status=404)


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def get_families_by_district(request, district_id):
    """API endpoint to fetch families in a district."""
    families = Family.objects.filter(district_id=district_id).values('id', 'family_code', 'head_of_family')
    data = [{'id': f['id'], 'text': f"{f['family_code']} - {f['head_of_family']}"} for f in families]
    return JsonResponse({'success': True, 'results': data})


@login_required
@permission_required('finance.manage_fees', raise_exception=True)
def get_students_by_district(request, district_id):
    """API endpoint to fetch students in a district."""
    students = Student.objects.filter(
        Q(family__district_id=district_id) |
        Q(partner__district_id=district_id) |
        Q(school__district_id=district_id)
    ).distinct().values('id', 'first_name', 'last_name', 'school_name')
    
    data = []
    for s in students:
        full_name = f"{s['first_name']} {s['last_name']}"
        school = s['school_name'] or "No School"
        data.append({'id': s['id'], 'text': f"{full_name} ({school})"})
        
    return JsonResponse({'success': True, 'results': data})
