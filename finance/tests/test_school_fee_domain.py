from io import BytesIO
from decimal import Decimal

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from core.models import AcademicYear, District, Partner, Province, School
from families.models import Family
from finance.models import SchoolFee, SchoolFeeDisbursement, SchoolFeePayment
from finance.services import get_or_create_school_fee_for_enrollment, record_school_fee_payment
from students.models import Student, StudentEnrollmentHistory


class SchoolFeeDomainTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='password123',
        )
        self.client.force_login(self.user)

        self.province = Province.objects.create(name='Kigali')
        self.district_a = District.objects.create(name='Gasabo', province=self.province)
        self.district_b = District.objects.create(name='Kicukiro', province=self.province)
        self.year_2024 = AcademicYear.objects.create(name='2024-2025', is_active=False)
        self.year_2025 = AcademicYear.objects.create(name='2025-2026', is_active=True)
        self.partner_a = Partner.objects.create(name='Partner A', district=self.district_a)
        self.partner_b = Partner.objects.create(name='Partner B', district=self.district_b)

        self.school_a = School.objects.create(
            name='Alpha Primary',
            district=self.district_a,
            bank_name='Bank A',
            bank_account_name='Alpha School',
            bank_account_number='111 222',
            fee_amount=Decimal('1200.00'),
        )
        self.school_b = School.objects.create(
            name='Beta Secondary',
            district=self.district_b,
            bank_name='Bank B',
            bank_account_name='Beta School',
            bank_account_number='333 444',
            fee_amount=Decimal('2400.00'),
        )

        self.family = Family.objects.create(
            head_of_family='Parent One',
            national_id='1199999999999999',
            phone_number='0780000000',
            province=self.province,
            district=self.district_a,
            total_family_members=4,
        )

        self.student = Student.objects.create(
            family=self.family,
            partner=self.partner_a,
            first_name='Aline',
            last_name='Uwase',
            gender='F',
            date_of_birth='2012-01-01',
            school=self.school_b,
            school_name=self.school_b.name,
            class_level='Senior 1',
            school_level='secondary',
            enrollment_status='enrolled',
            sponsorship_status='active',
            is_active=True,
        )

        self.enrollment_2024 = StudentEnrollmentHistory.objects.create(
            student=self.student,
            academic_year=self.year_2024,
            school=self.school_a,
            school_name=self.school_a.name,
            class_level='Primary 6',
            school_level='primary',
        )
        self.enrollment_2025 = StudentEnrollmentHistory.objects.create(
            student=self.student,
            academic_year=self.year_2025,
            school=self.school_b,
            school_name=self.school_b.name,
            class_level='Senior 1',
            school_level='secondary',
        )

    def _create_fee(self, enrollment, term='1', total='1200.00'):
        fee, _created = get_or_create_school_fee_for_enrollment(
            enrollment,
            term,
            total_fees=Decimal(total),
            actor=self.user,
        )
        return fee

    def test_one_fee_per_student_academic_year_and_term(self):
        self._create_fee(self.enrollment_2024, term='1')
        with self.assertRaises(IntegrityError):
            SchoolFee.objects.create(
                student=self.student,
                academic_year=self.year_2024,
                term='1',
                total_fees=Decimal('500.00'),
            )

    def test_fee_uses_historical_school_and_bank_snapshot(self):
        fee = self._create_fee(self.enrollment_2024, term='1')

        self.assertEqual(fee.school, self.school_a)
        self.assertEqual(fee.school_name, 'Alpha Primary')
        self.assertEqual(fee.class_level, 'Primary 6')
        self.assertEqual(fee.bank_name, 'Bank A')
        self.assertEqual(fee.bank_account_name, 'Alpha School')
        self.assertEqual(fee.bank_account_number, '111222')

    def test_bulk_entry_is_idempotent_for_same_cumulative_amount(self):
        url = reverse('finance:bulk_fee_entry')
        payload = {
            'academic_year': self.year_2024.id,
            'school': self.school_a.id,
            'term': '1',
            'category': 'all',
            'payment_date': '2025-01-15',
            'form-TOTAL_FORMS': '1',
            'form-INITIAL_FORMS': '0',
            'form-MIN_NUM_FORMS': '0',
            'form-MAX_NUM_FORMS': '1000',
            'form-0-student_id': str(self.student.id),
            'form-0-total_fees': '1200.00',
            'form-0-amount_paid': '500.00',
        }

        response_one = self.client.post(url, payload, follow=True)
        self.assertEqual(response_one.status_code, 200)
        response_two = self.client.post(url, payload, follow=True)
        self.assertEqual(response_two.status_code, 200)

        fee = SchoolFee.objects.get(student=self.student, academic_year=self.year_2024, term='1')
        self.assertEqual(SchoolFee.objects.count(), 1)
        self.assertEqual(fee.amount_paid, Decimal('500.00'))
        self.assertEqual(fee.payments.count(), 1)

    def test_payment_service_blocks_overpayment(self):
        fee = self._create_fee(self.enrollment_2024, term='1', total='1000.00')
        record_school_fee_payment(
            fee=fee,
            amount_paid=Decimal('900.00'),
            payment_date=self.year_2024.created_at.date(),
            payment_method='bank',
            reference_number='REF-1',
            recorded_by=self.user,
            idempotency_key='payment-1',
        )
        fee.refresh_payment_summary()

        with self.assertRaises(ValidationError):
            record_school_fee_payment(
                fee=fee,
                amount_paid=Decimal('200.00'),
                payment_date=self.year_2024.created_at.date(),
                payment_method='bank',
                reference_number='REF-2',
                recorded_by=self.user,
                idempotency_key='payment-2',
            )

    def test_payment_service_handles_fee_without_nullable_relations_loaded(self):
        fee = SchoolFee.objects.create(
            student=self.student,
            academic_year=self.year_2024,
            term='3',
            total_fees=Decimal('700.00'),
            school=None,
            enrollment_history=None,
            recorded_by=self.user,
        )

        payment, created = record_school_fee_payment(
            fee=fee,
            amount_paid=Decimal('200.00'),
            payment_date=self.year_2024.created_at.date(),
            payment_method='bank',
            reference_number='NULL-JOIN-1',
            recorded_by=self.user,
            idempotency_key='null-join-1',
        )

        fee.refresh_from_db()
        self.assertTrue(created)
        self.assertEqual(payment.school_fee_id, fee.id)
        self.assertEqual(fee.amount_paid, Decimal('200.00'))
        self.assertEqual(fee.balance, Decimal('500.00'))

    def test_queue_reconciles_after_payment(self):
        fee = self._create_fee(self.enrollment_2024, term='2', total='1000.00')
        record_school_fee_payment(
            fee=fee,
            amount_paid=Decimal('400.00'),
            payment_date=self.year_2024.created_at.date(),
            payment_method='bank',
            reference_number='QUEUE-1',
            recorded_by=self.user,
            idempotency_key='queue-1',
        )

        queue_url = reverse('finance:fee_disbursement_queue')
        response = self.client.get(queue_url, {'academic_year': self.year_2024.id, 'term': '2'})
        self.assertContains(response, 'Aline Uwase')

        record_school_fee_payment(
            fee=SchoolFee.objects.get(pk=fee.pk),
            amount_paid=Decimal('600.00'),
            payment_date=self.year_2024.created_at.date(),
            payment_method='bank',
            reference_number='QUEUE-2',
            recorded_by=self.user,
            idempotency_key='queue-2',
        )

        response = self.client.get(queue_url, {'academic_year': self.year_2024.id, 'term': '2'})
        self.assertNotContains(response, 'Aline Uwase')

    def test_exports_do_not_mutate_disbursement_state(self):
        fee = self._create_fee(self.enrollment_2024, term='3', total='1000.00')
        SchoolFeeDisbursement.objects.create(
            school_fee=fee,
            status='pending',
        )

        excel_url = reverse('finance:export_fee_disbursement_excel')
        pdf_url = reverse('finance:export_fee_disbursement_pdf')

        excel_response = self.client.get(excel_url, {'academic_year': self.year_2024.id, 'term': '3'})
        pdf_response = self.client.get(pdf_url, {'academic_year': self.year_2024.id, 'term': '3'})

        self.assertEqual(excel_response.status_code, 200)
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(
            SchoolFeeDisbursement.objects.get(school_fee=fee).status,
            'pending',
        )

    def test_filters_are_consistent_across_fee_pages(self):
        other_student = Student.objects.create(
            family=self.family,
            partner=self.partner_b,
            first_name='Brian',
            last_name='Mugisha',
            gender='M',
            date_of_birth='2011-02-02',
            school=self.school_b,
            school_name=self.school_b.name,
            class_level='Senior 2',
            school_level='secondary',
            enrollment_status='enrolled',
            sponsorship_status='active',
            is_active=True,
        )
        StudentEnrollmentHistory.objects.create(
            student=other_student,
            academic_year=self.year_2024,
            school=self.school_b,
            school_name=self.school_b.name,
            class_level='Senior 2',
            school_level='secondary',
        )
        self._create_fee(self.enrollment_2024, term='1')
        other_fee, _created = get_or_create_school_fee_for_enrollment(
            other_student.enrollment_history.get(academic_year=self.year_2024),
            '1',
            total_fees=Decimal('2400.00'),
            actor=self.user,
        )
        other_fee.save()

        params = {'academic_year': self.year_2024.id, 'term': '1', 'school': self.school_a.id}
        fees_response = self.client.get(reverse('finance:fees_list'), params)
        queue_response = self.client.get(reverse('finance:fee_disbursement_queue'), params)
        dashboard_response = self.client.get(reverse('finance:school_fees_dashboard'), params)

        fee_students = [fee.student.full_name for fee in fees_response.context['fees'].object_list]
        queue_students = [fee.student.full_name for fee in queue_response.context['disbursements'].object_list]
        dashboard_schools = [group['school_name'] for group in dashboard_response.context['school_fee_groups'].object_list]

        self.assertEqual(fee_students, ['Aline Uwase'])
        self.assertEqual(queue_students, ['Aline Uwase'])
        self.assertEqual(dashboard_schools, ['Alpha Primary'])

    def test_disbursement_export_groups_pending_payments_by_school_with_partner_filter(self):
        fee_alpha = self._create_fee(self.enrollment_2024, term='1', total='1200.00')

        other_student = Student.objects.create(
            family=self.family,
            partner=self.partner_b,
            first_name='Brian',
            last_name='Mugisha',
            gender='M',
            date_of_birth='2011-02-02',
            school=self.school_b,
            school_name=self.school_b.name,
            class_level='Senior 2',
            school_level='secondary',
            enrollment_status='enrolled',
            sponsorship_status='active',
            is_active=True,
        )
        other_enrollment = StudentEnrollmentHistory.objects.create(
            student=other_student,
            academic_year=self.year_2024,
            school=self.school_b,
            school_name=self.school_b.name,
            class_level='Senior 2',
            school_level='secondary',
        )
        fee_beta, _created = get_or_create_school_fee_for_enrollment(
            other_enrollment,
            '1',
            total_fees=Decimal('2400.00'),
            actor=self.user,
        )
        fee_beta.save()

        response = self.client.get(
            reverse('finance:export_fee_disbursement_excel'),
            {
                'academic_year': self.year_2024.id,
                'term': '1',
                'partner': self.partner_a.id,
                'school': self.school_a.id,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('pending_school_fee_payments_by_school.xlsx', response['Content-Disposition'])

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        flattened = ' | '.join(str(value) for row in rows for value in row if value is not None)

        self.assertIn(f'School: {self.school_a.name}', flattened)
        self.assertIn(f'Bank: {self.school_a.bank_name}', flattened)
        self.assertIn('Aline Uwase', flattened)
        self.assertIn('Partner A', flattened)
        self.assertNotIn('Brian Mugisha', flattened)
        self.assertNotIn(fee_beta.school_name, flattened)
        self.assertEqual(fee_alpha.school_name, self.school_a.name)

    def test_reconciliation_uses_explicit_scope_not_ui_filters(self):
        fee_alpha = self._create_fee(self.enrollment_2024, term='1')

        other_student = Student.objects.create(
            family=self.family,
            first_name='Cedric',
            last_name='Nshimiyimana',
            gender='M',
            date_of_birth='2011-03-03',
            school=self.school_b,
            school_name=self.school_b.name,
            class_level='Senior 2',
            school_level='secondary',
            enrollment_status='enrolled',
            sponsorship_status='active',
            is_active=True,
        )
        other_enrollment = StudentEnrollmentHistory.objects.create(
            student=other_student,
            academic_year=self.year_2024,
            school=self.school_b,
            school_name=self.school_b.name,
            class_level='Senior 2',
            school_level='secondary',
        )
        fee_beta, _created = get_or_create_school_fee_for_enrollment(
            other_enrollment,
            '1',
            total_fees=Decimal('2400.00'),
            actor=self.user,
        )
        fee_beta.save()

        list_response = self.client.get(
            reverse('finance:fee_disbursement_queue'),
            {'academic_year': self.year_2024.id, 'term': '1', 'school': self.school_a.id},
        )
        visible_students = [fee.student.full_name for fee in list_response.context['disbursements'].object_list]
        self.assertEqual(visible_students, ['Aline Uwase'])

        sync_response = self.client.post(
            reverse('finance:sync_fee_disbursement_queue'),
            {'academic_year': self.year_2024.id, 'term': '1'},
            follow=True,
        )
        self.assertEqual(sync_response.status_code, 200)
        self.assertEqual(SchoolFeeDisbursement.objects.filter(status='pending').count(), 2)
        self.assertTrue(SchoolFeeDisbursement.objects.filter(school_fee=fee_alpha).exists())
        self.assertTrue(SchoolFeeDisbursement.objects.filter(school_fee=fee_beta).exists())
