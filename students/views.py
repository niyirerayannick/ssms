from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.db import transaction
from django.db.models import (
    Q,
    Avg,
    Count,
    Sum,
    Case,
    When,
    Value,
    BooleanField,
    Prefetch,
    DecimalField,
)
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.views.generic import ListView, DetailView
from django.db.models.functions import Coalesce
from django.http import Http404, HttpResponse
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from core.models import District, School, Partner
from core.export_utils import (
    ExportNumberedCanvas,
    add_export_header,
    autosize_worksheet_columns,
    build_export_pdf_document,
    build_export_table,
    prepend_row_numbers,
    resolve_logo_path,
    style_excel_header,
    style_excel_table_rows,
    write_excel_report_header,
)
from django.utils import timezone
from django.core import signing
from django.views.decorators.http import require_POST
from django.contrib.auth import get_user_model
from django.forms import formset_factory
from urllib.parse import urlencode
from .models import (
    Student,
    StudentPhoto,
    StudentMark,
    StudentMaterial,
    sync_student_enrollment_history,
)
from core.models import Notification, AcademicYear
from core.academic_years import get_default_academic_year
from .forms import (
    StudentForm,
    StudentPhotoForm,
    StudentMarkForm,
    StudentMaterialForm,
    AcademicYearPromotionForm,
    BulkPerformanceFilterForm,
    BulkStudentMarkForm,
    BulkMaterialFilterForm,
    BulkStudentMaterialForm,
)
from families.models import FamilyStudent
from finance.models import SchoolFee
from insurance.models import FamilyInsurance
from students.services.promotion import promote_students_to_academic_year


@login_required
@permission_required('students.view_student', raise_exception=True)
def student_list(request):
    """List all students with search and filters."""
    students = Student.objects.select_related(
        'school', 'family__district', 'partner__district'
    ).all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(family__district__name__icontains=search_query) |
            Q(school__name__icontains=search_query)
        )
    
    # Filter by sponsorship status
    status_filter = request.GET.get('status', '')
    if status_filter:
        students = students.filter(sponsorship_status=status_filter)
    
    # Filter by gender
    gender_filter = request.GET.get('gender', '')
    if gender_filter:
        students = students.filter(gender=gender_filter)

    # Filter by district (family district)
    district_filter = request.GET.get('district', '')
    if district_filter:
        students = students.filter(
            Q(family__district_id=district_filter) |
            Q(partner__district_id=district_filter)
        )

    # Filter by partner
    partner_filter = request.GET.get('partner', '')
    if partner_filter:
        students = students.filter(partner_id=partner_filter)

    school_level_filter = request.GET.get('school_level', '')
    if school_level_filter:
        students = students.filter(school_level=school_level_filter)

    summary_queryset = students.distinct()
    boarding_counts = {
        item['boarding_status']: item['total']
        for item in summary_queryset.values('boarding_status').annotate(total=Count('id'))
    }
    level_counts = {
        item['school_level']: item['total']
        for item in summary_queryset.values('school_level').annotate(total=Count('id'))
    }
    
    # Pagination
    paginator = Paginator(summary_queryset.order_by('first_name', 'last_name'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'students': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'gender_filter': gender_filter,
        'district_filter': district_filter,
        'partner_filter': partner_filter,
        'school_level_filter': school_level_filter,
        'school_level_choices': Student.SCHOOL_LEVEL_CHOICES,
        'districts': District.objects.order_by('name'),
        'partners': Partner.objects.order_by('name'),
        'boarding_count': boarding_counts.get('boarding', 0),
        'non_boarding_count': boarding_counts.get('non_boarding', 0),
        'nursery_count': level_counts.get('nursery', 0),
        'primary_count': level_counts.get('primary', 0),
        'secondary_count': level_counts.get('secondary', 0),
        'tvet_count': level_counts.get('tvet', 0),
        'university_count': level_counts.get('university', 0),
    }
    return render(request, 'students/student_list.html', context)


def _notify_admins_and_exec(request, actor, verb, link, send_email=False):
    User = get_user_model()
    # Users who should receive notifications and emails (Admins, Execs, Superusers)
    admin_exec_users = User.objects.filter(
        Q(groups__name__in=['Admin', 'Executive Secretary']) |
        Q(is_superuser=True)
    ).distinct()

    # Set of IDs for notification creation
    recipient_ids = set(admin_exec_users.values_list('id', flat=True))
    
    # Also include staff in notifications (but maybe not email if not admin/exec)
    staff_ids = User.objects.filter(is_staff=True).values_list('id', flat=True)
    recipient_ids.update(staff_ids)

    if actor:
        recipient_ids.add(actor.id)
    
    recipients = User.objects.filter(id__in=recipient_ids)
    
    # Create DB Notifications
    notifications = [
        Notification(recipient=user, actor=actor, verb=verb, link=link)
        for user in recipients
    ]
    if notifications:
        Notification.objects.bulk_create(notifications)

    # Send Email if requested (Only to Admins and Executive Secretaries)
    if send_email:
        email_recipients = admin_exec_users.exclude(email='').values_list('email', flat=True)
        
        if email_recipients:
            subject = f"SIMS Notification: {verb}"
            full_link = request.build_absolute_uri(link)
            actor_name = actor.get_full_name() or actor.username
            
            # Prepare HTML content
            context = {
                'verb': verb,
                'actor_name': actor_name,
                'full_link': full_link,
            }
            html_message = render_to_string('emails/notification.html', context)
            plain_message = strip_tags(html_message)
            
            try:
                send_mail(
                    subject=subject,
                    message=plain_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=list(email_recipients),
                    html_message=html_message,
                    fail_silently=True
                )
            except Exception as e:
                print(f"Failed to send email: {e}")


def _can_approve_student(user):
    return (
        user.is_superuser or
        user.groups.filter(name__in=['Admin', 'Executive Secretary']).exists()
    )


def _build_pdf_table(data, col_widths=None, header_background='#0f766e', body_font_size=8):
    """Create a consistently styled PDF table."""
    return build_export_table(
        data,
        col_widths=col_widths,
        header_background=colors.HexColor(header_background),
        body_font_size=body_font_size,
    )


def _build_pdf_image(field_file, width=60, height=60):
    """Create a ReportLab image from Django storage without relying on absolute paths."""

    if not field_file:
        return None

    try:
        field_file.open('rb')
        image_buffer = BytesIO(field_file.read())
        image_buffer.seek(0)
        return Image(image_buffer, width=width, height=height)
    except Exception:
        return None
    finally:
        try:
            field_file.close()
        except Exception:
            pass


def _resolve_logo_path():
    """Return the first existing local logo asset path."""
    return resolve_logo_path()


def _pdf_text(value, fallback='N/A'):
    """Return a safe string for PDF output."""

    if value is None:
        return fallback
    if isinstance(value, str):
        value = value.strip()
        return value or fallback
    return str(value)


@login_required
@permission_required('students.add_student', raise_exception=True)
def student_create(request):
    """Create a new student."""
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save(commit=False)
            student.sponsorship_status = 'pending'
            student.save()
            form.save_m2m()
            default_year = get_default_academic_year()
            if default_year:
                sync_student_enrollment_history(student, default_year, overwrite=True)
            _notify_admins_and_exec(
                request=request,
                actor=request.user,
                verb=f"Added student {student.full_name}",
                link=reverse('students:student_detail', kwargs={'pk': student.pk}),
                send_email=True
            )
            messages.success(request, f'Student {student.full_name} created successfully!')
            return redirect('students:student_detail', pk=student.pk)
    else:
        form = StudentForm()
    
    return render(request, 'students/student_form.html', {'form': form, 'title': 'Add Student'})


@login_required
@permission_required('students.change_student', raise_exception=True)
def student_edit(request, pk):
    """Edit an existing student."""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            student = form.save()
            default_year = get_default_academic_year()
            if default_year:
                sync_student_enrollment_history(student, default_year, overwrite=True)
            _notify_admins_and_exec(
                request=request,
                actor=request.user,
                verb=f"Updated student {student.full_name}",
                link=reverse('students:student_detail', kwargs={'pk': student.pk})
            )
            messages.success(request, f'Student {student.full_name} updated successfully!')
            return redirect('students:student_detail', pk=student.pk)
    else:
        form = StudentForm(instance=student)
    
    return render(request, 'students/student_form.html', {'form': form, 'student': student, 'title': 'Edit Student'})


@login_required
@permission_required('students.view_studentmaterial', raise_exception=True)
def student_materials(request):
    """List sponsored students with material status, with export support."""
    context = _build_student_materials_context(request)
    export_format = request.GET.get('export', '').strip().lower()
    if export_format == 'pdf':
        return _export_student_materials_pdf(context)
    if export_format == 'excel':
        return _export_student_materials_excel(context)
    return render(request, 'students/student_materials.html', context)


def _material_filter_params(request):
    academic_year_param = request.GET.get('academic_year', '').strip()
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    district_filter = request.GET.get('district', '').strip()

    active_year = get_default_academic_year()
    academic_year_id = None
    if academic_year_param:
        try:
            academic_year_id = int(academic_year_param)
        except (TypeError, ValueError):
            academic_year_id = None
    if academic_year_id is None and active_year:
        academic_year_id = active_year.id
    if academic_year_id is None:
        academic_year_id = AcademicYear.objects.order_by('-name').values_list('id', flat=True).first()

    return {
        'academic_year_id': academic_year_id,
        'status_filter': status_filter,
        'search_query': search_query,
        'district_filter': district_filter,
    }


def _build_student_materials_context(request):
    filters = _material_filter_params(request)
    academic_year_id = filters['academic_year_id']
    status_filter = filters['status_filter']
    search_query = filters['search_query']
    district_filter = filters['district_filter']

    selected_academic_year = (
        AcademicYear.objects.filter(id=academic_year_id).first()
        if academic_year_id else None
    )
    selected_district = (
        District.objects.filter(id=district_filter).first()
        if district_filter else None
    )

    students_qs = (
        Student.objects.filter(sponsorship_status='active')
        .select_related('family__district', 'partner__district', 'school')
    )
    if search_query:
        students_qs = students_qs.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(family__family_code__icontains=search_query)
        )
    if district_filter:
        students_qs = students_qs.filter(
            Q(family__district_id=district_filter) |
            Q(partner__district_id=district_filter)
        )

    students = list(students_qs.order_by('first_name', 'last_name').distinct())
    materials_qs = StudentMaterial.objects.select_related('student', 'academic_year')
    if academic_year_id:
        materials_qs = materials_qs.filter(academic_year_id=academic_year_id)
    materials_by_student = {material.student_id: material for material in materials_qs}

    rows = []
    for student in students:
        material = materials_by_student.get(student.id)
        has_all_required = material.all_required_received if material else False
        if status_filter == 'complete' and not has_all_required:
            continue
        if status_filter == 'missing' and has_all_required:
            continue
        district_name = student.partner.district.name if student.partner and student.partner.district else (
            student.family.district.name if student.family and student.family.district else 'N/A'
        )
        rows.append({
            'student': student,
            'material': material,
            'has_all_required': has_all_required,
            'district_name': district_name,
        })

    boarding_counts = {
        item['boarding_status']: item['total']
        for item in students_qs.values('boarding_status').annotate(total=Count('id'))
    }
    level_counts = {
        item['school_level']: item['total']
        for item in students_qs.values('school_level').annotate(total=Count('id'))
    }

    total_rows = len(rows)
    complete_rows = sum(1 for row in rows if row['has_all_required'])
    missing_rows = total_rows - complete_rows

    params = request.GET.copy()
    params.pop('export', None)
    querystring = params.urlencode()
    path = request.path

    context = {
        'rows': rows,
        'selected_academic_year_id': academic_year_id,
        'selected_academic_year': selected_academic_year,
        'selected_district': selected_district,
        'status_filter': status_filter,
        'search_query': search_query,
        'district_filter': district_filter,
        'districts': District.objects.order_by('name'),
        'available_years': AcademicYear.objects.order_by('-name'),
        'academic_year_label': selected_academic_year.name if selected_academic_year else 'All Years',
        'district_label': selected_district.name if selected_district else 'All Districts',
        'total_rows': total_rows,
        'complete_rows': complete_rows,
        'missing_rows': missing_rows,
        'boarding_count': boarding_counts.get('boarding', 0),
        'non_boarding_count': boarding_counts.get('non_boarding', 0),
        'nursery_count': level_counts.get('nursery', 0),
        'primary_count': level_counts.get('primary', 0),
        'secondary_count': level_counts.get('secondary', 0),
        'tvet_count': level_counts.get('tvet', 0),
        'export_pdf_url': f"{path}?{querystring}&export=pdf" if querystring else f"{path}?export=pdf",
        'export_excel_url': f"{path}?{querystring}&export=excel" if querystring else f"{path}?export=excel",
    }
    return context


def _export_student_materials_pdf(context):
    buffer = BytesIO()
    doc = build_export_pdf_document(
        buffer,
        'Student Materials Report',
        pagesize=landscape(A4),
        left_margin=24,
        right_margin=24,
        top_margin=24,
        bottom_margin=24,
    )
    elements = []
    add_export_header(
        elements,
        'Student Materials Report',
        (
            f"Academic Year: {context['academic_year_label']} | "
            f"District: {context['district_label']} | "
            f"Status: {context['status_filter'] or 'all'}"
        ),
        generated_label=f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}",
    )
    table_data = [[
        'No.',
        'Name',
        'District',
        'Level',
        'Bag',
        'Books',
        'Pens',
        'Rulers',
        'Drawing',
        'Registers',
        'Math Set',
        'Calculator',
        'Periodic',
        'Dup Papers',
        'Pads',
    ]]

    for index, row in enumerate(context['rows'], start=1):
        student = row['student']
        material = row['material']
        table_data.append([
            str(index),
            student.full_name,
            row['district_name'],
            student.school_level or 'N/A',
            'V' if material and material.bag_received else 'X',
            'V' if material and material.books_received else 'X',
            'V' if material and material.pens_pencils_received else 'X',
            'V' if material and material.rulers_erasers_received else 'X',
            'V' if material and material.drawing_books_received else 'X',
            'V' if material and material.register_files_received else 'X',
            'V' if material and material.mathematical_sets_received else 'X',
            'V' if material and material.scientific_calculators_received else 'X',
            'V' if material and material.periodic_tables_received else 'X',
            'V' if material and material.duplicating_papers_received else 'X',
            'V' if material and material.sanitary_pads_received else 'X',
        ])

    table = _build_pdf_table(
        table_data,
        [30, 120, 80, 55, 38, 40, 38, 42, 42, 45, 45, 45, 42, 48, 38],
        body_font_size=6,
    )
    for row_index in range(1, len(table_data)):
        for col_index in range(4, len(table_data[0])):
            cell_value = table_data[row_index][col_index]
            color = colors.HexColor('#15803d') if cell_value == 'V' else colors.HexColor('#dc2626')
            table.setStyle(TableStyle([
                ('TEXTCOLOR', (col_index, row_index), (col_index, row_index), color),
                ('FONTNAME', (col_index, row_index), (col_index, row_index), 'Helvetica-Bold'),
                ('ALIGN', (col_index, row_index), (col_index, row_index), 'CENTER'),
            ]))
    elements.append(
        table
    )
    doc.build(elements, canvasmaker=ExportNumberedCanvas)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="student_materials_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf)
    return response


def _export_student_materials_excel(context):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Materials'

    subtitle = (
        f"Academic Year: {context['academic_year_label']} | "
        f"District: {context['district_label']} | "
        f"Status: {context['status_filter'] or 'all'}"
    )
    headers = [
        'No.',
        'Name',
        'District',
        'Level',
        'Bag',
        'Books',
        'Pens',
        'Rulers',
        'Drawing',
        'Registers',
        'Math Set',
        'Calculator',
        'Periodic',
        'Dup Papers',
        'Pads',
    ]
    header_row_idx = write_excel_report_header(worksheet, 'Student Materials Report', subtitle, len(headers))
    worksheet.append(headers)
    style_excel_header(worksheet, header_row_idx, fill_color='0F172A')

    data_start_row = header_row_idx + 1
    for index, row in enumerate(context['rows'], start=1):
        student = row['student']
        material = row['material']
        worksheet.append([
            index,
            student.full_name,
            row['district_name'],
            student.school_level or 'N/A',
            'V' if material and material.bag_received else 'X',
            'V' if material and material.books_received else 'X',
            'V' if material and material.pens_pencils_received else 'X',
            'V' if material and material.rulers_erasers_received else 'X',
            'V' if material and material.drawing_books_received else 'X',
            'V' if material and material.register_files_received else 'X',
            'V' if material and material.mathematical_sets_received else 'X',
            'V' if material and material.scientific_calculators_received else 'X',
            'V' if material and material.periodic_tables_received else 'X',
            'V' if material and material.duplicating_papers_received else 'X',
            'V' if material and material.sanitary_pads_received else 'X',
        ])

        current_row = worksheet.max_row
        for col_idx in range(5, 16):
            cell = worksheet.cell(row=current_row, column=col_idx)
            cell.font = Font(
                bold=True,
                color='15803D' if cell.value == 'V' else 'DC2626',
            )
            cell.alignment = Alignment(horizontal='center')

    column_widths = [8, 28, 18, 14, 10, 10, 10, 10, 12, 12, 12, 12, 12, 12, 10]
    for idx, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[chr(64 + idx)].width = width
    style_excel_table_rows(
        worksheet,
        header_row_idx=header_row_idx,
        data_start_row=data_start_row,
        data_end_row=worksheet.max_row,
        max_col=len(headers),
        centered_columns=[1, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15],
    )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="student_materials_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    response.write(buffer.read())
    buffer.close()
    return response


@login_required
@permission_required('students.add_studentmaterial', raise_exception=True)
def student_material_bulk_entry(request):
    """Bulk capture materials filtered by academic year and district."""
    data_source = request.POST if request.method == 'POST' else request.GET
    filter_form = BulkMaterialFilterForm(data_source or None)
    MaterialBulkFormSet = formset_factory(BulkStudentMaterialForm, extra=0)

    formset = None
    students_loaded = False
    table_rows = []
    selected_filters = {}
    selected_year = None
    selected_district = None
    summary = {
        'student_count': 0,
        'records_existing': 0,
        'pending_records': 0,
        'complete_records': 0,
    }

    tracked_fields = [
        'bag_received',
        'books_received',
        'pens_pencils_received',
        'rulers_erasers_received',
        'drawing_books_received',
        'register_files_received',
        'mathematical_sets_received',
        'scientific_calculators_received',
        'periodic_tables_received',
        'duplicating_papers_received',
        'sanitary_pads_received',
        'shoes_received',
        'uniforms_received',
        'received_date',
        'notes',
    ]

    if filter_form.is_valid():
        students_loaded = True
        academic_year = filter_form.cleaned_data['academic_year']
        district = filter_form.cleaned_data['district']
        selected_year = academic_year
        selected_district = district

        student_qs = (
            Student.objects.filter(
                sponsorship_status='active'
            )
            .filter(
                Q(family__district=district) | Q(partner__district=district)
            )
            .select_related('family__district', 'partner__district', 'school')
            .order_by('first_name', 'last_name')
            .distinct()
        )
        students = list(student_qs)
        student_lookup = {student.id: student for student in students}

        existing_materials = StudentMaterial.objects.filter(
            student__in=students,
            academic_year=academic_year,
        ).select_related('student')
        existing_map = {record.student_id: record for record in existing_materials}

        summary = {
            'student_count': len(students),
            'records_existing': len(existing_map),
            'pending_records': max(len(students) - len(existing_map), 0),
            'complete_records': sum(1 for record in existing_map.values() if record.all_required_received),
        }

        initial_data = []
        for student in students:
            material = existing_map.get(student.id)
            initial_data.append({
                'student_id': student.id,
                'bag_received': material.bag_received if material else False,
                'books_received': material.books_received if material else False,
                'pens_pencils_received': material.pens_pencils_received if material else False,
                'rulers_erasers_received': material.rulers_erasers_received if material else False,
                'drawing_books_received': material.drawing_books_received if material else False,
                'register_files_received': material.register_files_received if material else False,
                'mathematical_sets_received': material.mathematical_sets_received if material else False,
                'scientific_calculators_received': material.scientific_calculators_received if material else False,
                'periodic_tables_received': material.periodic_tables_received if material else False,
                'duplicating_papers_received': material.duplicating_papers_received if material else False,
                'sanitary_pads_received': material.sanitary_pads_received if material else False,
                'shoes_received': material.shoes_received if material else False,
                'uniforms_received': material.uniforms_received if material else False,
                'received_date': material.received_date if material else None,
                'notes': material.notes if material else '',
            })

        if request.method == 'POST':
            formset = MaterialBulkFormSet(request.POST)
            if formset.is_valid():
                created_count = 0
                updated_count = 0
                with transaction.atomic():
                    for form in formset:
                        student_id = form.cleaned_data.get('student_id')
                        if not student_id or student_id not in student_lookup:
                            continue

                        defaults = {field: form.cleaned_data.get(field) for field in tracked_fields}
                        record, created = StudentMaterial.objects.update_or_create(
                            student=student_lookup[student_id],
                            academic_year=academic_year,
                            defaults=defaults,
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                        existing_map[student_id] = record

                messages.success(
                    request,
                    f"Bulk materials saved for {district.name}: {created_count} new, {updated_count} updated."
                )
                query = urlencode({
                    'academic_year': academic_year.id,
                    'district': district.id,
                })
                return redirect(f"{reverse('students:student_material_bulk_entry')}?{query}")
        else:
            formset = MaterialBulkFormSet(initial=initial_data)

        if formset is not None:
            for idx, student in enumerate(students):
                if idx >= len(formset.forms):
                    break
                existing_record = existing_map.get(student.id)
                table_rows.append({
                    'student': student,
                    'form': formset.forms[idx],
                    'existing_record': existing_record,
                    'requires_secondary_materials': student.school_level == 'secondary',
                    'requires_sanitary_pads': student.gender == 'F' and (student.age or 0) >= 12,
                    'district_name': student.partner.district.name if student.partner and student.partner.district else (
                        student.family.district.name if student.family and student.family.district else 'N/A'
                    ),
                })

        selected_filters = {
            'academic_year': academic_year.id,
            'district': district.id,
        }

    context = {
        'filter_form': filter_form,
        'formset': formset,
        'table_rows': table_rows,
        'students_loaded': students_loaded,
        'selected_filters': selected_filters,
        'selected_year': selected_year,
        'selected_district': selected_district,
        'summary': summary,
    }
    return render(request, 'students/student_material_bulk_entry.html', context)


@login_required
@permission_required('students.add_studentmaterial', raise_exception=True)
def student_material_create(request):
    """Create a school material record."""
    initial = {}
    student_id = request.GET.get('student')
    if student_id:
        try:
            initial['student'] = int(student_id)
        except (TypeError, ValueError):
            initial['student'] = student_id
    academic_year_param = request.GET.get('academic_year')
    if academic_year_param:
        try:
            initial['academic_year'] = int(academic_year_param)
        except (TypeError, ValueError):
            pass
    if request.method == 'POST':
        form = StudentMaterialForm(request.POST)
        if form.is_valid():
            record = form.save()
            messages.success(request, f"Materials saved for {record.student.full_name}.")
            redirect_url = reverse('students:student_materials')
            if record.academic_year_id:
                redirect_url = f"{redirect_url}?academic_year={record.academic_year_id}"
            return redirect(redirect_url)
    else:
        form = StudentMaterialForm(initial=initial)
    return render(request, 'students/student_material_form.html', {'form': form, 'title': 'Add School Materials'})


@login_required
@permission_required('students.change_studentmaterial', raise_exception=True)
def student_material_edit(request, pk):
    """Edit a school material record."""
    record = get_object_or_404(StudentMaterial, pk=pk)
    if request.method == 'POST':
        form = StudentMaterialForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, f"Materials updated for {record.student.full_name}.")
            redirect_url = reverse('students:student_materials')
            if record.academic_year_id:
                redirect_url = f"{redirect_url}?academic_year={record.academic_year_id}"
            return redirect(redirect_url)
    else:
        form = StudentMaterialForm(instance=record)
    return render(request, 'students/student_material_form.html', {'form': form, 'title': 'Edit School Materials'})


@login_required
@permission_required('students.change_student', raise_exception=True)
def academic_year_promotion(request):
    """Run yearly student promotion from the UI."""
    promotion_summary = None

    if request.method == 'POST':
        form = AcademicYearPromotionForm(request.POST)
        if form.is_valid():
            promotion_summary = promote_students_to_academic_year(
                form.cleaned_data['source_year'],
                form.cleaned_data['target_year'],
                overwrite=form.cleaned_data['overwrite_existing'],
                include_inactive=form.cleaned_data['include_inactive'],
                activate_target=form.cleaned_data['activate_target'],
            )
            messages.success(
                request,
                (
                    f"Promotion completed from {promotion_summary.source_year.name} to {promotion_summary.target_year.name}: "
                    f"{promotion_summary.created_count} created, {promotion_summary.updated_count} updated, "
                    f"{promotion_summary.skipped_count} skipped, {promotion_summary.graduated_count} graduated."
                ),
            )
    else:
        form = AcademicYearPromotionForm()

    context = {
        'form': form,
        'title': 'Academic Year Promotion',
        'promotion_summary': promotion_summary,
    }
    return render(request, 'students/academic_year_promotion.html', context)


@login_required
@permission_required('students.view_student', raise_exception=True)
def student_detail(request, pk):
    """View student profile with family, fees, and insurance information."""
    student = get_object_or_404(Student, pk=pk)
    
    # Get related data
    family_member = getattr(student, 'family_member', None)
    family = student.family or (family_member.family if family_member else None)
    fees = student.fees.all()
    insurance_records = family.insurance_records.all() if family else []
    
    context = {
        'student': student,
        'family': family,
        'family_member': family_member,
        'fees': fees,
        'insurance_records': insurance_records,
        'can_approve_student': _can_approve_student(request.user),
    }
    return render(request, 'students/student_detail.html', context)


@login_required
@permission_required('students.view_student', raise_exception=True)
def student_full_report_pdf(request, pk):
    """Export a complete student report across all academic years."""

    student = get_object_or_404(
        Student.objects.select_related(
            'family',
            'family__district',
            'family__sector',
            'family__cell',
            'family__village',
            'school',
            'partner',
            'program_officer',
        ),
        pk=pk,
    )
    family_member = getattr(student, 'family_member', None)
    family = student.family or (family_member.family if family_member else None)

    enrollment_history = list(
        student.enrollment_history.select_related('academic_year', 'school').order_by('-academic_year__name')
    )
    fees = list(
        student.fees.select_related('academic_year').order_by('-academic_year__name', 'term')
    )
    marks = list(
        student.academic_records.select_related('academic_year').order_by('-academic_year__name', 'term', 'subject')
    )
    materials = list(
        student.material_records.select_related('academic_year').order_by('-academic_year__name')
    )
    photos = list(student.photos.all().order_by('-created_at'))
    insurance_records = list(
        family.insurance_records.select_related('insurance_year').order_by('-insurance_year__name')
    ) if family else []

    total_fees_required = sum((fee.total_fees for fee in fees), 0)
    total_fees_paid = sum((fee.amount_paid for fee in fees), 0)
    total_fees_balance = sum((fee.balance for fee in fees), 0)

    buffer = BytesIO()
    doc = build_export_pdf_document(
        buffer,
        'Student Comprehensive Report',
        pagesize=A4,
        left_margin=36,
        right_margin=36,
        top_margin=36,
        bottom_margin=36,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'StudentReportTitle',
        parent=styles['Title'],
        fontSize=20,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=6,
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        'StudentReportSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#475569'),
        spaceAfter=6,
        alignment=1,
    )
    section_style = ParagraphStyle(
        'StudentReportSection',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#0f766e'),
        spaceBefore=8,
        spaceAfter=8,
    )
    body_style = ParagraphStyle(
        'StudentReportBody',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=4,
    )
    small_style = ParagraphStyle(
        'StudentReportSmall',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#475569'),
        spaceAfter=4,
        alignment=1,
    )

    elements = []
    add_export_header(
        elements,
        'Student Comprehensive Report',
        student.full_name,
        generated_label=f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')} | Student ID {student.pk}",
    )

    photo_entries = []
    if getattr(student, 'profile_picture', None):
        photo_entries.append({
            'file': student.profile_picture,
            'caption': 'Profile Picture',
            'created_at': None,
            'source': 'Profile',
        })
    for photo in photos:
        photo_entries.append({
            'file': photo.image,
            'caption': photo.caption or 'Student Photo',
            'created_at': photo.created_at,
            'source': 'Camera' if photo.captured_via_camera else 'Upload',
        })

    if photo_entries:
        primary_photo = _build_pdf_image(photo_entries[0]['file'], width=110, height=110)
        if primary_photo is not None:
            primary_photo.hAlign = 'CENTER'
            elements.append(primary_photo)
            elements.append(Spacer(1, 8))
        caption_parts = [_pdf_text(photo_entries[0]['caption']), _pdf_text(photo_entries[0]['source'])]
        if photo_entries[0]['created_at']:
            caption_parts.insert(1, photo_entries[0]['created_at'].strftime('%Y-%m-%d'))
        elements.append(Paragraph(' | '.join(caption_parts), small_style))
        elements.append(Spacer(1, 10))

    personal_data = [
        ['Field', 'Value', 'Field', 'Value'],
        ['Full Name', student.full_name, 'Gender', student.get_gender_display()],
        ['Date of Birth', student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else 'N/A', 'Age', str(student.age or 'N/A')],
        ['Current School', student.school.name if student.school else (student.school_name or 'N/A'), 'Current Class', student.class_level or 'N/A'],
        ['School Level', student.get_school_level_display() if student.school_level else 'N/A', 'Boarding', student.get_boarding_status_display()],
        ['Enrollment Status', student.get_enrollment_status_display(), 'Sponsorship Status', student.get_sponsorship_status_display()],
        ['Partner', student.partner.name if student.partner else 'N/A', 'Program Officer', student.program_officer.get_full_name() if student.program_officer else 'N/A'],
        ['Location', student.location_display or 'N/A', 'Active Account', 'Yes' if student.is_active else 'No'],
    ]
    if student.sponsorship_reason:
        personal_data.append(['Support Reason', student.sponsorship_reason, 'Disability', student.disability_display])
    else:
        personal_data.append(['Disability', student.disability_display, '', ''])
    elements.append(Paragraph('Personal Information', section_style))
    elements.append(_build_pdf_table(personal_data, [95, 180, 95, 150], body_font_size=8))
    elements.append(Spacer(1, 10))

    family_data = [['Field', 'Value']]
    if family:
        family_data.extend([
            ['Family Code', family.family_code],
            ['Head of Family', family.head_of_family],
            ['Phone Number', family.phone_number],
            ['Alternative Phone', family.alternative_phone or 'N/A'],
            ['Father Name', family.father_name or 'N/A'],
            ['Mother Name', family.mother_name or 'N/A'],
            ['Guardian', family.guardian_name or 'N/A'],
            ['Guardian Phone', family.guardian_phone or 'N/A'],
            ['Total Family Members', str(family.total_family_members or 0)],
            ['Location', family.location_display or 'N/A'],
            ['Address Description', family.address_description or 'N/A'],
            ['Notes', family.notes or 'N/A'],
        ])
    else:
        family_data.append(['Family Information', 'No linked family record'])
    elements.append(Paragraph('Family Information', section_style))
    elements.append(_build_pdf_table(family_data, [140, 380], body_font_size=8))
    elements.append(Spacer(1, 10))

    history_data = [['Academic Year', 'School', 'Class', 'Level', 'Promoted On']]
    if enrollment_history:
        for row in enrollment_history:
            history_data.append([
                row.academic_year.name if row.academic_year else 'N/A',
                row.display_school_name,
                row.class_level or 'N/A',
                row.get_school_level_display() if row.school_level else 'N/A',
                row.promoted_on.strftime('%Y-%m-%d') if row.promoted_on else 'N/A',
            ])
    else:
        history_data.append(['No academic year history recorded', '', '', '', ''])
    elements.append(Paragraph('School History By Academic Year', section_style))
    elements.append(_build_pdf_table(history_data, [90, 220, 80, 90, 80], body_font_size=8))
    elements.append(Spacer(1, 10))

    fees_data = [['Academic Year', 'Term', 'School', 'Class', 'Required', 'Paid', 'Balance', 'Status']]
    if fees:
        for fee in fees:
            fees_data.append([
                fee.academic_year.name if fee.academic_year else 'N/A',
                fee.get_term_display(),
                fee.school_name or (student.school.name if student.school else 'N/A'),
                fee.class_level or 'N/A',
                f'{fee.total_fees:,.2f}',
                f'{fee.amount_paid:,.2f}',
                f'{fee.balance:,.2f}',
                fee.get_payment_status_display(),
            ])
        fees_data.append([
            'TOTAL', '', '', '',
            f'{total_fees_required:,.2f}',
            f'{total_fees_paid:,.2f}',
            f'{total_fees_balance:,.2f}',
            '',
        ])
    else:
        fees_data.append(['No fee records found', '', '', '', '', '', '', ''])
    elements.append(Paragraph('School Fees Across All Academic Years', section_style))
    elements.append(_build_pdf_table(fees_data, [75, 52, 140, 60, 55, 55, 55, 65], body_font_size=7))
    elements.append(Spacer(1, 10))

    marks_data = [['Academic Year', 'Term', 'Subject', 'Marks', 'Remark']]
    if marks:
        for record in marks:
            marks_data.append([
                record.academic_year.name if record.academic_year else 'N/A',
                record.term,
                record.subject,
                f'{record.marks}',
                record.teacher_remark or 'N/A',
            ])
    else:
        marks_data.append(['No marks found', '', '', '', ''])
    elements.append(Paragraph('Academic Marks Across All Academic Years', section_style))
    elements.append(_build_pdf_table(marks_data, [85, 60, 130, 55, 210], body_font_size=7))
    elements.append(Spacer(1, 10))

    materials_data = [[
        'Academic Year',
        'Standard Package',
        'Secondary Tools',
        'Girls 12+',
        'Optional Extras',
        'Received Date',
        'Notes',
    ]]
    if materials:
        for material in materials:
            standard_items = [
                ('Backpack', material.bag_received),
                ('Notebooks', material.books_received),
                ('Pens/Pencils', material.pens_pencils_received),
                ('Rulers/Erasers', material.rulers_erasers_received),
                ('Drawing Books', material.drawing_books_received),
                ('Register Files', material.register_files_received),
                ('Math Set', material.mathematical_sets_received),
            ]
            secondary_items = [
                ('Calculator', material.scientific_calculators_received),
                ('Periodic Table', material.periodic_tables_received),
                ('Duplicating Papers', material.duplicating_papers_received),
            ]
            optional_items = [
                ('Shoes', material.shoes_received),
                ('Uniforms', material.uniforms_received),
            ]
            materials_data.append([
                material.academic_year.name if material.academic_year else 'N/A',
                ', '.join(label for label, present in standard_items if present) or 'None',
                ', '.join(label for label, present in secondary_items if present) or 'N/A',
                'Pads included' if material.sanitary_pads_received else ('Required but missing' if material.requires_sanitary_pads else 'N/A'),
                ', '.join(label for label, present in optional_items if present) or 'None',
                material.received_date.strftime('%Y-%m-%d') if material.received_date else 'N/A',
                material.notes or material.special_request or 'N/A',
            ])
    else:
        materials_data.append(['No material records found', '', '', '', '', '', ''])
    elements.append(Paragraph('Student Materials', section_style))
    elements.append(_build_pdf_table(materials_data, [75, 130, 110, 70, 80, 65, 110], body_font_size=7))
    elements.append(Spacer(1, 10))

    insurance_data = [['Academic Year', 'Required', 'Paid', 'Balance', 'Coverage Status', 'Remarks']]
    if insurance_records:
        for record in insurance_records:
            insurance_data.append([
                record.insurance_year.name if record.insurance_year else 'N/A',
                f'{record.required_amount:,.2f}',
                f'{record.amount_paid:,.2f}',
                f'{record.balance:,.2f}',
                record.get_coverage_status_display(),
                record.remarks or 'N/A',
            ])
    else:
        insurance_data.append(['No family insurance records found', '', '', '', '', ''])
    elements.append(Paragraph('Family Insurance Records', section_style))
    elements.append(_build_pdf_table(insurance_data, [80, 65, 65, 65, 90, 155], body_font_size=7))

    doc.build(elements, canvasmaker=ExportNumberedCanvas)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    filename = f"{student.full_name.replace(' ', '_').lower()}_full_report.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response


@login_required
@require_POST
def student_approve(request, pk):
    """Approve a student (Executive Secretary or Admin)."""
    if not _can_approve_student(request.user):
        raise Http404()
    student = get_object_or_404(Student, pk=pk)
    student.sponsorship_status = 'active'
    student.save(update_fields=['sponsorship_status'])
    _notify_admins_and_exec(
        request=request,
        actor=request.user,
        verb=f"Approved student {student.full_name}",
        link=reverse('students:student_detail', kwargs={'pk': student.pk})
    )
    messages.success(request, f'Student {student.full_name} approved successfully!')
    return redirect('students:student_detail', pk=student.pk)


@login_required
@permission_required('students.view_student', raise_exception=True)
def photo_gallery(request):
    """View all student photos in a gallery."""
    district_filter = request.GET.get('district')
    school_filter = request.GET.get('school')
    level_filter = request.GET.get('level')

    photos = StudentPhoto.objects.select_related('student', 'student__family__district', 'student__school').all()

    if district_filter:
        photos = photos.filter(student__family__district_id=district_filter)
    if school_filter:
        photos = photos.filter(student__school_id=school_filter)
    if level_filter:
        photos = photos.filter(student__school_level=level_filter)

    photos = photos.order_by('-created_at')

    context = {
        'photos': photos,
        'districts': District.objects.order_by('name'),
        'schools': School.objects.order_by('name'),
        'levels': Student.SCHOOL_LEVELS,
        'selected_district': int(district_filter) if district_filter else None,
        'selected_school': int(school_filter) if school_filter else None,
        'selected_level': level_filter,
    }
    return render(request, 'students/photo_gallery.html', context)


@login_required
def add_photo(request, pk):
    """Add photo to student."""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        form = StudentPhotoForm(request.POST, request.FILES)
        if form.is_valid():
            photo = form.save(commit=False)
            photo.student = student
            photo.save()
            messages.success(request, 'Photo added successfully!')
            return redirect('students:student_photos', pk=student.pk)
    else:
        form = StudentPhotoForm()
    
    return render(request, 'students/add_photo.html', {
        'form': form,
        'student': student,
        'page_title': 'Add Student Photo',
        'submit_label': 'Upload Photo',
    })


@login_required
@permission_required('students.change_studentphoto', raise_exception=True)
def edit_photo(request, student_pk, photo_pk):
    """Edit an existing student photo."""
    student = get_object_or_404(Student, pk=student_pk)
    photo = get_object_or_404(StudentPhoto, pk=photo_pk, student=student)

    if request.method == 'POST':
        form = StudentPhotoForm(request.POST, request.FILES, instance=photo)
        if form.is_valid():
            updated_photo = form.save(commit=False)
            updated_photo.student = student
            updated_photo.save()
            messages.success(request, 'Photo updated successfully!')
            return redirect('students:student_photos', pk=student.pk)
    else:
        form = StudentPhotoForm(instance=photo)

    return render(request, 'students/add_photo.html', {
        'form': form,
        'student': student,
        'photo': photo,
        'page_title': 'Edit Student Photo',
        'submit_label': 'Update Photo',
    })


@login_required
@permission_required('students.delete_studentphoto', raise_exception=True)
def delete_photo(request, student_pk, photo_pk):
    """Delete a student photo."""
    student = get_object_or_404(Student, pk=student_pk)
    photo = get_object_or_404(StudentPhoto, pk=photo_pk, student=student)

    if request.method == 'POST':
        photo.delete()
        messages.success(request, 'Photo deleted successfully!')
        return redirect('students:student_photos', pk=student.pk)

    return render(request, 'students/delete_photo.html', {
        'student': student,
        'photo': photo,
    })


@login_required
@permission_required('students.view_student', raise_exception=True)
def student_photos(request, pk):
    """View photos for a specific student."""
    student = get_object_or_404(Student, pk=pk)
    photos = student.photos.all().order_by('-created_at')
    token = signing.dumps({'student_id': student.pk}, salt='student-photos')
    share_url = request.build_absolute_uri(
        reverse('students:student_photos_public', kwargs={'token': token})
    )
    return render(request, 'students/student_photos.html', {
        'student': student,
        'photos': photos,
        'share_url': share_url,
        'photo_count': photos.count(),
    })


def student_photos_public(request, token):
    """Public view for student photos via signed link."""
    try:
        data = signing.loads(token, salt='student-photos', max_age=60 * 60 * 24 * 7)
    except signing.SignatureExpired as exc:
        raise Http404('Share link expired.') from exc
    except signing.BadSignature as exc:
        raise Http404('Invalid share link.') from exc

    student_id = data.get('student_id')
    student = get_object_or_404(Student, pk=student_id)
    photos = student.photos.all().order_by('-created_at')
    return render(request, 'students/student_photos_public.html', {'student': student, 'photos': photos})

@login_required
def add_academic_record(request, pk):
    """Add academic record to student."""
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        form = StudentMarkForm(request.POST, request.FILES)
        if form.is_valid():
            record = form.save(commit=False)
            record.student = student
            record.save()
            messages.success(request, 'Academic record added successfully!')
            return redirect('students:student_detail', pk=student.pk)
    else:
        form = StudentMarkForm()
    
    return render(request, 'students/add_academic_record.html', {
        'form': form,
        'student': student,
        'page_title': 'Add Academic Record',
        'submit_label': 'Save Academic Record',
    })


@login_required
@permission_required('students.change_studentmark', raise_exception=True)
def edit_academic_record(request, student_pk, record_pk):
    """Edit an academic record linked to a student."""
    student = get_object_or_404(Student, pk=student_pk)
    record = get_object_or_404(StudentMark, pk=record_pk, student=student)

    if request.method == 'POST':
        form = StudentMarkForm(request.POST, request.FILES, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, 'Academic record updated successfully!')
            return redirect('students:student_detail', pk=student.pk)
    else:
        form = StudentMarkForm(instance=record)

    return render(request, 'students/add_academic_record.html', {
        'form': form,
        'student': student,
        'record': record,
        'page_title': 'Edit Academic Record',
        'submit_label': 'Update Academic Record',
    })


@login_required
@permission_required('students.delete_studentmark', raise_exception=True)
def delete_academic_record(request, student_pk, record_pk):
    """Delete an academic record linked to a student."""
    student = get_object_or_404(Student, pk=student_pk)
    record = get_object_or_404(StudentMark, pk=record_pk, student=student)

    if request.method == 'POST':
        record.delete()
        messages.success(request, 'Academic record deleted successfully!')
        return redirect('students:student_detail', pk=student.pk)

    return render(request, 'students/delete_academic_record.html', {
        'student': student,
        'record': record,
    })


@login_required
@permission_required('students.view_student', raise_exception=True)
def student_report_cards(request, pk):
    """View report card images for a specific student."""
    student = get_object_or_404(Student, pk=pk)
    records = student.academic_records.exclude(report_card_image='').exclude(report_card_image__isnull=True)
    return render(request, 'students/student_report_cards.html', {'student': student, 'records': records})


@login_required
@permission_required('students.add_studentmark', raise_exception=True)
def student_performance_bulk_entry(request):
    """Bulk capture student marks filtered by district, year, and term."""

    data_source = request.POST if request.method == 'POST' else request.GET
    filter_form = BulkPerformanceFilterForm(data_source or None)
    StudentBulkFormSet = formset_factory(BulkStudentMarkForm, extra=0)

    formset = None
    students_loaded = False
    table_rows = []
    selected_filters = {}
    selected_district = None
    selected_year = None
    term_display = None
    subject_label = None
    category_label = None
    class_label = None
    summary = {
        'student_count': 0,
        'records_existing': 0,
        'pending_records': 0,
        'average_marks': 0,
    }

    if filter_form.is_valid():
        students_loaded = True
        academic_year = filter_form.cleaned_data['academic_year']
        district = filter_form.cleaned_data['district']
        partner = filter_form.cleaned_data.get('partner')
        term = filter_form.cleaned_data['term']
        term_slug = term.lower().replace(' ', '')
        subject_value = f"{academic_year.name}-{term_slug}-marks"
        category = filter_form.cleaned_data['category']
        class_level = filter_form.cleaned_data.get('class_level', '').strip()

        selected_district = district
        selected_partner = partner
        selected_year = academic_year
        term_display = term
        subject_label = subject_value
        class_label = class_level or None
        category_label = dict(filter_form.fields['category'].choices).get(category, category)

        student_qs = (
            Student.objects.filter(is_active=True)
            .select_related('school', 'partner', 'family__district')
            .filter(
                Q(family__district=district) |
                Q(partner__district=district) |
                Q(school__district=district)
            )
            .order_by('first_name', 'last_name')
            .distinct()
        )
        if partner:
            student_qs = student_qs.filter(partner=partner)
        if category == 'primary':
            student_qs = student_qs.filter(school_level='primary')
        elif category == 'secondary':
            student_qs = student_qs.filter(school_level='secondary')
        if class_level:
            student_qs = student_qs.filter(class_level__iexact=class_level)

        students = list(student_qs)
        student_lookup = {student.id: student for student in students}

        existing_marks_qs = StudentMark.objects.filter(
            student__in=students,
            academic_year=academic_year,
            term=term,
        )
        existing_map = {}
        for mark in existing_marks_qs:
            current = existing_map.get(mark.student_id)
            if not current:
                existing_map[mark.student_id] = mark
                continue
            if current.subject != subject_value and mark.subject == subject_value:
                existing_map[mark.student_id] = mark
        existing_values = [float(mark.marks) for mark in existing_map.values() if mark.marks is not None]
        existing_count = len(existing_map)
        summary = {
            'student_count': len(students),
            'records_existing': existing_count,
            'pending_records': max(len(students) - existing_count, 0),
            'average_marks': round(sum(existing_values) / len(existing_values), 1) if existing_values else 0,
        }

        initial_data = []
        for student in students:
            mark = existing_map.get(student.id)
            initial_data.append({
                'student_id': student.id,
                'marks': mark.marks if mark else None,
                'teacher_remark': mark.teacher_remark if mark else '',
            })

        if request.method == 'POST':
            formset = StudentBulkFormSet(request.POST)
            if formset.is_valid():
                created_count = 0
                updated_count = 0
                with transaction.atomic():
                    for form in formset:
                        student_id = form.cleaned_data.get('student_id')
                        marks = form.cleaned_data.get('marks')
                        remark = form.cleaned_data.get('teacher_remark', '')

                        if not student_id or student_id not in student_lookup:
                            continue
                        if marks in (None, ''):
                            continue

                        existing_mark = existing_map.get(student_id)
                        if existing_mark:
                            existing_mark.subject = subject_value
                            existing_mark.marks = marks
                            existing_mark.teacher_remark = remark or ''
                            existing_mark.academic_year = academic_year
                            existing_mark.term = term
                            existing_mark.save(update_fields=['subject', 'marks', 'teacher_remark', 'academic_year', 'term', 'updated_at'])
                            updated_count += 1
                        else:
                            StudentMark.objects.create(
                                student=student_lookup[student_id],
                                academic_year=academic_year,
                                term=term,
                                subject=subject_value,
                                marks=marks,
                                teacher_remark=remark or '',
                            )
                            created_count += 1

                messages.success(
                    request,
                    f"Bulk marks saved for {district.name}{f' - {partner.name}' if partner else ''}: {created_count} new, {updated_count} updated."
                )
                params = {
                    'academic_year': academic_year.id,
                    'district': district.id,
                    'term': term,
                    'category': category,
                }
                if partner:
                    params['partner'] = partner.id
                if class_level:
                    params['class_level'] = class_level
                query = urlencode(params)
                return redirect(f"{reverse('students:student_performance_bulk_entry')}?{query}")
        else:
            formset = StudentBulkFormSet(initial=initial_data)

        if formset is not None:
            for idx, student in enumerate(students):
                if idx >= len(formset.forms):
                    break
                table_rows.append({
                    'student': student,
                    'form': formset.forms[idx],
                    'existing_mark': existing_map.get(student.id),
                })

        selected_filters = {
            'academic_year': academic_year.id,
            'district': district.id,
            'term': term,
            'category': category,
        }
        if partner:
            selected_filters['partner'] = partner.id
        if class_level:
            selected_filters['class_level'] = class_level

    context = {
        'filter_form': filter_form,
        'formset': formset,
        'table_rows': table_rows,
        'students_loaded': students_loaded,
        'selected_filters': selected_filters,
        'selected_district': selected_district,
        'selected_partner': selected_partner if 'selected_partner' in locals() else None,
        'selected_year': selected_year,
        'term_display': term_display,
        'subject_label': subject_label,
        'category_label': category_label,
        'class_label': class_label,
        'summary': summary,
    }
    return render(request, 'students/student_performance_bulk_entry.html', context)

class StudentPerformanceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Paginated student performance dashboard."""

    template_name = 'students/student_performance.html'
    context_object_name = 'performance_rows'
    paginate_by = 10
    permission_required = 'students.view_studentmark'
    raise_exception = True

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format in {'pdf', 'excel'}:
            self.object_list = self.get_queryset()
            return self.export_data(export_format)
        return super().get(request, *args, **kwargs)

    def get_filters(self):
        status = self.request.GET.get('status', 'all').lower()
        if status not in {'passed', 'failed'}:
            status = 'all'
        return {
            'academic_year': self.request.GET.get('academic_year', '').strip(),
            'term': self.request.GET.get('term', '').strip(),
            'class_level': self.request.GET.get('class_level', '').strip(),
            'search': self.request.GET.get('search', '').strip(),
            'status': status,
        }

    def get_queryset(self):
        if hasattr(self, '_performance_qs'):
            return self._performance_qs

        filters = self.get_filters()
        queryset = Student.objects.select_related('school').filter(academic_records__isnull=False)
        mark_filter = Q()

        if filters['class_level']:
            queryset = queryset.filter(class_level=filters['class_level'])

        if filters['search']:
            for term in filters['search'].split():
                queryset = queryset.filter(
                    Q(first_name__icontains=term) | Q(last_name__icontains=term)
                )

        if filters['academic_year']:
            queryset = queryset.filter(academic_records__academic_year_id=filters['academic_year'])
            mark_filter &= Q(academic_records__academic_year_id=filters['academic_year'])

        if filters['term']:
            queryset = queryset.filter(academic_records__term=filters['term'])
            mark_filter &= Q(academic_records__term=filters['term'])

        mark_filter = mark_filter if mark_filter.children else Q()

        queryset = queryset.annotate(
            avg_marks=Avg('academic_records__marks', filter=mark_filter),
            total_marks=Coalesce(
                Sum('academic_records__marks', filter=mark_filter),
                Value(0, output_field=DecimalField(max_digits=10, decimal_places=2))
            ),
            records_count=Count('academic_records', filter=mark_filter, distinct=True),
        )

        queryset = queryset.annotate(
            has_passed=Case(
                When(avg_marks__gte=Value(50), then=Value(True)),
                default=Value(False),
                output_field=BooleanField(),
            )
        )

        if filters['status'] == 'passed':
            queryset = queryset.filter(has_passed=True)
        elif filters['status'] == 'failed':
            queryset = queryset.filter(has_passed=False)

        queryset = queryset.order_by('first_name', 'last_name').distinct()

        self._performance_qs = queryset
        return queryset

    def get_summary_stats(self):
        if hasattr(self, '_summary_stats'):
            return self._summary_stats

        queryset = self.get_queryset()
        total_students = queryset.count()
        passed = queryset.filter(has_passed=True).count()
        failed = max(total_students - passed, 0)
        pass_rate = round((passed / total_students) * 100, 1) if total_students else 0

        self._summary_stats = {
            'total_students': total_students,
            'passed': passed,
            'failed': failed,
            'pass_rate': pass_rate,
        }
        return self._summary_stats

    def get_trend_data(self):
        filters = self.get_filters()
        trend_qs = StudentMark.objects.all()
        if filters['academic_year']:
            trend_qs = trend_qs.filter(academic_year_id=filters['academic_year'])
        if filters['class_level']:
            trend_qs = trend_qs.filter(student__class_level=filters['class_level'])
        if filters['search']:
            for term in filters['search'].split():
                trend_qs = trend_qs.filter(
                    Q(student__first_name__icontains=term) | Q(student__last_name__icontains=term)
                )

        trend_map = {
            row['term']: round(float(row['avg_marks'] or 0), 1)
            for row in trend_qs.values('term').annotate(avg_marks=Avg('marks'))
        }
        labels = [choice[0] for choice in StudentMark.TERM_CHOICES]
        values = [trend_map.get(label, 0) for label in labels]
        return labels, values

    def build_url(self, **overrides):
        params = self.request.GET.copy()
        params.pop('page', None)
        params.pop('export', None)
        for key, value in overrides.items():
            if key == 'status' and value == 'all':
                params.pop('status', None)
                continue
            if value in (None, ''):
                params.pop(key, None)
            else:
                params[key] = value
        query = params.urlencode()
        return f"{self.request.path}?{query}" if query else self.request.path

    def get_preserved_querystring(self):
        params = self.request.GET.copy()
        params.pop('page', None)
        params.pop('export', None)
        return params.urlencode()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        filters = self.get_filters()
        summary = self.get_summary_stats()
        trend_labels, trend_values = self.get_trend_data()

        available_classes = (
            StudentMark.objects.values_list('student__class_level', flat=True)
            .exclude(student__class_level__isnull=True)
            .exclude(student__class_level__exact='')
            .distinct()
            .order_by('student__class_level')
        )

        status_options = [
            {'label': 'All Students', 'value': 'all'},
            {'label': 'Passed Students', 'value': 'passed'},
            {'label': 'Failed Students', 'value': 'failed'},
        ]
        for option in status_options:
            option['url'] = self.build_url(status=option['value'])
            option['active'] = option['value'] == filters['status']

        filters_applied = any(
            value for key, value in filters.items() if key != 'status'
        ) or filters['status'] in {'passed', 'failed'}

        context.update({
            'available_years': AcademicYear.objects.order_by('-name'),
            'available_terms': [choice[0] for choice in StudentMark.TERM_CHOICES],
            'available_classes': available_classes,
            'selected_year': filters['academic_year'],
            'selected_term': filters['term'],
            'selected_class': filters['class_level'],
            'search_query': filters['search'],
            'status_filter': filters['status'],
            'status_filter_options': status_options,
            'trend_labels': trend_labels,
            'trend_values': trend_values,
            'querystring': self.get_preserved_querystring(),
            'filters_applied': filters_applied,
            'export_pdf_url': self.build_url(export='pdf'),
            'export_excel_url': self.build_url(export='excel'),
        })
        context.update(summary)
        return context

    def get_export_rows(self):
        rows = []
        for student in self.get_queryset():
            avg = student.avg_marks or 0
            rows.append({
                'name': student.full_name,
                'class_level': student.class_level or 'N/A',
                'terms_count': student.records_count or 0,
                'avg_marks': float(avg),
                'status': 'Pass' if student.has_passed else 'Fail',
            })
        return rows

    def describe_filters(self):
        filters = self.get_filters()
        year_label = 'All Years'
        if filters['academic_year']:
            year = AcademicYear.objects.filter(id=filters['academic_year']).values_list('name', flat=True).first()
            year_label = year or 'Selected Year'

        term_label = filters['term'] or 'All Terms'
        class_label = filters['class_level'] or 'All Classes'
        search_label = filters['search'] or 'All Students'
        status_map = {
            'all': 'All Students',
            'passed': 'Passed Students',
            'failed': 'Failed Students',
        }
        status_label = status_map.get(filters['status'], 'All Students')
        return f"Year: {year_label} | Term: {term_label} | Class: {class_label} | Search: {search_label} | Status: {status_label}"

    def export_data(self, export_format):
        rows = self.get_export_rows()
        if export_format == 'pdf':
            return self.generate_pdf(rows)
        return self.generate_excel(rows)

    def generate_pdf(self, rows):
        buffer = BytesIO()
        doc = build_export_pdf_document(
            buffer,
            'Student Performance Report',
            pagesize=landscape(A4),
            left_margin=24,
            right_margin=24,
            top_margin=24,
            bottom_margin=24,
        )
        elements = []
        add_export_header(
            elements,
            'Student Performance Report',
            self.describe_filters(),
            generated_label=f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}",
        )

        rows_data = []
        for row in rows:
            rows_data.append([
                row['name'],
                row['class_level'],
                str(row['terms_count']),
                f"{row['avg_marks']:.1f}",
                row['status'],
            ])
        table_data = prepend_row_numbers(
            ['Student Name', 'Class', 'Total Terms Recorded', 'Average (%)', 'Status'],
            rows_data,
        )
        table = build_export_table(
            table_data,
            col_widths=[36, 190, 110, 110, 95, 85],
            body_font_size=8,
            header_background=colors.HexColor('#0f172a'),
            centered_columns=[0, 3, 4],
        )

        for idx, row in enumerate(rows, start=1):
            bg_color = colors.HexColor('#d1fae5') if row['status'] == 'Pass' else colors.HexColor('#fee2e2')
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, idx), (-1, idx), bg_color),
            ]))

        elements.append(table)
        doc.build(elements, canvasmaker=ExportNumberedCanvas)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        filename = timezone.now().strftime('student_performance_%Y%m%d_%H%M%S.pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        return response

    def generate_excel(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Performance'

        headers = ['No.', 'Student Name', 'Class', 'Total Terms Recorded', 'Average (%)', 'Status']
        header_row_idx = write_excel_report_header(
            worksheet,
            'Student Performance Report',
            f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M')} | {self.describe_filters()}",
            len(headers),
        )
        worksheet.append(headers)
        style_excel_header(worksheet, header_row_idx, fill_color='0F172A')

        data_start_row = header_row_idx + 1
        for index, record in enumerate(rows, start=1):
            worksheet.append([
                index,
                record['name'],
                record['class_level'],
                record['terms_count'],
                round(record['avg_marks'], 1),
                record['status'],
            ])
            status_cell = worksheet.cell(row=worksheet.max_row, column=6)
            fill_color = 'D1FAE5' if record['status'] == 'Pass' else 'FEE2E2'
            status_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        column_widths = [8, 32, 18, 18, 18, 14]
        for idx, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[chr(64 + idx)].width = width
        style_excel_table_rows(
            worksheet,
            header_row_idx=header_row_idx,
            data_start_row=data_start_row,
            data_end_row=worksheet.max_row,
            max_col=len(headers),
            centered_columns=[1, 4, 5, 6],
        )
        for row_index, record in enumerate(rows, start=data_start_row):
            status_cell = worksheet.cell(row=row_index, column=6)
            fill_color = 'D1FAE5' if record['status'] == 'Pass' else 'FEE2E2'
            status_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = timezone.now().strftime('student_performance_%Y%m%d_%H%M%S.xlsx')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.read())
        buffer.close()
        return response


class StudentPerformanceDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Detailed performance history for a single student."""

    model = Student
    template_name = 'students/student_performance_detail.html'
    context_object_name = 'student'
    permission_required = 'students.view_studentmark'
    raise_exception = True

    def get_queryset(self):
        return (
            Student.objects.select_related('school')
            .prefetch_related(
                Prefetch(
                    'academic_records',
                    queryset=StudentMark.objects.select_related('academic_year')
                    .order_by('-academic_year__name', '-term', 'subject')
                )
            )
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        records = self.object.academic_records.all()
        stats = records.aggregate(
            total_marks=Coalesce(
                Sum('marks'),
                Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)),
            ),
            avg_marks=Avg('marks'),
            subjects_count=Count('id'),
        )
        avg_marks = stats.get('avg_marks') or 0
        context.update({
            'performance_history': records,
            'total_marks': stats.get('total_marks') or 0,
            'average_marks': avg_marks,
            'subjects_count': stats.get('subjects_count') or 0,
            'has_passed': avg_marks >= 50,
        })
        return context


@login_required
@permission_required('students.view_student', raise_exception=True)
def photo_gallery(request):
    """View student photos grouped into student albums."""
    search_query = request.GET.get('search', '').strip()
    district_id = request.GET.get('district')
    school_id = request.GET.get('school')
    level = request.GET.get('level')
    class_level = request.GET.get('class_level', '').strip()

    students = (
        Student.objects.select_related('family__district', 'school')
        .prefetch_related(
            Prefetch(
                'photos',
                queryset=StudentPhoto.objects.order_by('-created_at'),
                to_attr='album_photos',
            )
        )
        .annotate(photo_count=Count('photos'))
        .filter(photo_count__gt=0)
    )

    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(class_level__icontains=search_query) |
            Q(school__name__icontains=search_query)
        )
    if district_id:
        students = students.filter(
            Q(family__district_id=district_id) | Q(school__district_id=district_id)
        )
    if school_id:
        students = students.filter(school_id=school_id)
    if level:
        students = students.filter(school_level=level)
    if class_level:
        students = students.filter(class_level__icontains=class_level)

    students = students.order_by('first_name', 'last_name')

    context = {
        'students': students,
        'districts': District.objects.order_by('name'),
        'schools': School.objects.order_by('name'),
        'levels': Student.SCHOOL_LEVEL_CHOICES,
        'selected_district': int(district_id) if district_id and district_id.isdigit() else None,
        'selected_school': int(school_id) if school_id and school_id.isdigit() else None,
        'selected_level': level,
        'search_query': search_query,
        'class_level_filter': class_level,
        'album_count': students.count(),
        'photo_total': sum(student.photo_count for student in students),
    }
    return render(request, 'students/photo_gallery.html', context)
