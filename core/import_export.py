"""
Excel Import/Export functionality for Students, Families, and Schools
"""
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re
from io import BytesIO

from students.models import Student
from families.models import Family
from core.models import School, Province, District, Sector, Cell, Village


# ========== TEMPLATE GENERATION ==========

def _normalize_row(row, expected_len):
    row_values = list(row or [])
    if len(row_values) < expected_len:
        row_values.extend([None] * (expected_len - len(row_values)))
    return row_values


def _normalize_header(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = text.replace('*', '')
    text = re.sub(r'\([^)]*\)', '', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _build_header_index(header_row):
    header_index = {}
    for idx, cell in enumerate(header_row or []):
        key = _normalize_header(cell)
        if key:
            header_index[key] = idx
    return header_index

@login_required
def download_student_template(request):
    """Generate and download Excel template for student import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students Template"
    
    # Header styling
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")  # emerald-500
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Define headers
    headers = [
        'First Name*', 'Last Name*', 'Gender (M/F)*', 'Date of Birth (YYYY-MM-DD)*',
        'Family Code', 'School Name*', 'Class Level*',
        'Enrollment Status', 'Sponsorship Status',
        'Has Disability (Yes/No)', 'Disability Types', 'Disability Description',
        'Province', 'District', 'Sector', 'Cell', 'Village',
        'Program Officer Name', 'National ID'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add example data in row 2
    example_data = [
        'John', 'Doe', 'M', '2010-05-15',
        'FAM-2024-ABC123', 'Example Primary School', 'P4',
        'enrolled', 'active',
        'No', '', '',
        'Kigali City', 'Gasabo', 'Remera', 'Rukiri I', 'Amahoro',
        'Jane Smith', '1234567890123456'
    ]
    
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.font = Font(italic=True, color="666666")
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ("Excel Import Template - Students", "Instructions for importing student data"),
        ("", ""),
        ("Required Fields (marked with *):", ""),
        ("- First Name", "Student's first name"),
        ("- Last Name", "Student's last name"),
        ("- Gender", "Must be 'M' for Male or 'F' for Female"),
        ("- Date of Birth", "Format: YYYY-MM-DD (e.g., 2010-05-15)"),
        ("- School Name", "Name of the school"),
        ("- Class Level", "e.g., P1, P2, S1, S2, etc."),
        ("", ""),
        ("Optional Fields:", ""),
        ("- Family Code", "Leave blank to auto-assign or provide existing family code"),
        ("- Enrollment Status", "Options: enrolled, transferred, graduated, dropped_out (default: enrolled)"),
        ("- Sponsorship Status", "Options: active, pending, graduated (default: pending)"),
        ("- Has Disability", "Yes or No (default: No)"),
        ("- Disability Types", "Comma-separated: visual, hearing, mobility, intellectual, autism, speech, learning, emotional, other"),
        ("- Province/District/Sector/Cell/Village", "Rwanda location hierarchy"),
        ("- Program Officer Name", "Name of assigned program officer"),
        ("- National ID", "Student's national ID if available"),
        ("", ""),
        ("Tips:", ""),
        ("- Row 2 contains example data (delete before importing)", ""),
        ("- Import will skip rows with errors and show summary", ""),
        ("- Duplicate checks: National ID (if provided)", ""),
    ]
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        ws_instructions.cell(row=row_num, column=1, value=col1).font = Font(bold=True if col1 and not col1.startswith('-') else False)
        ws_instructions.cell(row=row_num, column=2, value=col2)
    
    # Auto-adjust column widths
    for sheet in [ws, ws_instructions]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Student_Import_Template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def download_family_template(request):
    """Generate and download Excel template for family import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Families Template"
    
    # Header styling
    header_fill = PatternFill(start_color="14B8A6", end_color="14B8A6", fill_type="solid")  # teal-500
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Define headers
    headers = [
        'Head of Family Name*', 'National ID*', 'Phone Number*', 'Alternative Phone',
        'Province*', 'District*', 'Sector', 'Cell', 'Village',
        'Total Family Members*', 'Address Description', 'Notes'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add example data
    example_data = [
        'Jean Pierre Mukasa', '1198012345678901', '+250788123456', '+250722654321',
        'Kigali City', 'Gasabo', 'Remera', 'Rukiri I', 'Amahoro',
        '5', 'Near the main road, blue gate', 'Family needs support'
    ]
    
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.font = Font(italic=True, color="666666")
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ("Excel Import Template - Families", "Instructions for importing family data"),
        ("", ""),
        ("Required Fields (marked with *):", ""),
        ("- Head of Family Name", "Full name of family head"),
        ("- National ID", "16-digit National ID (must be unique)"),
        ("- Phone Number", "Primary contact number"),
        ("- Province", "Rwanda province name"),
        ("- District", "Rwanda district name"),
        ("- Total Family Members", "Number of people in the family (minimum 1)"),
        ("", ""),
        ("Optional Fields:", ""),
        ("- Alternative Phone", "Secondary contact number"),
        ("- Sector/Cell/Village", "Additional location details"),
        ("- Address Description", "Detailed address or landmarks"),
        ("- Notes", "Additional comments"),
        ("", ""),
        ("Important Notes:", ""),
        ("- Family Code will be auto-generated (FAM-YYYY-XXXX format)", ""),
        ("- National ID must be unique in the system", ""),
        ("- Location names must match existing database entries", ""),
        ("- Row 2 contains example data (delete before importing)", ""),
    ]
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        ws_instructions.cell(row=row_num, column=1, value=col1).font = Font(bold=True if col1 and not col1.startswith('-') else False)
        ws_instructions.cell(row=row_num, column=2, value=col2)
    
    # Auto-adjust column widths
    for sheet in [ws, ws_instructions]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Family_Import_Template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def download_school_template(request):
    """Generate and download Excel template for school import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Schools Template"
    
    # Header styling
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")  # emerald-600
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Define headers
    headers = [
        'School Name*', 'Province', 'District*', 'Sector',
        'Headteacher Name', 'Headteacher Mobile', 'Headteacher Email',
        'Bank Name', 'Bank Account Name', 'Bank Account Number',
        'Fee Amount (RWF)'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add example data
    example_data = [
        'GS Kigali', 'Kigali City', 'Gasabo', 'Remera',
        'Dr. Marie Uwase', '+250788999888', 'headteacher@gskigali.ac.rw',
        'Bank of Kigali', 'GS Kigali School Account', '4000-1234567-89',
        '150000'
    ]
    
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.font = Font(italic=True, color="666666")
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ("Excel Import Template - Schools", "Instructions for importing school data"),
        ("", ""),
        ("Required Fields (marked with *):", ""),
        ("- School Name", "Full name of the school (must be unique)"),
        ("- District", "Rwanda district name"),
        ("", ""),
        ("Optional Fields:", ""),
        ("- Province", "Rwanda province name"),
        ("- Sector", "Rwanda sector name"),
        ("- Headteacher Name", "Full name of school headteacher"),
        ("- Headteacher Mobile", "Contact phone number"),
        ("- Headteacher Email", "Email address"),
        ("- Bank Name", "Name of the bank"),
        ("- Bank Account Name", "Account holder name"),
        ("- Bank Account Number", "Bank account number"),
        ("- Fee Amount (RWF)", "Standard fee per student (numbers only)"),
        ("", ""),
        ("Important Notes:", ""),
        ("- School names must be unique in the system", ""),
        ("- Location names must match existing database entries", ""),
        ("- Fee Amount should be numbers only (e.g., 150000 not 150,000)", ""),
        ("- Row 2 contains example data (delete before importing)", ""),
    ]
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        ws_instructions.cell(row=row_num, column=1, value=col1).font = Font(bold=True if col1 and not col1.startswith('-') else False)
        ws_instructions.cell(row=row_num, column=2, value=col2)
    
    # Auto-adjust column widths
    for sheet in [ws, ws_instructions]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=School_Import_Template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


# ========== IMPORT VIEWS ==========

@login_required
def import_students(request):
    """Handle student Excel file import."""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        expected_columns = 19
        expected_headers = [
            ('first name', True),
            ('last name', True),
            ('gender', True),
            ('date of birth', True),
            ('family code', False),
            ('school name', True),
            ('class level', True),
            ('enrollment status', False),
            ('sponsorship status', False),
            ('has disability', False),
            ('disability types', False),
            ('disability description', False),
            ('province', False),
            ('district', False),
            ('sector', False),
            ('cell', False),
            ('village', False),
            ('program officer name', False),
            ('national id', False),
        ]
        
        try:
            # Load workbook
            wb = load_workbook(excel_file)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_index = _build_header_index(header_row)
            missing_required = [name for name, required in expected_headers if required and name not in header_index]
            if missing_required:
                messages.error(
                    request,
                    "Missing required columns: "
                    + ", ".join(missing_required)
                    + ". Please download the latest template and try again."
                )
                return redirect('students:student_list')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Skip header row (row 1) and example row (row 2 if it's example)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not any(row):
                    continue
                
                try:
                    row = _normalize_row(row, expected_columns)
                    get_value = lambda key: row[header_index.get(key, -1)] if header_index.get(key, -1) >= 0 and header_index.get(key, -1) < len(row) else None
                    # Extract data
                    first_name = get_value('first name')
                    last_name = get_value('last name')
                    gender = get_value('gender')
                    dob_str = get_value('date of birth')
                    family_code = get_value('family code')
                    school_name = get_value('school name')
                    class_level = get_value('class level')
                    enrollment_status = get_value('enrollment status') or 'enrolled'
                    sponsorship_status = get_value('sponsorship status') or 'pending'
                    has_disability_str = get_value('has disability')
                    disability_types = get_value('disability types')
                    disability_description = get_value('disability description')
                    province_name = get_value('province')
                    district_name = get_value('district')
                    sector_name = get_value('sector')
                    cell_name = get_value('cell')
                    village_name = get_value('village')
                    program_officer = get_value('program officer name')
                    national_id = get_value('national id')
                    
                    # Validate required fields
                    if not all([first_name, last_name, gender, dob_str, school_name, class_level]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        error_count += 1
                        continue
                    
                    # Parse date
                    if isinstance(dob_str, datetime):
                        date_of_birth = dob_str.date()
                    else:
                        date_of_birth = datetime.strptime(str(dob_str), '%Y-%m-%d').date()
                    
                    # Get or create family
                    family = None
                    if family_code:
                        try:
                            family = Family.objects.get(family_code=family_code)
                        except Family.DoesNotExist:
                            errors.append(f"Row {row_num}: Family code '{family_code}' not found")
                    
                    # Get school
                    school = School.objects.filter(name=school_name).first()
                    if not school:
                        errors.append(f"Row {row_num}: School '{school_name}' not found (student will be created without school link)")
                    
                    # Get locations
                    province = Province.objects.filter(name=province_name).first() if province_name else None
                    district = District.objects.filter(name=district_name).first() if district_name else None
                    sector = Sector.objects.filter(name=sector_name).first() if sector_name else None
                    cell = Cell.objects.filter(name=cell_name).first() if cell_name else None
                    village = Village.objects.filter(name=village_name).first() if village_name else None
                    
                    # Parse disability
                    has_disability = has_disability_str and str(has_disability_str).lower() in ['yes', 'true', '1']
                    
                    # Create student
                    student = Student.objects.create(
                        family=family,
                        first_name=first_name,
                        last_name=last_name,
                        gender=gender.upper()[0],  # M or F
                        date_of_birth=date_of_birth,
                        school_name=school_name,
                        school=school,
                        class_level=class_level,
                        enrollment_status=enrollment_status,
                        sponsorship_status=sponsorship_status,
                        has_disability=has_disability,
                        disability_types=disability_types or '',
                        disability_description=disability_description or '',
                        province=province,
                        district=district,
                        sector=sector,
                        cell=cell,
                        village=village,
                        program_officer_name=program_officer or '',
                        national_id=national_id or ''
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
            
            # Show results
            if success_count > 0:
                messages.success(request, f"Successfully imported {success_count} student(s)!")
            if error_count > 0:
                messages.warning(request, f"Failed to import {error_count} row(s). See details below.")
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f"...and {len(errors) - 10} more errors")
            
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {str(e)}")
        
        return redirect('students:student_list')
    
    return render(request, 'core/import_form.html', {
        'title': 'Import Students',
        'download_url': 'core:download_student_template',
        'import_type': 'students'
    })


@login_required
def import_families(request):
    """Handle family Excel file import."""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        expected_columns = 12
        expected_headers = [
            ('head of family name', True),
            ('national id', True),
            ('phone number', True),
            ('alternative phone', False),
            ('province', True),
            ('district', True),
            ('sector', False),
            ('cell', False),
            ('village', False),
            ('total family members', True),
            ('address description', False),
            ('notes', False),
        ]
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_index = _build_header_index(header_row)
            missing_required = [name for name, required in expected_headers if required and name not in header_index]
            if missing_required:
                messages.error(
                    request,
                    "Missing required columns: "
                    + ", ".join(missing_required)
                    + ". Please download the latest template and try again."
                )
                return redirect('families:family_list')
            
            success_count = 0
            error_count = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                
                try:
                    row = _normalize_row(row, expected_columns)
                    get_value = lambda key: row[header_index.get(key, -1)] if header_index.get(key, -1) >= 0 and header_index.get(key, -1) < len(row) else None
                    head_of_family = get_value('head of family name')
                    national_id = get_value('national id')
                    phone_number = get_value('phone number')
                    alternative_phone = get_value('alternative phone')
                    province_name = get_value('province')
                    district_name = get_value('district')
                    sector_name = get_value('sector')
                    cell_name = get_value('cell')
                    village_name = get_value('village')
                    total_family_members = get_value('total family members')
                    address_description = get_value('address description')
                    notes = get_value('notes')
                    
                    # Validate required fields
                    if not all([head_of_family, national_id, phone_number, province_name, district_name, total_family_members]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        error_count += 1
                        continue
                    
                    # Check for duplicate National ID
                    if Family.objects.filter(national_id=national_id).exists():
                        errors.append(f"Row {row_num}: National ID '{national_id}' already exists")
                        error_count += 1
                        continue
                    
                    # Get locations
                    province = Province.objects.filter(name=province_name).first()
                    district = District.objects.filter(name=district_name).first()
                    sector = Sector.objects.filter(name=sector_name).first() if sector_name else None
                    cell = Cell.objects.filter(name=cell_name).first() if cell_name else None
                    village = Village.objects.filter(name=village_name).first() if village_name else None
                    
                    if not province:
                        errors.append(f"Row {row_num}: Province '{province_name}' not found")
                        error_count += 1
                        continue
                    if not district:
                        errors.append(f"Row {row_num}: District '{district_name}' not found")
                        error_count += 1
                        continue
                    
                    # Create family
                    family = Family.objects.create(
                        head_of_family=head_of_family,
                        national_id=national_id,
                        phone_number=phone_number,
                        alternative_phone=alternative_phone or '',
                        province=province,
                        district=district,
                        sector=sector,
                        cell=cell,
                        village=village,
                        total_family_members=int(total_family_members),
                        address_description=address_description or '',
                        notes=notes or ''
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f"Successfully imported {success_count} family/families!")
            if error_count > 0:
                messages.warning(request, f"Failed to import {error_count} row(s).")
                for error in errors[:10]:
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f"...and {len(errors) - 10} more errors")
            
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {str(e)}")
        
        return redirect('families:family_list')
    
    return render(request, 'core/import_form.html', {
        'title': 'Import Families',
        'download_url': 'core:download_family_template',
        'import_type': 'families'
    })


@login_required
def import_schools(request):
    """Handle school Excel file import."""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        expected_columns = 11
        expected_headers = [
            ('school name', True),
            ('province', False),
            ('district', True),
            ('sector', False),
            ('headteacher name', False),
            ('headteacher mobile', False),
            ('headteacher email', False),
            ('bank name', False),
            ('bank account name', False),
            ('bank account number', False),
            ('fee amount', False),
        ]
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_index = _build_header_index(header_row)
            missing_required = [name for name, required in expected_headers if required and name not in header_index]
            if missing_required:
                messages.error(
                    request,
                    "Missing required columns: "
                    + ", ".join(missing_required)
                    + ". Please download the latest template and try again."
                )
                return redirect('core:school_list')
            
            success_count = 0
            error_count = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                
                try:
                    row = _normalize_row(row, expected_columns)
                    get_value = lambda key: row[header_index.get(key, -1)] if header_index.get(key, -1) >= 0 and header_index.get(key, -1) < len(row) else None
                    school_name = get_value('school name')
                    province_name = get_value('province')
                    district_name = get_value('district')
                    sector_name = get_value('sector')
                    headteacher_name = get_value('headteacher name')
                    headteacher_mobile = get_value('headteacher mobile')
                    headteacher_email = get_value('headteacher email')
                    bank_name = get_value('bank name')
                    bank_account_name = get_value('bank account name')
                    bank_account_number = get_value('bank account number')
                    fee_amount = get_value('fee amount')
                    
                    # Validate required fields
                    if not all([school_name, district_name]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        error_count += 1
                        continue
                    
                    # Check for duplicate school name
                    if School.objects.filter(name=school_name).exists():
                        errors.append(f"Row {row_num}: School '{school_name}' already exists")
                        error_count += 1
                        continue
                    
                    # Get locations
                    province = Province.objects.filter(name=province_name).first() if province_name else None
                    district = District.objects.filter(name=district_name).first()
                    sector = Sector.objects.filter(name=sector_name).first() if sector_name else None
                    
                    if not district:
                        errors.append(f"Row {row_num}: District '{district_name}' not found")
                        error_count += 1
                        continue
                    
                    # Create school
                    school = School.objects.create(
                        name=school_name,
                        province=province,
                        district=district,
                        sector=sector,
                        headteacher_name=headteacher_name or '',
                        headteacher_mobile=headteacher_mobile or '',
                        headteacher_email=headteacher_email or '',
                        bank_name=bank_name or '',
                        bank_account_name=bank_account_name or '',
                        bank_account_number=bank_account_number or '',
                        fee_amount=float(fee_amount) if fee_amount else 0
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f"Successfully imported {success_count} school(s)!")
            if error_count > 0:
                messages.warning(request, f"Failed to import {error_count} row(s).")
                for error in errors[:10]:
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.error(request, f"...and {len(errors) - 10} more errors")
            
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {str(e)}")
        
        return redirect('core:school_list')
    
    return render(request, 'core/import_form.html', {
        'title': 'Import Schools',
        'download_url': 'core:download_school_template',
        'import_type': 'schools'
    })
