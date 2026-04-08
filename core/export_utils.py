from datetime import datetime
import os

from django.conf import settings
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


EXPORT_PRIMARY = colors.HexColor('#0f766e')
EXPORT_PRIMARY_DARK = colors.HexColor('#115e59')
EXPORT_TEXT_MUTED = colors.HexColor('#64748b')
EXPORT_ALT_ROW = colors.HexColor('#f8fafc')
EXPORT_BORDER = colors.HexColor('#cbd5e1')
EXPORT_TOTAL_BG = colors.HexColor('#d1fae5')
EXPORT_TOTAL_TEXT = colors.HexColor('#065f46')
EXPORT_FONT = 'Helvetica'
EXPORT_FONT_BOLD = 'Helvetica-Bold'


class ExportNumberedCanvas(canvas.Canvas):
    """Canvas that adds a consistent footer and page numbers."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        page_count = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(page_count)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        page_width, _page_height = self._pagesize
        self.setFont(EXPORT_FONT, 9)
        self.setFillColor(colors.grey)
        self.drawRightString(
            page_width - 0.5 * inch,
            0.5 * inch,
            f"Page {self._pageNumber} of {page_count}"
        )
        self.setStrokeColor(EXPORT_PRIMARY)
        self.setLineWidth(1)
        self.line(0.75 * inch, 0.65 * inch, page_width - 0.75 * inch, 0.65 * inch)


def resolve_logo_path():
    candidates = [
        os.path.join(settings.BASE_DIR, 'static', 'image', 'logo.jpg'),
        os.path.join(settings.BASE_DIR, 'static', 'image', 'logo.png'),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def build_export_pdf_document(
    target,
    title,
    pagesize=A4,
    left_margin=0.75 * inch,
    right_margin=0.75 * inch,
    top_margin=0.75 * inch,
    bottom_margin=1 * inch,
):
    return SimpleDocTemplate(
        target,
        pagesize=pagesize,
        title=title,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )


def add_export_header(elements, report_title, report_subtitle=None, generated_label=None):
    styles = getSampleStyleSheet()
    org_style = ParagraphStyle(
        'ExportOrg',
        parent=styles['Normal'],
        fontSize=11,
        textColor=EXPORT_TEXT_MUTED,
        alignment=TA_CENTER,
        spaceAfter=5,
    )
    title_style = ParagraphStyle(
        'ExportTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=EXPORT_PRIMARY,
        alignment=TA_CENTER,
        fontName=EXPORT_FONT_BOLD,
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        'ExportSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=EXPORT_TEXT_MUTED,
        alignment=TA_CENTER,
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        'ExportMeta',
        parent=styles['Normal'],
        fontSize=9,
        textColor=EXPORT_TEXT_MUTED,
        alignment=TA_RIGHT,
        spaceAfter=6,
    )

    logo_path = resolve_logo_path()
    if logo_path:
        logo = Image(logo_path, width=1.0 * inch, height=1.0 * inch)
        logo.hAlign = 'CENTER'
        elements.append(logo)
        elements.append(Spacer(1, 6))

    elements.append(Paragraph('Solidact Foundation', org_style))
    elements.append(Paragraph(report_title, title_style))
    if report_subtitle:
        elements.append(Paragraph(report_subtitle, subtitle_style))

    generated_text = generated_label or f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    elements.append(Paragraph(generated_text, meta_style))
    elements.append(Spacer(1, 14))


def build_export_table(
    data,
    col_widths=None,
    body_font_size=8,
    header_background=EXPORT_PRIMARY,
    row_backgrounds=None,
    centered_columns=None,
    right_aligned_columns=None,
    total_row_indexes=None,
):
    table = Table(data, colWidths=col_widths, repeatRows=1)
    row_backgrounds = row_backgrounds or [colors.white, EXPORT_ALT_ROW]
    centered_columns = centered_columns or []
    right_aligned_columns = right_aligned_columns or []
    total_row_indexes = total_row_indexes or []

    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), header_background),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), EXPORT_FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), body_font_size + 1),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 1), (-1, -1), EXPORT_FONT),
        ('FONTSIZE', (0, 1), (-1, -1), body_font_size),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), row_backgrounds),
        ('GRID', (0, 0), (-1, -1), 0.45, EXPORT_BORDER),
        ('BOX', (0, 0), (-1, -1), 0.8, header_background),
    ]

    for column_index in centered_columns:
        style_commands.append(('ALIGN', (column_index, 1), (column_index, -1), 'CENTER'))

    for column_index in right_aligned_columns:
        style_commands.append(('ALIGN', (column_index, 1), (column_index, -1), 'RIGHT'))

    for row_index in total_row_indexes:
        style_commands.extend([
            ('BACKGROUND', (0, row_index), (-1, row_index), EXPORT_TOTAL_BG),
            ('TEXTCOLOR', (0, row_index), (-1, row_index), EXPORT_TOTAL_TEXT),
            ('FONTNAME', (0, row_index), (-1, row_index), EXPORT_FONT_BOLD),
        ])

    table.setStyle(TableStyle(style_commands))
    return table


def prepend_row_numbers(headers, rows, label='No.'):
    numbered_rows = [[str(index), *list(row)] for index, row in enumerate(rows, start=1)]
    return [[label, *list(headers)], *numbered_rows]


def write_excel_report_header(worksheet, title, subtitle, total_columns):
    end_column = get_column_letter(total_columns)
    worksheet.merge_cells(f'A1:{end_column}1')
    worksheet['A1'] = title
    worksheet['A1'].font = Font(size=14, bold=True)
    worksheet['A1'].alignment = Alignment(horizontal='center')

    worksheet.merge_cells(f'A2:{end_column}2')
    worksheet['A2'] = subtitle
    worksheet['A2'].alignment = Alignment(horizontal='center')
    worksheet['A2'].font = Font(size=10)


def style_excel_header(worksheet, row_idx, fill_color='0F766E'):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in worksheet[row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def autosize_worksheet_columns(worksheet, max_width=32):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)
