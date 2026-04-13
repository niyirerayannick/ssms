from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime

from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.core.mail import EmailMessage
from django.db.models import Prefetch, Q, Value, CharField
from django.db.models.functions import Coalesce

from openpyxl import Workbook
from openpyxl.styles import Alignment
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph

from core.export_utils import (
    ExportNumberedCanvas,
    autosize_worksheet_columns,
    build_export_pdf_document,
    build_export_table,
    prepend_row_numbers,
    style_excel_header,
    style_excel_table_rows,
    write_excel_report_header,
)
from core.models import School
from core.utils import normalize_identifier_value, format_money
from families.models import Family
from finance.models import SchoolFee
from insurance.models import FamilyInsurance
from students.models import Student


@dataclass(frozen=True)
class ReportDefinition:
    key: str
    label: str
    permission: str
    filters: tuple[str, ...]
    formats: tuple[str, ...] = ("pdf", "excel")


REPORT_DEFINITIONS = {
    "students": ReportDefinition(
        key="students",
        label="Students List",
        permission="students.view_student",
        filters=("academic_year", "district", "sector", "school", "gender", "school_level", "sponsorship_status", "enrollment_status", "payment_ability", "mutuelle_support_status", "age_from", "age_to"),
    ),
    "families": ReportDefinition(
        key="families",
        label="Families Directory",
        permission="families.view_family",
        filters=("district", "sector", "payment_ability", "mutuelle_support_status"),
    ),
    "schools": ReportDefinition(
        key="schools",
        label="Schools Directory",
        permission="core.view_school",
        filters=("district", "sector"),
    ),
    "fees": ReportDefinition(
        key="fees",
        label="School Fees Summary",
        permission="finance.view_schoolfee",
        filters=("academic_year", "district", "school", "payment_status"),
    ),
    "insurance": ReportDefinition(
        key="insurance",
        label="Mutuelle Coverage",
        permission="insurance.view_familyinsurance",
        filters=("academic_year", "district", "sector", "coverage_status"),
    ),
    "supported_mutuelle_families": ReportDefinition(
        key="supported_mutuelle_families",
        label="Supported Mutuelle Families",
        permission="families.view_family",
        filters=("district", "sector"),
    ),
}

ARRANGEMENT_LABELS = {
    "default": "Default arrangement",
    "academic_year": "Academic Year",
    "age": "Age",
    "district": "District",
    "sector": "Sector",
    "school": "School",
    "gender": "Gender",
    "school_level": "Education Level",
    "sponsorship_status": "Sponsorship Status",
    "enrollment_status": "Academic Status",
    "payment_ability": "Payment Ability",
    "mutuelle_support_status": "Mutuelle Support",
    "payment_status": "School Fees Status",
    "coverage_status": "Coverage Status",
}

REPORT_ARRANGEMENT_FIELDS = {
    "students": ("age", "district", "sector", "school", "gender", "school_level", "sponsorship_status", "enrollment_status", "payment_ability", "mutuelle_support_status"),
    "families": ("district", "sector", "payment_ability", "mutuelle_support_status"),
    "schools": ("district", "sector"),
    "fees": ("academic_year", "district", "school", "payment_status"),
    "insurance": ("academic_year", "district", "sector", "coverage_status"),
    "supported_mutuelle_families": ("district", "sector"),
}


def get_available_reports_for_user(user):
    return [report for report in REPORT_DEFINITIONS.values() if user.has_perm(report.permission)]


def get_report_definition(report_key):
    return REPORT_DEFINITIONS.get(report_key)


def get_arrangement_choices(report_key):
    options = [("", ARRANGEMENT_LABELS["default"])]
    for field_name in REPORT_ARRANGEMENT_FIELDS.get(report_key, ()):
        options.append((field_name, ARRANGEMENT_LABELS.get(field_name, field_name.replace("_", " ").title())))
    return options


def _arrangement_label(arrangement):
    return ARRANGEMENT_LABELS.get(arrangement, arrangement.replace("_", " ").title())


def _apply_report_arrangement(queryset, report_key, arrangement):
    if not arrangement:
        return queryset

    if report_key == "students":
        if arrangement == "age":
            return queryset.order_by("-date_of_birth", "last_name", "first_name")
        if arrangement == "district":
            queryset = queryset.annotate(
                arrangement_district=Coalesce(
                    "partner__district__name",
                    "family__district__name",
                    "school__district__name",
                    Value("", output_field=CharField()),
                )
            )
            return queryset.order_by("arrangement_district", "last_name", "first_name")
        if arrangement == "sector":
            queryset = queryset.annotate(
                arrangement_sector=Coalesce(
                    "partner__sector__name",
                    "family__sector__name",
                    "school__sector__name",
                    Value("", output_field=CharField()),
                )
            )
            return queryset.order_by("arrangement_sector", "last_name", "first_name")
        if arrangement == "school":
            queryset = queryset.annotate(
                arrangement_school=Coalesce(
                    "school__name",
                    "school_name",
                    Value("", output_field=CharField()),
                )
            )
            return queryset.order_by("arrangement_school", "last_name", "first_name")
        if arrangement == "gender":
            return queryset.order_by("gender", "last_name", "first_name")
        if arrangement == "school_level":
            return queryset.order_by("school_level", "class_level", "last_name", "first_name")
        if arrangement == "sponsorship_status":
            return queryset.order_by("sponsorship_status", "last_name", "first_name")
        if arrangement == "enrollment_status":
            return queryset.order_by("enrollment_status", "last_name", "first_name")
        if arrangement == "payment_ability":
            return queryset.order_by("family__payment_ability", "last_name", "first_name")
        if arrangement == "mutuelle_support_status":
            return queryset.order_by("family__mutuelle_support_status", "last_name", "first_name")
        return queryset.order_by("last_name", "first_name")

    if report_key in {"families", "supported_mutuelle_families"}:
        if arrangement == "district":
            return queryset.order_by("district__name", "head_of_family")
        if arrangement == "sector":
            return queryset.order_by("sector__name", "head_of_family")
        if arrangement == "payment_ability":
            return queryset.order_by("payment_ability", "head_of_family")
        if arrangement == "mutuelle_support_status":
            return queryset.order_by("mutuelle_support_status", "head_of_family")
        return queryset.order_by("head_of_family")

    if report_key == "schools":
        if arrangement == "district":
            return queryset.order_by("district__name", "name")
        if arrangement == "sector":
            return queryset.order_by("sector__name", "name")
        return queryset.order_by("name")

    if report_key == "fees":
        if arrangement == "academic_year":
            return queryset.order_by("academic_year__name", "student__last_name", "student__first_name", "term")
        if arrangement == "district":
            return queryset.order_by("student__family__district__name", "student__last_name", "student__first_name", "term")
        if arrangement == "school":
            return queryset.order_by("student__school__name", "student__last_name", "student__first_name", "term")
        if arrangement == "payment_status":
            return queryset.order_by("payment_status", "student__last_name", "student__first_name", "term")
        return queryset.order_by("student__last_name", "student__first_name", "term")

    if report_key == "insurance":
        if arrangement == "academic_year":
            return queryset.order_by("insurance_year__name", "family__head_of_family")
        if arrangement == "district":
            return queryset.order_by("family__district__name", "family__head_of_family", "insurance_year__name")
        if arrangement == "sector":
            return queryset.order_by("family__sector__name", "family__head_of_family", "insurance_year__name")
        if arrangement == "coverage_status":
            return queryset.order_by("coverage_status", "family__head_of_family", "insurance_year__name")
        return queryset.order_by("family__head_of_family", "insurance_year__name")

    return queryset


def ensure_report_permission(user, report_key):
    report = get_report_definition(report_key)
    if not report or not user.has_perm(report.permission):
        raise PermissionDenied("You do not have permission to send this report.")
    return report


def _safe_year_shift(years_ago: int):
    today = date.today()
    try:
        return today.replace(year=today.year - years_ago)
    except ValueError:
        return today.replace(month=2, day=28, year=today.year - years_ago)


def _normalize_age_bounds(age_from, age_to):
    parsed_from = parsed_to = None
    if age_from not in (None, ""):
        try:
            parsed_from = max(int(age_from), 0)
        except (TypeError, ValueError):
            parsed_from = None
    if age_to not in (None, ""):
        try:
            parsed_to = max(int(age_to), 0)
        except (TypeError, ValueError):
            parsed_to = None
    if parsed_from is not None and parsed_to is not None and parsed_from > parsed_to:
        parsed_from, parsed_to = parsed_to, parsed_from
    return parsed_from, parsed_to


def _student_guardian_parent_label(student):
    family = student.family
    if family:
        if family.guardian_name:
            return family.guardian_name
        parent_names = [name for name in [family.father_name, family.mother_name] if name]
        if parent_names:
            return " / ".join(parent_names)
        return family.head_of_family or "N/A"
    if student.partner and student.partner.contact_person:
        return student.partner.contact_person
    return "N/A"


def _student_phone_label(student):
    family = student.family
    if family:
        if family.guardian_phone:
            return normalize_identifier_value(family.guardian_phone, "N/A")
        if family.phone_number:
            return normalize_identifier_value(family.phone_number, "N/A")
        if family.alternative_phone:
            return normalize_identifier_value(family.alternative_phone, "N/A")
    if student.partner and student.partner.phone:
        return normalize_identifier_value(student.partner.phone, "N/A")
    return "N/A"


def _student_sector_label(student):
    if student.partner and student.partner.sector:
        return student.partner.sector.name
    if student.family and student.family.sector:
        return student.family.sector.name
    if student.school and student.school.sector:
        return student.school.sector.name
    return "N/A"


def _student_export_row(student):
    return [
        student.full_name,
        student.get_gender_display(),
        str(student.age if student.age is not None else "N/A"),
        student.get_school_level_display() if student.school_level else "N/A",
        student.class_level or "N/A",
        student.school.name if student.school else (student.school_name or "N/A"),
        student.family_district_name,
        _student_sector_label(student),
        _student_guardian_parent_label(student),
        _student_phone_label(student),
        student.get_sponsorship_status_display(),
    ]


def _student_queryset(cleaned_data):
    students = Student.objects.select_related(
        "school",
        "family__district",
        "family__sector",
        "partner__district",
        "partner__sector",
    ).prefetch_related(
        Prefetch(
            "fees",
            queryset=SchoolFee.objects.select_related("academic_year").order_by("-academic_year__name", "-created_at"),
            to_attr="prefetched_fees",
        )
    )
    subtitle_parts = ["All Students"]

    academic_year = cleaned_data.get("academic_year")
    if academic_year:
        students = students.filter(
            Q(academic_records__academic_year=academic_year) | Q(fees__academic_year=academic_year)
        ).distinct()
        subtitle_parts = [f"Academic Year: {academic_year.name}"]

    district = cleaned_data.get("district")
    if district:
        students = students.filter(
            Q(family__district=district) | Q(partner__district=district) | Q(school__district=district)
        )
        subtitle_parts.append(district.name)

    sector = cleaned_data.get("sector")
    if sector:
        students = students.filter(
            Q(family__sector=sector) | Q(partner__sector=sector) | Q(school__sector=sector)
        )
        subtitle_parts.append(sector.name)

    school = cleaned_data.get("school")
    if school:
        students = students.filter(school=school)
        subtitle_parts.append(school.name)

    for field_name in ("gender", "school_level", "sponsorship_status", "enrollment_status"):
        field_value = cleaned_data.get(field_name)
        if field_value:
            students = students.filter(**{field_name: field_value})

    if cleaned_data.get("payment_ability"):
        students = students.filter(family__payment_ability=cleaned_data["payment_ability"])
    if cleaned_data.get("mutuelle_support_status"):
        students = students.filter(family__mutuelle_support_status=cleaned_data["mutuelle_support_status"])

    age_from, age_to = _normalize_age_bounds(cleaned_data.get("age_from"), cleaned_data.get("age_to"))
    if age_from is not None:
        students = students.filter(date_of_birth__lte=_safe_year_shift(age_from))
    if age_to is not None:
        students = students.filter(date_of_birth__gt=_safe_year_shift(age_to + 1))
    if age_from is not None or age_to is not None:
        subtitle_parts.append(
            f"Age {age_from if age_from is not None else 0} to {age_to if age_to is not None else 'above'}"
        )

    students = students.distinct()
    arrangement = cleaned_data.get("arrangement")
    students = _apply_report_arrangement(students, "students", arrangement)
    return students, " - ".join(subtitle_parts)


def _families_queryset(cleaned_data, supported_only=False):
    families = Family.objects.select_related("district", "sector").all()
    subtitle_parts = ["All Families" if not supported_only else "Families Unable to Pay and Supported"]

    if supported_only:
        families = families.filter(
            payment_ability=Family.PAYMENT_ABILITY_UNABLE,
            mutuelle_support_status=Family.MUTUELLE_SUPPORT_STATUS_SUPPORTED,
        )

    if cleaned_data.get("district"):
        families = families.filter(district=cleaned_data["district"])
        subtitle_parts.append(cleaned_data["district"].name)
    if cleaned_data.get("sector"):
        families = families.filter(sector=cleaned_data["sector"])
        subtitle_parts.append(cleaned_data["sector"].name)

    if not supported_only:
        if cleaned_data.get("payment_ability"):
            families = families.filter(payment_ability=cleaned_data["payment_ability"])
        if cleaned_data.get("mutuelle_support_status"):
            families = families.filter(mutuelle_support_status=cleaned_data["mutuelle_support_status"])

    arrangement = cleaned_data.get("arrangement")
    families = _apply_report_arrangement(families, "supported_mutuelle_families" if supported_only else "families", arrangement)
    return families, " - ".join(subtitle_parts)


def _schools_queryset(cleaned_data):
    schools = School.objects.select_related("district", "sector").all()
    subtitle_parts = ["All Schools"]
    if cleaned_data.get("district"):
        schools = schools.filter(district=cleaned_data["district"])
        subtitle_parts.append(cleaned_data["district"].name)
    if cleaned_data.get("sector"):
        schools = schools.filter(sector=cleaned_data["sector"])
        subtitle_parts.append(cleaned_data["sector"].name)
    arrangement = cleaned_data.get("arrangement")
    schools = _apply_report_arrangement(schools, "schools", arrangement)
    return schools, " - ".join(subtitle_parts)


def _fees_queryset(cleaned_data):
    fees = SchoolFee.objects.select_related("student", "student__school", "student__family__district").all()
    subtitle_parts = ["All Fee Records"]
    if cleaned_data.get("academic_year"):
        fees = fees.filter(academic_year=cleaned_data["academic_year"])
        subtitle_parts = [f"Academic Year: {cleaned_data['academic_year'].name}"]
    if cleaned_data.get("district"):
        fees = fees.filter(student__family__district=cleaned_data["district"])
        subtitle_parts.append(cleaned_data["district"].name)
    if cleaned_data.get("school"):
        fees = fees.filter(student__school=cleaned_data["school"])
        subtitle_parts.append(cleaned_data["school"].name)
    if cleaned_data.get("payment_status"):
        fees = fees.filter(payment_status=cleaned_data["payment_status"])
    arrangement = cleaned_data.get("arrangement")
    fees = _apply_report_arrangement(fees, "fees", arrangement)
    return fees, " - ".join(subtitle_parts)


def _insurance_queryset(cleaned_data):
    records = FamilyInsurance.objects.select_related("family", "insurance_year", "family__district", "family__sector").all()
    subtitle_parts = ["All Insurance Records"]
    if cleaned_data.get("academic_year"):
        records = records.filter(insurance_year=cleaned_data["academic_year"])
        subtitle_parts = [f"Academic Year: {cleaned_data['academic_year'].name}"]
    if cleaned_data.get("district"):
        records = records.filter(family__district=cleaned_data["district"])
        subtitle_parts.append(cleaned_data["district"].name)
    if cleaned_data.get("sector"):
        records = records.filter(family__sector=cleaned_data["sector"])
        subtitle_parts.append(cleaned_data["sector"].name)
    if cleaned_data.get("coverage_status"):
        records = records.filter(coverage_status=cleaned_data["coverage_status"])
    arrangement = cleaned_data.get("arrangement")
    records = _apply_report_arrangement(records, "insurance", arrangement)
    return records, " - ".join(subtitle_parts)


def _build_students_pdf(queryset, subtitle):
    from reports.views import create_letterhead

    buffer = io.BytesIO()
    doc = build_export_pdf_document(buffer, "Students List Report", pagesize=landscape(A4), left_margin=36, right_margin=36, top_margin=40, bottom_margin=48)
    elements = []
    create_letterhead(elements, "Students List Report", f"{subtitle} (Total: {queryset.count()})")
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("StudentExportCell", parent=styles["BodyText"], fontSize=6.2, leading=7.2, wordWrap="CJK")
    centered_style = ParagraphStyle("StudentExportCellCentered", parent=cell_style, alignment=TA_CENTER)
    rows = []
    for student in queryset:
        row = _student_export_row(student)
        rows.append([
            Paragraph(row[0], cell_style),
            Paragraph(row[1], centered_style),
            Paragraph(row[2], centered_style),
            Paragraph(row[3], centered_style),
            Paragraph(row[4], centered_style),
            Paragraph(row[5], cell_style),
            Paragraph(row[6], centered_style),
            Paragraph(row[7], centered_style),
            Paragraph(row[8], cell_style),
            Paragraph(row[9], centered_style),
            Paragraph(row[10], centered_style),
        ])
    data = prepend_row_numbers(["Full Name", "Gender", "Age", "Education Level", "Class/Year", "School", "District", "Sector", "Guardian/Parent", "Phone", "Sponsorship Status"], rows)
    elements.append(build_export_table(data, col_widths=[24, 92, 34, 26, 54, 50, 104, 54, 54, 126, 72, 70], body_font_size=6.2, centered_columns=[0, 2, 3, 4, 5, 7, 10, 11]))
    doc.build(elements, canvasmaker=ExportNumberedCanvas)
    return buffer.getvalue()


def _build_students_excel(queryset, subtitle):
    wb = Workbook()
    ws = wb.active
    ws.title = "Students List"
    headers = ["No.", "Full Name", "Gender", "Age", "Education Level", "Class/Year", "School", "District", "Sector", "Guardian/Parent", "Phone", "Sponsorship Status"]
    header_row = write_excel_report_header(ws, "Students List Report", subtitle, len(headers))
    ws.append(headers)
    style_excel_header(ws, header_row)
    data_start_row = header_row + 1
    for index, student in enumerate(queryset, start=1):
        ws.append([index, *_student_export_row(student)])
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_index, width in enumerate([8, 24, 12, 8, 16, 14, 24, 16, 14, 22, 16, 18], start=1):
        ws.column_dimensions[chr(64 + column_index)].width = width
    style_excel_table_rows(ws, header_row_idx=header_row, data_start_row=data_start_row, data_end_row=ws.max_row, max_col=len(headers), centered_columns=[1, 3, 4, 5, 6, 8, 11, 12])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_families_pdf(queryset, subtitle, title):
    from reports.views import create_letterhead

    buffer = io.BytesIO()
    doc = build_export_pdf_document(buffer, title, pagesize=landscape(A4))
    elements = []
    create_letterhead(elements, title, f"{subtitle} (Total: {queryset.count()})")
    rows = [[family.family_code, family.head_of_family, normalize_identifier_value(family.phone_number, "N/A"), str(family.total_family_members or 0), family.district.name if family.district else "N/A", family.sector.name if family.sector else "N/A", family.get_payment_ability_display(), family.get_mutuelle_support_status_display()] for family in queryset]
    data = prepend_row_numbers(["Family Code", "Head of Family", "Phone", "Members", "District", "Sector", "Payment Ability", "Mutuelle Support"], rows)
    elements.append(build_export_table(data, col_widths=[26, 95, 120, 80, 45, 72, 60, 74, 82], body_font_size=7, centered_columns=[0, 4, 5, 6, 7, 8]))
    doc.build(elements, canvasmaker=ExportNumberedCanvas)
    return buffer.getvalue()


def _build_families_excel(queryset, subtitle, title, sheet_title):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    headers = ["No.", "Family Code", "Head of Family", "Phone", "Members", "District", "Sector", "Payment Ability", "Mutuelle Support"]
    header_row = write_excel_report_header(ws, title, subtitle, len(headers))
    ws.append(headers)
    style_excel_header(ws, header_row)
    data_start_row = header_row + 1
    for index, family in enumerate(queryset, start=1):
        ws.append([index, family.family_code, family.head_of_family, normalize_identifier_value(family.phone_number, "N/A"), family.total_family_members or 0, family.district.name if family.district else "N/A", family.sector.name if family.sector else "N/A", family.get_payment_ability_display(), family.get_mutuelle_support_status_display()])
    autosize_worksheet_columns(ws, max_width=24)
    style_excel_table_rows(ws, header_row_idx=header_row, data_start_row=data_start_row, data_end_row=ws.max_row, max_col=len(headers), centered_columns=[1, 5, 6, 7, 8, 9])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_schools_pdf(queryset, subtitle):
    from reports.views import create_letterhead

    buffer = io.BytesIO()
    doc = build_export_pdf_document(buffer, "Schools Directory Report", pagesize=landscape(A4))
    elements = []
    create_letterhead(elements, "Schools Directory Report", f"{subtitle} (Total: {queryset.count()})")
    rows = [[school.name, school.headteacher_name or "N/A", normalize_identifier_value(school.headteacher_mobile, "N/A"), school.district.name if school.district else "N/A", school.sector.name if school.sector else "N/A", format_money(school.fee_amount), school.bank_name or "N/A"] for school in queryset]
    data = prepend_row_numbers(["School Name", "Headteacher", "Phone", "District", "Sector", "Fee Amount", "Bank"], rows)
    elements.append(build_export_table(data, col_widths=[26, 132, 100, 92, 60, 56, 66, 100], body_font_size=7, centered_columns=[0, 4, 5, 6], right_aligned_columns=[6]))
    doc.build(elements, canvasmaker=ExportNumberedCanvas)
    return buffer.getvalue()


def _build_schools_excel(queryset, subtitle):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schools"
    headers = ["No.", "School Name", "Headteacher", "Phone", "District", "Sector", "Fee Amount", "Bank"]
    header_row = write_excel_report_header(ws, "Schools Directory Report", subtitle, len(headers))
    ws.append(headers)
    style_excel_header(ws, header_row)
    data_start_row = header_row + 1
    for index, school in enumerate(queryset, start=1):
        ws.append([index, school.name, school.headteacher_name or "N/A", normalize_identifier_value(school.headteacher_mobile, "N/A"), school.district.name if school.district else "N/A", school.sector.name if school.sector else "N/A", float(school.fee_amount or 0), school.bank_name or "N/A"])
    autosize_worksheet_columns(ws, max_width=24)
    style_excel_table_rows(ws, header_row_idx=header_row, data_start_row=data_start_row, data_end_row=ws.max_row, max_col=len(headers), centered_columns=[1, 5, 6], right_aligned_columns=[7])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_fees_pdf(queryset, subtitle):
    from reports.views import create_letterhead

    buffer = io.BytesIO()
    doc = build_export_pdf_document(buffer, "School Fees Summary Report")
    elements = []
    create_letterhead(elements, "School Fees Summary Report", f"{subtitle} (Total: {queryset.count()})")
    rows = [[fee.student.full_name, f"Term {fee.term}", fee.student.school.name if fee.student.school else "N/A", f"{fee.total_fees:,.0f}", f"{fee.amount_paid:,.0f}", f"{fee.balance:,.0f}", fee.get_payment_status_display()] for fee in queryset]
    data = prepend_row_numbers(["Student Name", "Term", "School", "Required (RWF)", "Paid (RWF)", "Balance (RWF)", "Status"], rows)
    elements.append(build_export_table(data, col_widths=[32, 122, 54, 104, 72, 72, 72, 68], body_font_size=7, centered_columns=[0, 2, 4, 5, 6, 7]))
    doc.build(elements, canvasmaker=ExportNumberedCanvas)
    return buffer.getvalue()


def _build_fees_excel(queryset, subtitle):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fees Summary"
    headers = ["No.", "Student Name", "Term", "Required Fees", "Amount Paid", "Balance", "Status"]
    header_row = write_excel_report_header(ws, "School Fees Summary Report", subtitle, len(headers))
    ws.append(headers)
    style_excel_header(ws, header_row)
    data_start_row = header_row + 1
    for index, fee in enumerate(queryset, start=1):
        ws.append([index, fee.student.full_name, fee.term, float(fee.total_fees), float(fee.amount_paid), float(fee.balance), fee.get_payment_status_display()])
    autosize_worksheet_columns(ws, max_width=50)
    style_excel_table_rows(ws, header_row_idx=header_row, data_start_row=data_start_row, data_end_row=ws.max_row, max_col=len(headers), centered_columns=[1, 3, 7], right_aligned_columns=[4, 5, 6])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_insurance_pdf(queryset, subtitle):
    from reports.views import create_letterhead

    buffer = io.BytesIO()
    doc = build_export_pdf_document(buffer, "Mutuelle de Sante Coverage Report")
    elements = []
    create_letterhead(elements, "Mutuelle de Sante Coverage Report", f"{subtitle} (Total: {queryset.count()})")
    rows = []
    for insurance in queryset:
        balance = float(insurance.required_amount) - float(insurance.amount_paid)
        rows.append([insurance.family.head_of_family, insurance.insurance_year.name if insurance.insurance_year else "", f"{insurance.required_amount:,.0f}", f"{insurance.amount_paid:,.0f}", f"{balance:,.0f}", insurance.get_coverage_status_display()])
    data = prepend_row_numbers(["Family Head", "Year", "Required (RWF)", "Paid (RWF)", "Balance (RWF)", "Status"], rows)
    elements.append(build_export_table(data, col_widths=[32, 151, 72, 90, 90, 90, 80], body_font_size=8, centered_columns=[0, 2, 3, 4, 5, 6]))
    doc.build(elements, canvasmaker=ExportNumberedCanvas)
    return buffer.getvalue()


def _build_insurance_excel(queryset, subtitle):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mutuelle Coverage"
    headers = ["No.", "Family Head", "Year", "Required Amount", "Amount Paid", "Balance", "Status"]
    header_row = write_excel_report_header(ws, "Mutuelle de Sante Coverage Report", subtitle, len(headers))
    ws.append(headers)
    style_excel_header(ws, header_row)
    data_start_row = header_row + 1
    for index, insurance in enumerate(queryset, start=1):
        balance = float(insurance.required_amount) - float(insurance.amount_paid)
        ws.append([index, insurance.family.head_of_family, insurance.insurance_year.name if insurance.insurance_year else "", float(insurance.required_amount), float(insurance.amount_paid), balance, insurance.get_coverage_status_display()])
    autosize_worksheet_columns(ws, max_width=28)
    style_excel_table_rows(ws, header_row_idx=header_row, data_start_row=data_start_row, data_end_row=ws.max_row, max_col=len(headers), centered_columns=[1, 3, 7], right_aligned_columns=[4, 5, 6])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_report_attachment(report_key, export_format, cleaned_data):
    report = get_report_definition(report_key)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    arrangement = cleaned_data.get("arrangement")

    if report_key == "students":
        queryset, subtitle = _student_queryset(cleaned_data)
        content = _build_students_pdf(queryset, subtitle) if export_format == "pdf" else _build_students_excel(queryset, subtitle)
        filename = f"students_list_{timestamp}.{'pdf' if export_format == 'pdf' else 'xlsx'}"
    elif report_key == "families":
        queryset, subtitle = _families_queryset(cleaned_data)
        content = _build_families_pdf(queryset, subtitle, "Families Directory Report") if export_format == "pdf" else _build_families_excel(queryset, subtitle, "Families Directory Report", "Families")
        filename = f"families_directory_{timestamp}.{'pdf' if export_format == 'pdf' else 'xlsx'}"
    elif report_key == "schools":
        queryset, subtitle = _schools_queryset(cleaned_data)
        content = _build_schools_pdf(queryset, subtitle) if export_format == "pdf" else _build_schools_excel(queryset, subtitle)
        filename = f"schools_directory_{timestamp}.{'pdf' if export_format == 'pdf' else 'xlsx'}"
    elif report_key == "fees":
        queryset, subtitle = _fees_queryset(cleaned_data)
        content = _build_fees_pdf(queryset, subtitle) if export_format == "pdf" else _build_fees_excel(queryset, subtitle)
        filename = f"school_fees_{timestamp}.{'pdf' if export_format == 'pdf' else 'xlsx'}"
    elif report_key == "insurance":
        queryset, subtitle = _insurance_queryset(cleaned_data)
        content = _build_insurance_pdf(queryset, subtitle) if export_format == "pdf" else _build_insurance_excel(queryset, subtitle)
        filename = f"mutuelle_coverage_{timestamp}.{'pdf' if export_format == 'pdf' else 'xlsx'}"
    elif report_key == "supported_mutuelle_families":
        queryset, subtitle = _families_queryset(cleaned_data, supported_only=True)
        content = _build_families_pdf(queryset, subtitle, "Supported Mutuelle Families Report") if export_format == "pdf" else _build_families_excel(queryset, subtitle, "Supported Mutuelle Families Report", "Supported Families")
        filename = f"supported_mutuelle_families_{timestamp}.{'pdf' if export_format == 'pdf' else 'xlsx'}"
    else:
        raise ValueError("Unsupported report type selected.")

    if arrangement:
        subtitle = f"{subtitle} - Arranged by {_arrangement_label(arrangement)}"

    return {
        "filename": filename,
        "content_type": "application/pdf" if export_format == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "content": content,
        "report": report,
        "subtitle": subtitle,
        "record_count": queryset.count(),
    }


def build_filter_preview(report_key, cleaned_data):
    report = get_report_definition(report_key)
    if not report:
        return []
    labels = {
        "academic_year": "Academic Year",
        "district": "District",
        "sector": "Sector",
        "school": "School",
        "gender": "Gender",
        "school_level": "Education Level",
        "sponsorship_status": "Sponsorship Status",
        "enrollment_status": "Academic Status",
        "payment_ability": "Payment Ability",
        "mutuelle_support_status": "Mutuelle Support",
        "coverage_status": "Coverage Status",
        "payment_status": "School Fees Status",
        "age_from": "Age From",
        "age_to": "Age To",
        "arrangement": "Arrangement",
    }
    choices_lookup = {
        "gender": dict(Student.GENDER_CHOICES),
        "school_level": dict(Student.SCHOOL_LEVEL_CHOICES),
        "sponsorship_status": dict(Student.SPONSORSHIP_STATUS_CHOICES),
        "enrollment_status": dict(Student.ENROLLMENT_STATUS_CHOICES),
        "payment_ability": dict(Family.PAYMENT_ABILITY_CHOICES),
        "mutuelle_support_status": dict(Family.MUTUELLE_SUPPORT_STATUS_CHOICES),
        "payment_status": dict(SchoolFee.PAYMENT_STATUS_CHOICES),
        "coverage_status": dict(FamilyInsurance.COVERAGE_STATUS_CHOICES),
    }
    preview = []
    arrangement = cleaned_data.get("arrangement")
    if arrangement:
        preview.append({
            "label": labels["arrangement"],
            "value": _arrangement_label(arrangement),
        })
    for field_name in report.filters:
        value = cleaned_data.get(field_name)
        if value in (None, "", []):
            continue
        display_value = getattr(value, "name", str(value))
        if field_name in choices_lookup:
            display_value = choices_lookup[field_name].get(value, display_value)
        preview.append({"label": labels.get(field_name, field_name.replace("_", " ").title()), "value": display_value})
    return preview


def send_report_email(*, recipients, subject, body, attachment_name, attachment_bytes, attachment_content_type):
    sender_address = settings.EMAIL_HOST_USER or "noreply@sims.com"
    from_email = getattr(
        settings,
        "REPORTS_FROM_EMAIL",
        f"SAF-IMS Notifications <{sender_address}>",
    )
    email = EmailMessage(subject=subject, body=body, from_email=from_email, to=recipients)
    email.attach(attachment_name, attachment_bytes, attachment_content_type)
    email.send(fail_silently=False)
