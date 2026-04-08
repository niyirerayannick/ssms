from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.db.models import Prefetch
from django.db.models import Sum, Count, Avg, Q
from students.models import Student, StudentMark, StudentMaterial
from finance.models import SchoolFee
from insurance.models import FamilyInsurance
from families.models import Family
from core.models import AcademicYear, Partner, District, School
from core.export_utils import (
    ExportNumberedCanvas,
    add_export_header,
    autosize_worksheet_columns,
    build_export_pdf_document,
    build_export_table,
    prepend_row_numbers,
    resolve_logo_path,
    style_excel_header,
    style_excel_table_rows,
    write_excel_report_header,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, date
import io


NumberedCanvas = ExportNumberedCanvas


def _resolve_logo_path():
    return resolve_logo_path()


def _build_age_band_analysis(students_queryset):
    """Return age-band rows and chart data from student DOB values."""

    today = date.today()
    age_bands = [
        {'label': 'Under 6', 'min': None, 'max': 5, 'count': 0},
        {'label': '6-10', 'min': 6, 'max': 10, 'count': 0},
        {'label': '11-15', 'min': 11, 'max': 15, 'count': 0},
        {'label': '16-20', 'min': 16, 'max': 20, 'count': 0},
        {'label': '21-25', 'min': 21, 'max': 25, 'count': 0},
        {'label': '26+', 'min': 26, 'max': None, 'count': 0},
    ]

    students_with_dob = 0
    students_missing_dob = 0
    university_with_dob = 0
    university_missing_dob = 0

    total_students = students_queryset.count()
    for school_level, date_of_birth in students_queryset.values_list('school_level', 'date_of_birth'):
        if not date_of_birth:
            students_missing_dob += 1
            if school_level == 'university':
                university_missing_dob += 1
            continue

        students_with_dob += 1
        if school_level == 'university':
            university_with_dob += 1

        age_years = today.year - date_of_birth.year - (
            (today.month, today.day) < (date_of_birth.month, date_of_birth.day)
        )

        for band in age_bands:
            lower_ok = band['min'] is None or age_years >= band['min']
            upper_ok = band['max'] is None or age_years <= band['max']
            if lower_ok and upper_ok:
                band['count'] += 1
                break

    for band in age_bands:
        band['percentage'] = round((band['count'] / total_students * 100) if total_students else 0, 1)

    return {
        'rows': age_bands,
        'labels': [band['label'] for band in age_bands],
        'counts': [band['count'] for band in age_bands],
        'students_with_dob': students_with_dob,
        'students_missing_dob': students_missing_dob,
        'university_with_dob': university_with_dob,
        'university_missing_dob': university_missing_dob,
    }


def create_letterhead(elements, report_title, report_subtitle=None):
    """Create professional letterhead for reports"""
    add_export_header(elements, report_title, report_subtitle)


def _safe_year_shift(years_ago):
    """Return today's date shifted back by whole years, handling leap years safely."""
    today = date.today()
    try:
        return today.replace(year=today.year - years_ago)
    except ValueError:
        return today.replace(month=2, day=28, year=today.year - years_ago)


def _apply_student_report_filters(request, queryset=None):
    """Apply shared student report filters, including age range."""
    students = queryset or Student.objects.select_related('school', 'program_officer', 'family', 'partner').all()

    year_id = request.GET.get('year')
    district_id = request.GET.get('district')
    age_from = request.GET.get('age_from')
    age_to = request.GET.get('age_to')

    subtitle_parts = ["All Students"]

    if year_id:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        students = students.filter(
            Q(academic_records__academic_year=academic_year) |
            Q(fees__academic_year=academic_year)
        ).distinct()
        subtitle_parts = [f"Academic Year: {academic_year.name}"]

    if district_id:
        district = get_object_or_404(District, id=district_id)
        students = students.filter(
            Q(family__district_id=district_id) | Q(partner__district_id=district_id)
        )
        subtitle_parts.append(district.name)

    parsed_age_from = None
    parsed_age_to = None

    if age_from:
        try:
            parsed_age_from = max(int(age_from), 0)
            born_on_or_before = _safe_year_shift(parsed_age_from)
            students = students.filter(date_of_birth__lte=born_on_or_before)
        except (TypeError, ValueError):
            parsed_age_from = None

    if age_to:
        try:
            parsed_age_to = max(int(age_to), 0)
            born_after = _safe_year_shift(parsed_age_to + 1)
            students = students.filter(date_of_birth__gt=born_after)
        except (TypeError, ValueError):
            parsed_age_to = None

    if parsed_age_from is not None and parsed_age_to is not None and parsed_age_from > parsed_age_to:
        parsed_age_from, parsed_age_to = parsed_age_to, parsed_age_from
        students = queryset or Student.objects.select_related('school', 'program_officer', 'family', 'partner').all()

        if year_id:
            academic_year = get_object_or_404(AcademicYear, id=year_id)
            students = students.filter(
                Q(academic_records__academic_year=academic_year) |
                Q(fees__academic_year=academic_year)
            ).distinct()
        if district_id:
            students = students.filter(
                Q(family__district_id=district_id) | Q(partner__district_id=district_id)
            )
        students = students.filter(
            date_of_birth__lte=_safe_year_shift(parsed_age_from),
            date_of_birth__gt=_safe_year_shift(parsed_age_to + 1),
        )

    if parsed_age_from is not None or parsed_age_to is not None:
        age_label = f"Age {parsed_age_from if parsed_age_from is not None else 0} to {parsed_age_to if parsed_age_to is not None else 'above'}"
        subtitle_parts.append(age_label)

    subtitle = " - ".join(subtitle_parts)
    return students.distinct(), subtitle, {
        'year_id': year_id,
        'district_id': district_id,
        'age_from': parsed_age_from,
        'age_to': parsed_age_to,
    }


def _apply_directory_district_filter(request, queryset, base_label="All Records"):
    """Apply the shared district filter to family and school directory reports."""
    district_id = request.GET.get('district')
    subtitle = base_label
    if district_id:
        district = get_object_or_404(District, id=district_id)
        queryset = queryset.filter(district_id=district_id)
        subtitle = f"{base_label} - {district.name}"
    return queryset, subtitle


def _student_guardian_parent_label(student):
    family = student.family
    if family:
        if family.guardian_name:
            return family.guardian_name
        parent_names = [name for name in [family.father_name, family.mother_name] if name]
        if parent_names:
            return " / ".join(parent_names)
        return family.head_of_family or 'N/A'
    if student.partner and student.partner.contact_person:
        return student.partner.contact_person
    return 'N/A'


def _student_phone_label(student):
    family = student.family
    if family:
        if family.guardian_phone:
            return family.guardian_phone
        if family.phone_number:
            return family.phone_number
        if family.alternative_phone:
            return family.alternative_phone
    if student.partner and student.partner.phone:
        return student.partner.phone
    return 'N/A'


def _student_sector_label(student):
    if student.partner and student.partner.sector:
        return student.partner.sector.name
    if student.family and student.family.sector:
        return student.family.sector.name
    if student.school and student.school.sector:
        return student.school.sector.name
    return 'N/A'


def _student_mutuelle_support_label(student):
    if student.family:
        return student.family.get_mutuelle_support_status_display()
    return 'N/A'


def _student_latest_fee_status(student):
    fee_records = getattr(student, 'prefetched_fees', None)
    if fee_records:
        return fee_records[0].get_payment_status_display()
    return 'No Fee Record'


def _student_export_row(student):
    return [
        student.full_name,
        student.get_gender_display(),
        str(student.age if student.age is not None else 'N/A'),
        student.get_school_level_display() if student.school_level else 'N/A',
        student.class_level or 'N/A',
        student.school.name if student.school else (student.school_name or 'N/A'),
        student.family_district_name,
        _student_sector_label(student),
        _student_guardian_parent_label(student),
        _student_phone_label(student),
        student.get_sponsorship_status_display(),
    ]


@login_required
@permission_required('students.view_student', raise_exception=True)
def students_pdf(request):
    """Export students list as PDF."""
    base_queryset = Student.objects.select_related(
        'school',
        'family__district',
        'family__sector',
        'partner__district',
        'partner__sector',
    ).prefetch_related(
        Prefetch(
            'fees',
            queryset=SchoolFee.objects.select_related('academic_year').order_by('-academic_year__name', '-created_at'),
            to_attr='prefetched_fees',
        )
    )
    students, subtitle, _filters = _apply_student_report_filters(request, queryset=base_queryset)

    buffer = io.BytesIO()
    doc = build_export_pdf_document(
        buffer,
        "Students List Report",
        pagesize=landscape(A4),
        left_margin=36,
        right_margin=36,
        top_margin=40,
        bottom_margin=48,
    )
    elements = []

    create_letterhead(
        elements,
        "Students List Report",
        f"{subtitle} (Total: {students.count()})"
    )

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        'StudentExportCell',
        parent=styles['BodyText'],
        fontSize=6.2,
        leading=7.2,
        wordWrap='CJK',
    )
    centered_cell_style = ParagraphStyle(
        'StudentExportCellCentered',
        parent=cell_style,
        alignment=TA_CENTER,
    )

    rows = []
    for student in students.order_by('last_name', 'first_name'):
        row = _student_export_row(student)
        rows.append([
            Paragraph(row[0], cell_style),
            Paragraph(row[1], centered_cell_style),
            Paragraph(row[2], centered_cell_style),
            Paragraph(row[3], centered_cell_style),
            Paragraph(row[4], centered_cell_style),
            Paragraph(row[5], cell_style),
            Paragraph(row[6], centered_cell_style),
            Paragraph(row[7], centered_cell_style),
            Paragraph(row[8], cell_style),
            Paragraph(row[9], centered_cell_style),
            Paragraph(row[10], centered_cell_style),
        ])

    data = prepend_row_numbers(
        [
            'Full Name',
            'Gender',
            'Age',
            'Education Level',
            'Class/Year',
            'School',
            'District',
            'Sector',
            'Guardian/Parent',
            'Phone',
            'Sponsorship Status',
        ],
        rows,
    )
    table = build_export_table(
        data,
        col_widths=[24, 92, 34, 26, 54, 50, 104, 54, 54, 126, 72, 70],
        body_font_size=6.2,
        centered_columns=[0, 2, 3, 4, 5, 7, 10, 11],
    )
    elements.append(table)
    doc.build(elements, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="students_list_report.pdf"'
    return response


@login_required
@permission_required('students.view_student', raise_exception=True)
def students_excel(request):
    """Export students list as Excel with optional age range filters."""
    base_queryset = Student.objects.select_related(
        'school',
        'family__district',
        'family__sector',
        'partner__district',
        'partner__sector',
    ).prefetch_related(
        Prefetch(
            'fees',
            queryset=SchoolFee.objects.select_related('academic_year').order_by('-academic_year__name', '-created_at'),
            to_attr='prefetched_fees',
        )
    )
    students, subtitle, filters = _apply_student_report_filters(request, queryset=base_queryset)
    students = students.order_by('last_name', 'first_name')

    wb = Workbook()
    ws = wb.active
    ws.title = "Students List"

    headers = [
        'No.',
        'Full Name',
        'Gender',
        'Age',
        'Education Level',
        'Class/Year',
        'School',
        'District',
        'Sector',
        'Guardian/Parent',
        'Phone',
        'Sponsorship Status',
    ]
    header_row = write_excel_report_header(ws, 'Students List Report', subtitle, len(headers))
    ws.append(headers)
    style_excel_header(ws, header_row)

    data_start_row = header_row + 1
    for index, student in enumerate(students, start=1):
        ws.append([index, *_student_export_row(student)])
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    preferred_widths = [8, 24, 12, 8, 16, 14, 24, 16, 14, 22, 16, 18]
    for column_index, width in enumerate(preferred_widths, start=1):
        ws.column_dimensions[chr(64 + column_index)].width = width
    style_excel_table_rows(
        ws,
        header_row_idx=header_row,
        data_start_row=data_start_row,
        data_end_row=ws.max_row,
        max_col=len(headers),
        centered_columns=[1, 3, 4, 5, 6, 8, 11, 12],
    )

    filename_parts = ["students_list"]
    if filters['age_from'] is not None or filters['age_to'] is not None:
        filename_parts.append(f"age_{filters['age_from'] if filters['age_from'] is not None else 0}_to_{filters['age_to'] if filters['age_to'] is not None else 'all'}")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{"_".join(filename_parts)}.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('students.view_student', raise_exception=True)
def sponsored_students_report(request):
    """Detailed report for sponsored (active) students."""
    year_id = request.GET.get('year')
    district_id = request.GET.get('district')
    students = (
        Student.objects.select_related('family', 'school', 'program_officer', 'partner')
        .filter(sponsorship_status='active')
    )
    
    subtitle = "All Sponsored Students"
    if year_id:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        students = students.filter(
            Q(academic_records__academic_year=academic_year) | 
            Q(fees__academic_year=academic_year)
        ).distinct()
        subtitle = f"Academic Year: {academic_year.name}"
    
    if district_id:
        district = get_object_or_404(District, id=district_id)
        # Include sponsored students coming from partners whose district matches as well
        students = students.filter(
            Q(family__district_id=district_id) | Q(partner__district_id=district_id)
        )
        subtitle += f" - {district.name}"
    
    students = students.order_by('last_name', 'first_name')
    
    total = students.count()
    boys = students.filter(gender='M').count()
    girls = students.filter(gender='F').count()
    with_disability = students.filter(has_disability=True).count()
    without_disability = students.filter(has_disability=False).count()

    context = {
        'students': students,
        'total': total,
        'boys': boys,
        'girls': girls,
        'with_disability': with_disability,
        'without_disability': without_disability,
        'subtitle': subtitle,
    }
    return render(request, 'reports/sponsored_students.html', context)


@login_required
@permission_required('finance.view_schoolfee', raise_exception=True)
def fees_pdf(request):
    """Export school fees summary as PDF."""
    year_id = request.GET.get('year')
    district_id = request.GET.get('district')
    fees = SchoolFee.objects.select_related('student', 'student__school').all()
    
    subtitle = "All Records"
    if year_id:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        fees = fees.filter(academic_year=academic_year)
        subtitle = f"Academic Year: {academic_year.name}"
    
    if district_id:
        district = get_object_or_404(District, id=district_id)
        fees = fees.filter(student__family__district_id=district_id)
        subtitle += f" - {district.name}"
    
    # Calculate summary statistics
    total_required = sum(float(fee.total_fees) for fee in fees)
    total_paid = sum(float(fee.amount_paid) for fee in fees)
    total_balance = sum(float(fee.balance) for fee in fees)
    
    paid_count = fees.filter(payment_status='paid').count()
    partial_count = fees.filter(payment_status='partial').count()
    pending_count = fees.filter(payment_status='pending').count()
    
    buffer = io.BytesIO()
    doc = build_export_pdf_document(
        buffer,
        "School Fees Summary Report",
    )
    elements = []
    
    # Add letterhead
    create_letterhead(
        elements,
        "School Fees Summary Report",
        f"{subtitle} (Total: {fees.count()})"
    )
    
    # Summary statistics box
    styles = getSampleStyleSheet()
    summary_style = ParagraphStyle(
        'SummaryBox',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#047857'),
        alignment=TA_CENTER
    )
    
    summary_data = [[
        Paragraph(f"<b>Paid:</b> {paid_count}", summary_style),
        Paragraph(f"<b>Partial:</b> {partial_count}", summary_style),
        Paragraph(f"<b>Pending:</b> {pending_count}", summary_style),
    ]]
    
    summary_table = build_export_table(
        summary_data,
        col_widths=[2.4 * inch, 2.4 * inch, 2.4 * inch],
        body_font_size=9,
        centered_columns=[0, 1, 2],
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    rows = []
    for fee in fees.order_by('student__last_name', 'student__first_name', 'term'):
        rows.append([
            fee.student.full_name,
            f"Term {fee.term}",
            fee.student.school.name if fee.student.school else 'N/A',
            f"{fee.total_fees:,.0f}",
            f"{fee.amount_paid:,.0f}",
            f"{fee.balance:,.0f}",
            fee.get_payment_status_display(),
        ])

    data = prepend_row_numbers(
        ['Student Name', 'Term', 'School', 'Required (RWF)', 'Paid (RWF)', 'Balance (RWF)', 'Status'],
        rows,
    )
    data.append([
        '',
        'TOTAL',
        '',
        '',
        f"{total_required:,.0f}",
        f"{total_paid:,.0f}",
        f"{total_balance:,.0f}",
        ''
    ])
    table = build_export_table(
        data,
        col_widths=[0.45 * inch, 1.7 * inch, 0.75 * inch, 1.45 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 0.95 * inch],
        body_font_size=7,
        centered_columns=[0, 2, 4, 5, 6, 7],
        total_row_indexes=[len(data) - 1],
    )
    elements.append(table)
    doc.build(elements, canvasmaker=NumberedCanvas)
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="school_fees_report.pdf"'
    return response


@login_required
@permission_required('finance.view_schoolfee', raise_exception=True)
def fees_excel(request):
    """Export fees summary as Excel."""
    year_id = request.GET.get('year')
    district_id = request.GET.get('district')
    fees = SchoolFee.objects.select_related('student').all()
    
    if year_id:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        fees = fees.filter(academic_year=academic_year)
    
    if district_id:
        district = get_object_or_404(District, id=district_id)
        fees = fees.filter(student__family__district_id=district_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fees Summary"
    
    headers = ['No.', 'Student Name', 'Term', 'Required Fees', 'Amount Paid', 'Balance', 'Status']
    header_row = write_excel_report_header(ws, 'School Fees Summary Report', 'Filtered fee records export', len(headers))
    ws.append(headers)
    style_excel_header(ws, header_row)
    
    # Data rows
    total_required = 0
    total_paid = 0
    total_balance = 0
    
    data_start_row = header_row + 1
    for index, fee in enumerate(fees.order_by('student__last_name', 'student__first_name', 'term'), start=1):
        ws.append([
            index,
            fee.student.full_name,
            fee.term,
            float(fee.total_fees),
            float(fee.amount_paid),
            float(fee.balance),
            fee.get_payment_status_display(),
        ])
        total_required += float(fee.total_fees)
        total_paid += float(fee.amount_paid)
        total_balance += float(fee.balance)
    
    # Summary row
    ws.append([])
    ws.append(['', 'TOTAL', '', total_required, total_paid, total_balance, ''])
    autosize_worksheet_columns(ws, max_width=50)
    style_excel_table_rows(
        ws,
        header_row_idx=header_row,
        data_start_row=data_start_row,
        data_end_row=ws.max_row,
        max_col=len(headers),
        centered_columns=[1, 3, 7],
        right_aligned_columns=[4, 5, 6],
        total_row_indexes=[ws.max_row],
    )
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="fees_summary.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('insurance.view_familyinsurance', raise_exception=True)
def insurance_pdf(request):
    """Export insurance coverage as PDF."""
    year_id = request.GET.get('year')
    district_id = request.GET.get('district')
    insurance_records = FamilyInsurance.objects.select_related('family').all()
    
    subtitle = "All Records"
    if year_id:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        insurance_records = insurance_records.filter(insurance_year=academic_year)
        subtitle = f"Academic Year: {academic_year.name}"
    
    if district_id:
        district = get_object_or_404(District, id=district_id)
        insurance_records = insurance_records.filter(family__district_id=district_id)
        subtitle += f" - {district.name}"
    
    # Calculate summary statistics
    covered = insurance_records.filter(coverage_status='covered').count()
    partially_covered = insurance_records.filter(coverage_status='partially_covered').count()
    not_covered = insurance_records.filter(coverage_status='not_covered').count()
    total_required = sum(float(i.required_amount) for i in insurance_records)
    total_paid = sum(float(i.amount_paid) for i in insurance_records)
    
    buffer = io.BytesIO()
    doc = build_export_pdf_document(
        buffer,
        "Mutuelle de Sante Coverage Report",
    )
    elements = []
    
    # Add letterhead
    create_letterhead(
        elements,
        "Mutuelle de Santé Coverage Report",
        f"{subtitle} (Total: {insurance_records.count()})"
    )
    
    # Summary statistics box
    styles = getSampleStyleSheet()
    summary_style = ParagraphStyle(
        'SummaryBox',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#047857'),
        alignment=TA_CENTER
    )
    
    summary_data = [[
        Paragraph(f"<b>Covered:</b> {covered}", summary_style),
        Paragraph(f"<b>Partially:</b> {partially_covered}", summary_style),
        Paragraph(f"<b>Not Covered:</b> {not_covered}", summary_style),
    ]]
    
    summary_table = build_export_table(
        summary_data,
        col_widths=[2.4 * inch, 2.4 * inch, 2.4 * inch],
        body_font_size=9,
        centered_columns=[0, 1, 2],
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    rows = []
    for insurance in insurance_records.order_by('family__head_of_family', 'insurance_year__name'):
        balance = float(insurance.required_amount) - float(insurance.amount_paid)
        rows.append([
            insurance.family.head_of_family,
            insurance.insurance_year.name if insurance.insurance_year else '',
            f"{insurance.required_amount:,.0f}",
            f"{insurance.amount_paid:,.0f}",
            f"{balance:,.0f}",
            insurance.get_coverage_status_display(),
        ])

    data = prepend_row_numbers(
        ['Family Head', 'Year', 'Required (RWF)', 'Paid (RWF)', 'Balance (RWF)', 'Status'],
        rows,
    )
    total_balance = total_required - total_paid
    data.append([
        '',
        'TOTAL',
        '',
        f"{total_required:,.0f}",
        f"{total_paid:,.0f}",
        f"{total_balance:,.0f}",
        ''
    ])
    table = build_export_table(
        data,
        col_widths=[0.45 * inch, 2.1 * inch, 1.0 * inch, 1.25 * inch, 1.25 * inch, 1.25 * inch, 1.1 * inch],
        body_font_size=8,
        centered_columns=[0, 2, 3, 4, 5, 6],
        total_row_indexes=[len(data) - 1],
    )
    elements.append(table)
    doc.build(elements, canvasmaker=NumberedCanvas)
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="mutuelle_coverage_report.pdf"'
    return response


@login_required
@permission_required('families.view_family', raise_exception=True)
def supported_mutuelle_families_pdf(request):
    """Export families unable to pay and supported in Mutuelle de Sante."""
    families = (
        Family.objects.select_related('province', 'district', 'sector')
        .filter(
            payment_ability=Family.PAYMENT_ABILITY_UNABLE,
            mutuelle_support_status=Family.MUTUELLE_SUPPORT_STATUS_SUPPORTED,
        )
        .order_by('head_of_family')
    )
    families, subtitle = _apply_directory_district_filter(
        request,
        families,
        base_label="Families Unable to Pay and Supported by Solidact",
    )

    buffer = io.BytesIO()
    doc = build_export_pdf_document(
        buffer,
        "Supported Mutuelle Families Report",
        pagesize=landscape(A4),
    )
    elements = []
    create_letterhead(
        elements,
        "Supported Mutuelle Families Report",
        f"{subtitle} (Total: {families.count()})"
    )

    rows = []
    for family in families:
        rows.append([
            family.family_code,
            family.head_of_family,
            family.phone_number or 'N/A',
            str(family.total_family_members or 0),
            family.district.name if family.district else 'N/A',
            family.sector.name if family.sector else 'N/A',
            family.get_payment_ability_display(),
            family.get_mutuelle_support_status_display(),
        ])

    data = prepend_row_numbers(
        ['Family Code', 'Head of Family', 'Phone', 'Members', 'District', 'Sector', 'Payment Ability', 'Mutuelle Support'],
        rows,
    )
    table = build_export_table(
        data,
        col_widths=[26, 95, 120, 80, 45, 72, 60, 74, 82],
        body_font_size=7,
        centered_columns=[0, 4, 5, 6, 7, 8],
    )
    elements.append(table)
    doc.build(elements, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="supported_mutuelle_families_report.pdf"'
    return response


@login_required
@permission_required('families.view_family', raise_exception=True)
def families_pdf(request):
    """Export family directory as PDF."""
    families = Family.objects.select_related('province', 'district', 'sector').all().order_by('head_of_family')
    families, subtitle = _apply_directory_district_filter(request, families, base_label="All Families")

    buffer = io.BytesIO()
    doc = build_export_pdf_document(
        buffer,
        "Families Directory Report",
        pagesize=landscape(A4),
    )
    elements = []
    create_letterhead(
        elements,
        "Families Directory Report",
        f"{subtitle} (Total: {families.count()})"
    )

    rows = []
    for family in families:
        rows.append([
            family.family_code,
            family.head_of_family,
            family.phone_number or 'N/A',
            str(family.total_family_members or 0),
            family.district.name if family.district else 'N/A',
            family.sector.name if family.sector else 'N/A',
            family.get_payment_ability_display(),
            family.get_mutuelle_support_status_display(),
        ])

    data = prepend_row_numbers(
        ['Family Code', 'Head of Family', 'Phone', 'Members', 'District', 'Sector', 'Payment Ability', 'Mutuelle Support'],
        rows,
    )
    table = build_export_table(
        data,
        col_widths=[26, 95, 120, 80, 45, 72, 60, 74, 82],
        body_font_size=7,
        centered_columns=[0, 4, 5, 6, 7, 8],
    )
    elements.append(table)
    doc.build(elements, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="families_directory_report.pdf"'
    return response


@login_required
@permission_required('families.view_family', raise_exception=True)
def families_excel(request):
    """Export family directory as Excel."""
    families = Family.objects.select_related('province', 'district', 'sector').all().order_by('head_of_family')
    families, subtitle = _apply_directory_district_filter(request, families, base_label="All Families")

    wb = Workbook()
    ws = wb.active
    ws.title = "Families"

    headers = ['No.', 'Family Code', 'Head of Family', 'Phone', 'Members', 'District', 'Sector', 'Payment Ability', 'Mutuelle Support']
    header_row = write_excel_report_header(ws, 'Families Directory Report', subtitle, len(headers))
    ws.append(headers)
    style_excel_header(ws, header_row)
    data_start_row = header_row + 1

    for index, family in enumerate(families, start=1):
        ws.append([
            index,
            family.family_code,
            family.head_of_family,
            family.phone_number or 'N/A',
            family.total_family_members or 0,
            family.district.name if family.district else 'N/A',
            family.sector.name if family.sector else 'N/A',
            family.get_payment_ability_display(),
            family.get_mutuelle_support_status_display(),
        ])

    autosize_worksheet_columns(ws, max_width=24)
    style_excel_table_rows(
        ws,
        header_row_idx=header_row,
        data_start_row=data_start_row,
        data_end_row=ws.max_row,
        max_col=len(headers),
        centered_columns=[1, 5, 6, 7, 8, 9],
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="families_directory_report.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('core.view_school', raise_exception=True)
def schools_pdf(request):
    """Export schools directory as PDF."""
    schools = School.objects.select_related('province', 'district', 'sector').all().order_by('name')
    schools, subtitle = _apply_directory_district_filter(request, schools, base_label="All Schools")

    buffer = io.BytesIO()
    doc = build_export_pdf_document(
        buffer,
        "Schools Directory Report",
        pagesize=landscape(A4),
    )
    elements = []
    create_letterhead(
        elements,
        "Schools Directory Report",
        f"{subtitle} (Total: {schools.count()})"
    )

    rows = []
    for school in schools:
        rows.append([
            school.name,
            school.headteacher_name or 'N/A',
            school.headteacher_mobile or 'N/A',
            school.district.name if school.district else 'N/A',
            school.sector.name if school.sector else 'N/A',
            f"{school.fee_amount:,.2f}",
            school.bank_name or 'N/A',
        ])

    data = prepend_row_numbers(
        ['School Name', 'Headteacher', 'Phone', 'District', 'Sector', 'Fee Amount', 'Bank'],
        rows,
    )
    table = build_export_table(
        data,
        col_widths=[26, 132, 100, 92, 60, 56, 66, 100],
        body_font_size=7,
        centered_columns=[0, 4, 5, 6],
        right_aligned_columns=[6],
    )
    elements.append(table)
    doc.build(elements, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="schools_directory_report.pdf"'
    return response


@login_required
@permission_required('core.view_school', raise_exception=True)
def schools_excel(request):
    """Export schools directory as Excel."""
    schools = School.objects.select_related('province', 'district', 'sector').all().order_by('name')
    schools, subtitle = _apply_directory_district_filter(request, schools, base_label="All Schools")

    wb = Workbook()
    ws = wb.active
    ws.title = "Schools"

    headers = ['No.', 'School Name', 'Headteacher', 'Phone', 'District', 'Sector', 'Fee Amount', 'Bank']
    header_row = write_excel_report_header(ws, 'Schools Directory Report', subtitle, len(headers))
    ws.append(headers)
    style_excel_header(ws, header_row)
    data_start_row = header_row + 1

    for index, school in enumerate(schools, start=1):
        ws.append([
            index,
            school.name,
            school.headteacher_name or 'N/A',
            school.headteacher_mobile or 'N/A',
            school.district.name if school.district else 'N/A',
            school.sector.name if school.sector else 'N/A',
            float(school.fee_amount or 0),
            school.bank_name or 'N/A',
        ])

    autosize_worksheet_columns(ws, max_width=24)
    style_excel_table_rows(
        ws,
        header_row_idx=header_row,
        data_start_row=data_start_row,
        data_end_row=ws.max_row,
        max_col=len(headers),
        centered_columns=[1, 5, 6],
        right_aligned_columns=[7],
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="schools_directory_report.xlsx"'
    wb.save(response)
    return response


@login_required
def reports_index(request):
    """Reports index page."""
    counts = Student.objects.aggregate(
        boarding_count=Count('id', filter=Q(boarding_status='boarding')),
        non_boarding_count=Count('id', filter=Q(boarding_status='non_boarding')),
        nursery_count=Count('id', filter=Q(school_level='nursery')),
        primary_count=Count('id', filter=Q(school_level='primary')),
        secondary_count=Count('id', filter=Q(school_level='secondary')),
        tvet_count=Count('id', filter=Q(school_level='tvet')),
    )
    
    # Get all academic years for filter dropdown
    academic_years = AcademicYear.objects.all().order_by('-name')
    # Get all districts for filter dropdown
    districts = District.objects.all().order_by('name')

    context = {
        'boarding_count': counts['boarding_count'],
        'non_boarding_count': counts['non_boarding_count'],
        'nursery_count': counts['nursery_count'],
        'primary_count': counts['primary_count'],
        'secondary_count': counts['secondary_count'],
        'tvet_count': counts['tvet_count'],
        'academic_years': academic_years,
        'districts': districts,
        'age_options': range(1, 31),
    }
    return render(request, 'reports/index.html', context)


@login_required
@permission_required('finance.view_schoolfee', raise_exception=True)
def financial_report_pdf(request):
    """Generate comprehensive financial report PDF."""
    year_id = request.GET.get('year')
    district_id = request.GET.get('district')
    
    # Base querysets
    fees_qs = SchoolFee.objects.all()
    insurance_qs = FamilyInsurance.objects.all()
    
    subtitle = "All Years"
    
    if year_id:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        fees_qs = fees_qs.filter(academic_year=academic_year)
        insurance_qs = insurance_qs.filter(insurance_year=academic_year)
        subtitle = f"Academic Year: {academic_year.name}"
    
    if district_id:
        district = get_object_or_404(District, id=district_id)
        fees_qs = fees_qs.filter(student__family__district_id=district_id)
        insurance_qs = insurance_qs.filter(family__district_id=district_id)
        subtitle += f" - {district.name}"
    
    # Calculate Fees Totals
    fees_total_req = fees_qs.aggregate(Sum('total_fees'))['total_fees__sum'] or 0
    fees_total_paid = fees_qs.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    fees_total_bal = fees_qs.aggregate(Sum('balance'))['balance__sum'] or 0
    
    # Calculate Insurance Totals
    insurance_total_req = insurance_qs.aggregate(Sum('required_amount'))['required_amount__sum'] or 0
    insurance_total_paid = insurance_qs.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    # Insurance balance is calculated differently in model save, but we can sum it up or calc here
    # Since FamilyInsurance has 'balance' field:
    insurance_total_bal = insurance_qs.aggregate(Sum('balance'))['balance__sum'] or 0
    
    # Grand Totals
    grand_total_req = fees_total_req + insurance_total_req
    grand_total_paid = fees_total_paid + insurance_total_paid
    grand_total_bal = fees_total_bal + insurance_total_bal
    
    # Create PDF
    buffer = io.BytesIO()
    doc = build_export_pdf_document(
        buffer,
        "Comprehensive Financial Report",
    )
    elements = []
    
    # Letterhead
    create_letterhead(elements, "Comprehensive Financial Report", subtitle)
    
    styles = getSampleStyleSheet()
    h2_style = ParagraphStyle(
        'Heading2Custom',
        parent=styles['Heading2'],
        textColor=colors.HexColor('#047857'),
        spaceBefore=12,
        spaceAfter=6
    )
    
    # 1. Executive Summary Table
    elements.append(Paragraph("Executive Summary", h2_style))
    
    summary_data = [
        ['Category', 'Required Amount (RWF)', 'Paid Amount (RWF)', 'Outstanding Balance (RWF)', 'Collection Rate'],
        ['School Fees', f"{fees_total_req:,.0f}", f"{fees_total_paid:,.0f}", f"{fees_total_bal:,.0f}", 
         f"{(fees_total_paid/fees_total_req*100 if fees_total_req else 0):.1f}%"],
        ['Insurance', f"{insurance_total_req:,.0f}", f"{insurance_total_paid:,.0f}", f"{insurance_total_bal:,.0f}",
         f"{(insurance_total_paid/insurance_total_req*100 if insurance_total_req else 0):.1f}%"],
        ['TOTAL', f"{grand_total_req:,.0f}", f"{grand_total_paid:,.0f}", f"{grand_total_bal:,.0f}",
         f"{(grand_total_paid/grand_total_req*100 if grand_total_req else 0):.1f}%"]
    ]
    
    summary_table = build_export_table(
        summary_data,
        col_widths=[2 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch, 1 * inch],
        body_font_size=9,
        right_aligned_columns=[1, 2, 3, 4],
        total_row_indexes=[len(summary_data) - 1],
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # 2. Detailed Breakdown by Term (for Fees) if available
    # Only if specific year is selected or generally useful
    if fees_qs.exists():
        elements.append(Paragraph("School Fees Breakdown by Term", h2_style))
        term_stats = []
        for term_code, term_name in SchoolFee.TERM_CHOICES:
            term_fees = fees_qs.filter(term=term_code)
            t_req = term_fees.aggregate(Sum('total_fees'))['total_fees__sum'] or 0
            t_paid = term_fees.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
            t_bal = term_fees.aggregate(Sum('balance'))['balance__sum'] or 0
            term_stats.append([term_name, t_req, t_paid, t_bal])
            
        term_data = [['Term', 'Required', 'Paid', 'Balance']]
        for t in term_stats:
            term_data.append([
                t[0],
                f"{t[1]:,.0f}",
                f"{t[2]:,.0f}",
                f"{t[3]:,.0f}"
            ])
            
        term_table = build_export_table(
            term_data,
            col_widths=[2 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch],
            body_font_size=8,
            right_aligned_columns=[1, 2, 3],
        )
        elements.append(term_table)
        elements.append(Spacer(1, 20))

    # 3. Insurance Status Breakdown
    if insurance_qs.exists():
        elements.append(Paragraph("Insurance Coverage Status", h2_style))
        covered = insurance_qs.filter(coverage_status='covered').count()
        partial = insurance_qs.filter(coverage_status='partially_covered').count()
        not_covered = insurance_qs.filter(coverage_status='not_covered').count()
        
        ins_data = [
            ['Status', 'Count', 'Percentage'],
            ['Covered', str(covered), f"{(covered/insurance_qs.count()*100):.1f}%"],
            ['Partially Covered', str(partial), f"{(partial/insurance_qs.count()*100):.1f}%"],
            ['Not Covered', str(not_covered), f"{(not_covered/insurance_qs.count()*100):.1f}%"],
        ]
        
        ins_table = build_export_table(
            ins_data,
            col_widths=[3 * inch, 2 * inch, 2 * inch],
            body_font_size=8,
            centered_columns=[1, 2],
        )
        elements.append(ins_table)

    # Build PDF
    doc.build(elements, canvasmaker=NumberedCanvas)
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"financial_report_{year_id if year_id else 'all_time'}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@permission_required(('finance.view_schoolfee', 'insurance.view_familyinsurance', 'students.view_student'), raise_exception=True)
def analysis_dashboard(request):
    """
    Dashboard view for statistical analysis and visualizations with filters.
    """
    # Get Filter Parameters
    year_id = request.GET.get('year')
    term_val = request.GET.get('term')
    district_id = request.GET.get('district')
    partner_id = request.GET.get('partner')
    level_val = request.GET.get('level')

    # Base QuerySets
    students_qs = Student.objects.all()
    marks_qs = StudentMark.objects.all()
    materials_qs = StudentMaterial.objects.all()
    fees_qs = SchoolFee.objects.all()

    # Apply Student Filters (Level, Partner, District)
    if level_val:
        students_qs = students_qs.filter(school_level=level_val)
    if partner_id:
        students_qs = students_qs.filter(partner_id=partner_id)
    if district_id:
        students_qs = students_qs.filter(family__district_id=district_id)

    # Filter related data by filtered students
    # We apply this even if no student filters are set, to ensure consistency 
    # (though if no student filters, students_qs is all, so it's a no-op efficiently handled by DB usually)
    if level_val or partner_id or district_id:
        marks_qs = marks_qs.filter(student__in=students_qs)
        materials_qs = materials_qs.filter(student__in=students_qs)
        fees_qs = fees_qs.filter(student__in=students_qs)

    # Apply Year Filter
    if year_id:
        marks_qs = marks_qs.filter(academic_year_id=year_id)
        materials_qs = materials_qs.filter(academic_year_id=year_id)
        fees_qs = fees_qs.filter(academic_year_id=year_id)

    # Apply Term Filter
    if term_val:
        # StudentMark uses 'Term X' format
        marks_qs = marks_qs.filter(term=f"Term {term_val}")
        # SchoolFee uses 'X' format
        fees_qs = fees_qs.filter(term=term_val)
        # StudentMaterial is annual, so term filter doesn't apply directly

    # Student Analysis
    total_students = students_qs.count()
    university_students = students_qs.filter(school_level='university').count()
    
    # Gender Distribution
    gender_map = dict(Student.GENDER_CHOICES)
    gender_data = students_qs.values('gender').annotate(count=Count('id'))
    gender_labels = [gender_map.get(item['gender'], item['gender']) for item in gender_data]
    gender_counts = [item['count'] for item in gender_data]
    
    # School Level Distribution
    level_map = dict(Student.SCHOOL_LEVEL_CHOICES)
    level_data = students_qs.values('school_level').annotate(count=Count('id'))
    level_labels = [level_map.get(item['school_level'], item['school_level']) for item in level_data]
    level_counts = [item['count'] for item in level_data]
    level_count_map = {item['school_level']: item['count'] for item in level_data}
    school_level_rows = [
        {
            'label': label,
            'count': level_count_map.get(value, 0),
            'percentage': round(
                (
                    level_count_map.get(value, 0)
                    / total_students * 100
                ) if total_students else 0,
                1,
            ),
            'is_university': value == 'university',
        }
        for value, label in Student.SCHOOL_LEVEL_CHOICES
    ]

    age_analysis = _build_age_band_analysis(students_qs)
    age_range_rows = age_analysis['rows']
    age_labels = age_analysis['labels']
    age_counts = age_analysis['counts']

    university_breakdown = {
        'male': students_qs.filter(school_level='university', gender='M').count(),
        'female': students_qs.filter(school_level='university', gender='F').count(),
        'active': students_qs.filter(school_level='university', sponsorship_status='active').count(),
        'pending': students_qs.filter(school_level='university', sponsorship_status='pending').count(),
        'graduated': students_qs.filter(school_level='university', sponsorship_status='graduated').count(),
    }
    university_chart_labels = ['Male', 'Female', 'Active', 'Pending', 'Graduated']
    university_chart_counts = [
        university_breakdown['male'],
        university_breakdown['female'],
        university_breakdown['active'],
        university_breakdown['pending'],
        university_breakdown['graduated'],
    ]

    # Finance Analysis
    total_fees_expected = fees_qs.aggregate(total=Sum('total_fees'))['total'] or 0
    total_fees_paid = fees_qs.aggregate(total=Sum('amount_paid'))['total'] or 0
    total_balance = fees_qs.aggregate(total=Sum('balance'))['total'] or 0
    
    # Payment Status Distribution
    status_map = dict(SchoolFee.PAYMENT_STATUS_CHOICES)
    payment_status_data = fees_qs.values('payment_status').annotate(count=Count('id'))
    status_labels = [status_map.get(item['payment_status'], item['payment_status']) for item in payment_status_data]
    status_counts = [item['count'] for item in payment_status_data]

    # Performance Analysis
    avg_marks = marks_qs.aggregate(avg=Avg('marks'))['avg'] or 0
    total_marks_count = marks_qs.count()
    passed_count = marks_qs.filter(marks__gte=50).count()
    pass_rate = (passed_count / total_marks_count * 100) if total_marks_count > 0 else 0

    # Best District Performance
    district_performance = marks_qs.values(
        'student__family__district__name'
    ).annotate(
        avg_marks=Avg('marks'),
        success_rate=Count('id', filter=Q(marks__gte=50)) * 100.0 / Count('id')
    ).order_by('-success_rate', '-avg_marks')
    
    best_district = district_performance.first() if district_performance else None

    # Student Categories & Levels
    # Categories: lower primary(p1,p2,p3), upper primary(p4,p5,p6), ordinary level(s1,s2,s3), advanced level(s4,s5,s6)
    lower_primary_levels = ['P1', 'P2', 'P3', 'p1', 'p2', 'p3']
    upper_primary_levels = ['P4', 'P5', 'P6', 'p4', 'p5', 'p6']
    ordinary_level_levels = ['S1', 'S2', 'S3', 's1', 's2', 's3']
    advanced_level_levels = ['S4', 'S5', 'S6', 's4', 's5', 's6']

    category_counts = {
        'lower_primary': students_qs.filter(class_level__in=lower_primary_levels).count(),
        'upper_primary': students_qs.filter(class_level__in=upper_primary_levels).count(),
        'ordinary_level': students_qs.filter(class_level__in=ordinary_level_levels).count(),
        'advanced_level': students_qs.filter(class_level__in=advanced_level_levels).count(),
    }

    # Graduates (S6 students)
    graduating_students = students_qs.filter(class_level='S6').select_related('family__district', 'school')
    graduates_count = graduating_students.count()

    # Materials Analysis
    # If term filter is active, materials might show annual data still, which is fine, 
    # or we could clear it. For now, we show annual data filtered by year/student.
    total_materials_records = materials_qs.count()
    
    materials_stats = {
        'books': materials_qs.filter(books_received=True).count(),
        'bags': materials_qs.filter(bag_received=True).count(),
        'shoes': materials_qs.filter(shoes_received=True).count(),
        'uniforms': materials_qs.filter(uniforms_received=True).count(),
    }
    materials_labels = ['Books', 'Bags', 'Shoes', 'Uniforms']
    materials_counts = [materials_stats['books'], materials_stats['bags'], materials_stats['shoes'], materials_stats['uniforms']]

    # Partner Analysis
    partner_data = students_qs.values('partner__name').annotate(count=Count('id')).order_by('-count')
    partner_labels = [item['partner__name'] or 'No Partner' for item in partner_data]
    partner_counts = [item['count'] for item in partner_data]

    # Context for Filters
    academic_years = AcademicYear.objects.all().order_by('-name')
    districts = District.objects.all().order_by('name')
    partners = Partner.objects.all().order_by('name')
    school_levels = Student.SCHOOL_LEVEL_CHOICES
    terms = [('1', 'Term 1'), ('2', 'Term 2'), ('3', 'Term 3')]

    context = {
        'page_title': 'Analysis Dashboard',
        'total_students': total_students,
        'university_students': university_students,
        'gender_labels': gender_labels,
        'gender_counts': gender_counts,
        'level_labels': level_labels,
        'level_counts': level_counts,
        'school_level_rows': school_level_rows,
        'age_range_rows': age_range_rows,
        'age_labels': age_labels,
        'age_counts': age_counts,
        'students_with_dob': age_analysis['students_with_dob'],
        'students_missing_dob': age_analysis['students_missing_dob'],
        'university_with_dob': age_analysis['university_with_dob'],
        'university_missing_dob': age_analysis['university_missing_dob'],
        'university_breakdown': university_breakdown,
        'university_chart_labels': university_chart_labels,
        'university_chart_counts': university_chart_counts,
        'total_fees_expected': float(total_fees_expected),
        'total_fees_paid': float(total_fees_paid),
        'total_balance': float(total_balance),
        'status_labels': status_labels,
        'status_counts': status_counts,
        # New Stats
        'avg_marks': round(float(avg_marks), 1),
        'pass_rate': round(float(pass_rate), 1),
        'materials_labels': materials_labels,
        'materials_counts': materials_counts,
        'partner_labels': partner_labels,
        'partner_counts': partner_counts,
        # Performance Updates
        'best_district': best_district,
        'category_counts': category_counts,
        'graduating_students': graduating_students,
        'graduates_count': graduates_count,
        # Filters
        'academic_years': academic_years,
        'districts': districts,
        'partners': partners,
        'school_levels': school_levels,
        'terms': terms,
        'selected_year': int(year_id) if year_id else None,
        'selected_term': term_val,
        'selected_district': int(district_id) if district_id else None,
        'selected_partner': int(partner_id) if partner_id else None,
        'selected_level': level_val,
    }
    return render(request, 'reports/analysis.html', context)

